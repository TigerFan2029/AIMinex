import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, scrolledtext, IntVar, Menu
from tkinter import *
from tkinter import filedialog as fd
from customtkinter import CTkImage

import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import MouseButton
from matplotlib.lines import Line2D

from sklearn.decomposition import PCA, KernelPCA
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler, FunctionTransformer
from sklearn.cluster import KMeans
from sklearn.pipeline import Pipeline, make_pipeline
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.cluster import AgglomerativeClustering, DBSCAN, SpectralClustering, MeanShift, AffinityPropagation, AgglomerativeClustering, Birch
from sklearn.mixture import GaussianMixture
from yellowbrick.cluster import KElbowVisualizer  

import re
from PIL import Image, ImageTk
from tktooltip import ToolTip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import webbrowser

from typing import Union, Callable

from terminal import TerminalApp
from pca import PCA_class



class main(ctk.CTk):   
    def __init__(self):
        super().__init__()
        self.title("MineralAI")
        ctk.set_appearance_mode("Light")
        style = ttk.Style()
        style.theme_use("default")
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        self.geometry(f"{screen_width}x{screen_height}")
        self.filtered_df = pd.DataFrame()
        self.create_widgets()

        # Get the system default color for selection background
        system_select_bg = self.option_get('selectBackground', '')

        # If the system default is empty, use a fallback color
        if not system_select_bg:
            system_select_bg = 'LightSkyBlue1'
        
        # Set the palette with the system default selection background color
        self.tk_setPalette(background='white', foreground='black', selectBackground=system_select_bg, activeForeground='black')

    
    def create_widgets(self):    
        # Create the main layout frames
        self.selection_frame = ctk.CTkFrame(self, height = 250)
        self.selection_frame.grid(row=0, column=0, rowspan=2, sticky="nsew", pady=(5,0), padx=5)
        
        self.legend_frame = ctk.CTkFrame(self, width=250)
        self.legend_frame.grid(row=0, column=3, sticky="ns", padx=5, pady=(5,0))
    
        self.box_frame = ctk.CTkFrame(self, width=150)
        self.box_frame.grid(row=2, column=0, rowspan=2, sticky="nsew", padx=5, pady=5)
    
        self.box_frame_sub = ctk.CTkFrame(self, width=200)
        self.box_frame_sub.pack_propagate(False)
        self.box_frame_sub.grid(row=0, column=1, rowspan=4, sticky="ns", pady=5)
    
        self.bargraph_frame = ctk.CTkFrame(self)
        self.bargraph_frame.grid(row=0, column=2, rowspan=4, sticky="nsew", pady=5, padx=(5,0))
        
        self.output_frame = ctk.CTkFrame(self, width=250)
        self.output_frame.pack_propagate(False)
        self.output_frame.grid(row=1, column=3, rowspan=3, sticky="nsew", padx=5, pady=5)
        
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(2, weight=1)
        
        self.output_text = scrolledtext.ScrolledText(
            self.output_frame, 
            wrap="word", 
            bd=0,  # Border thickness
            highlightthickness=0,  # Removes the focus border
            relief="flat"  # Makes the widget flat without any 3D effect
        )
        self.output_text.pack(fill="both", expand=True)

    

        menubar = Menu(self)
        self.filemenu = Menu(menubar, tearoff=0)

        # Create file menu options
        self.filemenu.add_command(label="Open File", command=self.load_data)
        self.filemenu.add_command(label="New Tab", command=lambda: self.shared_container.create_tab("New Tab"))
        self.filemenu.entryconfig("New Tab", state=tk.DISABLED)
        
        self.save_pc_menu = tk.Menu(self.filemenu, tearoff=0)
        
        # Create save PC options
        self.save_pc_menu.add_command(label="Save PC by sample excel", command=self.export_pc_by_sample)
        self.save_pc_menu.add_command(label="Save PC by element excel", command=self.export_pc_by_element)
        
        self.save_pc_menu.entryconfig("Save PC by sample excel", state=tk.DISABLED)
        self.save_pc_menu.entryconfig("Save PC by element excel", state=tk.DISABLED)
        
        self.filemenu.add_cascade(label="Save PC", menu=self.save_pc_menu)
        
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Exit", command=self.destroy)
        menubar.add_cascade(label="File", menu=self.filemenu)
        
        self.editmenu = Menu(menubar, tearoff=0)
        self.editmenu.add_command(label="Edit Data Point Color/Shape", command=self.open_color_window, state="disabled")

        menubar.add_cascade(label="Edit", menu=self.editmenu)
        
        helpmenu = Menu(menubar, tearoff=0)
        # Create help menu options
        helpmenu.add_command(label="Help Index", command=self.open_help_html)
        helpmenu.add_separator()
        helpmenu.add_command(label="Short-Cuts:", command=None)
        helpmenu.add_command(label="Deselect_all", command=None, accelerator="Cmd+d")
        helpmenu.add_separator()
        helpmenu.add_command(label="About...", command=None)
        menubar.add_cascade(label="Help", menu=helpmenu)

        self.config(menu=menubar)
        
        self.terminal = TerminalApp(self.output_text, self)  # Pass self to TerminalApp

        style = ttk.Style()
        style.map('TCombobox', 
        fieldbackground=[('readonly', 'white')],
        background=[('readonly', 'lightgrey')])

    def open_color_window(self):
        global dc, ds, color_map, color_map1
    
        # Open a new window for manual color editing
        color_window = tk.Toplevel(self)
        color_window.title("Edit Colors")
    
        # Lithologies Section
        lithology_label = ctk.CTkLabel(color_window, text="Lithology")
        lithology_label.grid(row=0, column=0, padx=10, pady=10)
    
        color_label = ctk.CTkLabel(color_window, text="Color")
        color_label.grid(row=0, column=1, padx=10, pady=10)
    
        visual_label = ctk.CTkLabel(color_window, text="Selected Color")
        visual_label.grid(row=0, column=2, padx=10, pady=10)
    
        # Generate the list of comboboxes for lithology color selection
        for idx, lithology in enumerate(self.lithologies):
            # Label for lithology
            litho_label = ctk.CTkLabel(color_window, text=lithology)
            litho_label.grid(row=idx + 1, column=0, padx=10, pady=5)

            # Small rectangle to visualize the selected color
            color_visual = ttk.Label(color_window, text="   ", background=color_map.get(lithology, "black"))
            color_visual.grid(row=idx + 1, column=2, padx=10, pady=5)
    
            # Combobox for selecting color
            color_combo = ctk.CTkComboBox(
                color_window, 
                values=["lime", "blue", "cyan", "deeppink", "pink", "black", "grey", "sienna", "dimgrey", "lightgrey"], 
                state="readonly", 
                command=lambda selected_value, lithology=lithology, color_visual=color_visual: self.update_color_visual(selected_value, lithology, color_visual)
            )
            color_combo.set(color_map.get(lithology, "black"))
            color_combo.grid(row=idx + 1, column=1, padx=10, pady=5)

                
        # Rock Units Section
        rock_label = ctk.CTkLabel(color_window, text="Rock Unit")
        rock_label.grid(row=0, column=4, padx=10, pady=10)
        
        shape_label = ctk.CTkLabel(color_window, text="Shape")
        shape_label.grid(row=0, column=5, padx=10, pady=10)
        
        shape_visual_label = ctk.CTkLabel(color_window, text="Selected Shape")
        shape_visual_label.grid(row=0, column=6, padx=10, pady=10)
        
        # Define a mapping between user-friendly names and symbols
        shape_name_to_symbol = {
            "triangle": "^",
            "star": "*",
            "circle": "o",
            "square": "s",
            "diamond": "D"
        }
        
        # Define the reverse mapping (symbols to names) for setting the default value
        symbol_to_shape_name = {v: k for k, v in shape_name_to_symbol.items()}
        
        # Rock Units Section
        for idx, rock_unit in enumerate(ds['Shape'].unique()):
            # Label for rock unit
            rock_label = ctk.CTkLabel(color_window, text=rock_unit)
            rock_label.grid(row=idx + 1, column=4, padx=10, pady=5)
        
            # Get the default shape symbol
            default_shape_symbol = color_map1.get(rock_unit, ds.loc[ds['Shape'] == rock_unit, 'Shapes'].values[0])
            
            # Convert the symbol to a user-friendly name for display in the combobox
            default_shape_name = symbol_to_shape_name.get(default_shape_symbol)
                    
            # Text to visualize the selected shape symbol
            rock_shape_visual = ttk.Label(color_window, text=default_shape_symbol)
            rock_shape_visual.grid(row=idx + 1, column=6, padx=10, pady=5)
        
            # Combobox for selecting shape (display names instead of symbols)
            rock_shape_combo = ctk.CTkComboBox(
                color_window, 
                values=list(shape_name_to_symbol.keys()),  # Display the names in the combobox
                state="readonly", 
                command=lambda selected_value, rock_unit=rock_unit, rock_shape_visual=rock_shape_visual: self.update_rock_shape_visual(selected_value, rock_unit, rock_shape_visual, shape_name_to_symbol)
            )
            rock_shape_combo.set(default_shape_name)  # Set the default name for the combobox
            rock_shape_combo.grid(row=idx + 1, column=5, padx=10, pady=5)


        # Apply Button to save the changes
        apply_button = ctk.CTkButton(color_window, text="Apply Colors", command=self.apply_color_change)
        apply_button.grid(row=max(len(self.lithologies), len(ds['Shape'].unique())) + 2, column=0, columnspan=7, stick = "ew", padx=10, pady=(10,0))

        label = ctk.CTkLabel(color_window, text="Redraw graphs to apply change")
        label.grid(row=max(len(self.lithologies), len(ds['Shape'].unique())) + 3, column=0, columnspan=7, stick = "ew", padx=10, pady=(0,10))
        
    def update_color_visual(self, selected_value, lithology, color_visual):
        # Declare dc as global to access the global dataframe
        global dc, color_map
        
        # Update the rectangle color preview with the selected color
        color_visual.configure(background=selected_value)
        
        # Update the color map and global dc dataframe immediately
        color_map[lithology] = selected_value
        dc.loc[dc['Lithology'] == lithology, 'Color'] = selected_value

    def update_rock_shape_visual(self, selected_shape_name, rock_unit, rock_shape_visual, shape_name_to_symbol):
        global ds, color_map1
        # Convert the selected name back to the corresponding symbol
        selected_symbol = shape_name_to_symbol.get(selected_shape_name)
        
        # Update the shape visualization (display the selected symbol)
        rock_shape_visual.configure(text=selected_symbol)
        
        # Update the shape map and global ds dataframe with the symbol, not the name
        color_map1[rock_unit] = selected_symbol
        ds.loc[ds['Shape'] == rock_unit, 'Shapes'] = selected_symbol


    def apply_color_change(self):
        global dc, ds, color_map, color_map1
        
        # Get selected lithology or rock unit from the listbox and the chosen color
        selected_lithology = self.lithology_listbox_color.get(tk.ACTIVE)
        selected_color = self.color_combo.get()
        
        if selected_lithology and selected_color:
            # Update the color map for lithologies (color_map) and global `dc`
            if selected_lithology in self.lithologies:
                dc.loc[dc['Lithology'] == selected_lithology, 'Color'] = selected_color
                color_map[selected_lithology] = selected_color
            
            # Update the color map for rock units (color_map1) and global `ds`
            if selected_lithology in ds['Shape'].values:
                ds.loc[ds['Shape'] == selected_lithology, 'Shapes'] = selected_color
                color_map1[selected_lithology] = selected_color
    
            # Update visuals globally to reflect the changes
            self.update_visuals_globally()

        
    def open_help_html(self):
        import urllib.parse
        
        # Open HTML help file
        html_path = '_build/html/index.html'
        absolute_path = os.path.abspath(html_path)
        file_url = urllib.parse.urljoin('file:', urllib.request.pathname2url(absolute_path))
        print(f"Opening HTML file at: {file_url}")
        
        try:
            if os.path.exists(absolute_path):
                webbrowser.open_new_tab(file_url)
                print("HTML file opened successfully.")
            else:
                print(f"File does not exist: {absolute_path}")
        except Exception as e:
            print(f"Error opening HTML file: {e}")

    def button_0(self):
        # Create sheet combobox and load button
        self.sheet_combobox = ctk.CTkComboBox(self.selection_frame, values=[], state="readonly")
        self.sheet_combobox.pack(padx=5, pady=5)
        self.sheet_combobox.set("Select a sheet")
        self.load_sheet_button = ctk.CTkButton(self.selection_frame, text="Load Sheet", command=self.load_selected_sheet)
        self.load_sheet_button.pack(pady=5)

    def create_lithobuttons(self):
        # Create lithology selection buttons
        ctk.CTkLabel(self.selection_frame, text="Select Lithology:").grid(row=0, column=0, columnspan=4, sticky="w", padx=5, pady=(5,0))
        self.lithology_listbox.grid(row=1, column=0, columnspan=4, sticky="we", padx=5, pady=(5,0))
        
        def deselect_all():
            self.lithology_listbox.selection_clear(0, tk.END)
            
        self.lithology_listbox.bind('<Command-d>', lambda event: deselect_all())
        self.create_buttons()

    def create_buttons(self):
        #initialize tab function
        self.filemenu.entryconfig("New Tab", state=tk.NORMAL)
        self.shared_container = SharedContainer(self.bargraph_frame)
        
        # Create PCA and scaling options
        self.current_button = None
        
        self.scaler_combo = ctk.CTkComboBox(self.selection_frame, values=["Standard Scaler", "Logarithmic Scaler"], state="readonly")
        self.scaler_combo.grid(row=2, column=0, columnspan=4, sticky="we", padx=5, pady=(5,0))
        self.scaler_combo.set("Select PCA Scaler:")

        self.pca_type_combo = ctk.CTkComboBox(self.selection_frame, values=["PCA", "Kernel PCA"], command=self.kernelstat, state="readonly")
        self.pca_type_combo.grid(row=3, column=0, columnspan=4, sticky="we", padx=5, pady=(5,0))
        self.pca_type_combo.set("PCA")
        

        ctk.CTkLabel(self.selection_frame, text="Number of PCA Components:").grid(row=8, column=0, columnspan=4, sticky="we", padx=5, pady=(5,0))
        self.slider = ctk.CTkSlider(self.selection_frame, from_=1, to=8, number_of_steps=7, command=self.on_slider_change)
        self.slider.set(6)
        self.slider.grid(row=9, column=0, columnspan=4, padx=10, pady=0)
        self.label = ctk.CTkLabel(self.selection_frame, text=f"Current value: {int(self.slider.get())}")
        self.label.grid(row=10, column=0, columnspan=4, padx=5, pady=0)

        # initiate widgets for kernal parameters
        self.gamma_slider = ctk.CTkSlider(self.selection_frame, from_=0, to=1, width=120, command=self.gamma_change)
        self.degree_slider = ctk.CTkSlider(self.selection_frame, from_=0, to=10, width=120, command=self.degree_change)
        self.coef_slider = ctk.CTkSlider(self.selection_frame, from_=0, to=10, width=120, command=self.coef_change)
        
        self.label1 = ctk.CTkLabel(self.selection_frame, text=f"Gamma: {self.gamma_slider.get():.2f}")
        self.label2 = ctk.CTkLabel(self.selection_frame, text=f"Degree: {self.degree_slider.get():.2f}")
        self.label3 = ctk.CTkLabel(self.selection_frame, text=f"Coef: {self.coef_slider.get():.2f}")

        # set defaults
        filtered_columns = [col for col in self.df_0.columns if '_ppm' in col or '_pct' in col]
        num_filtered_features = len(filtered_columns)
        
        self.gamma_slider.set(1/num_filtered_features)
        self.degree_slider.set(3)
        self.coef_slider.set(1)

        self.gamma = 1/num_filtered_features
        self.degree = 3
        self.coef = 1
        
        self.label1.configure(text=f"Gamma: {self.gamma:.2f}")
        self.label2.configure(text=f"Degree: {self.degree:.2f}")
        self.label3.configure(text=f"Coef: {self.coef:.2f}")
        
        self.kernel_combo = ctk.CTkComboBox(self.selection_frame, values=["linear", "poly", "rbf", "sigmoid", "cosine", "precomputed"], command=self.kernel_param, state="readonly")
        self.kernel_text = ctk.CTkLabel(self.selection_frame, text="Kernel:")
        self.kernel_combo.set("linear")

        self.apply_button = ctk.CTkButton(self.selection_frame, text="Apply", command= self.filter_dataframe)
        self.apply_button.grid(row=11, column=0, columnspan=4, padx=5, pady=(0,5))

        self.kernel_param()

    def kernel_param(self, *arg):
        self.kernel = self.kernel_combo.get()
         
        # clear all previous sliders
        try:
            self.gamma_slider.grid_forget()
            self.degree_slider.grid_forget()
            self.coef_slider.grid_forget()
            self.label1.grid_forget()
            self.label2.grid_forget()
            self.label3.grid_forget()
        except:
            pass

        if self.kernel == "rbf":
            self.gamma_slider.grid(row=5, column=1, columnspan=3, sticky="e", padx=(10,0), pady=0)
            self.label1.grid(row=5, column=0, sticky="w", columnspan=2, padx=(10,0), pady=0)

        elif self.kernel == "poly":
            self.gamma_slider.grid(row=5, column=1, columnspan=3, sticky="e", padx=(10,0), pady=0)
            self.label1.grid(row=5, column=0, sticky="w", columnspan=2, padx=(10,0), pady=0)
            
            self.degree_slider.grid(row=6, column=1, columnspan=3, sticky="e", padx=(10,0), pady=0)
            self.label2.grid(row=6, column=0, sticky="w", columnspan=2, padx=(10,0), pady=0)

            self.coef_slider.grid(row=7, column=1, columnspan=3, sticky="e", padx=(10,0), pady=0)
            self.label3.grid(row=7, column=0, sticky="w", columnspan=2, padx=(10,0), pady=0)

        elif self.kernel == "sigmoid":
            self.gamma_slider.grid(row=5, column=1, columnspan=3, sticky="e", padx=(10,0), pady=0)
            self.label1.grid(row=5, column=0, sticky="w", columnspan=2, padx=(10,0), pady=0)
            
            self.coef_slider.grid(row=5, column=1, columnspan=3, sticky="e", padx=(10,0), pady=0)
            self.label3.grid(row=5, column=0, sticky="w", columnspan=2, padx=(10,0), pady=0)

        else:
            pass
                
    def kernelstat(self, *args):
        # Show kernel options if Kernel PCA is selected
        pca_type = self.pca_type_combo.get()
        if pca_type == "Kernel PCA":
            self.kernel_combo.grid(row=4, column=1, columnspan=3, sticky="we", padx=5, pady=(5,0))
            self.kernel_text.grid(row=4, column=0, sticky="w", padx=(10,0), pady=(5,0))
            self.kernel_param()
        else:
            self.kernel_combo.grid_forget()
            self.kernel_text.grid_forget()
            # clear all previous sliders
            try:
                self.gamma_slider.grid_forget()
                self.degree_slider.grid_forget()
                self.coef_slider.grid_forget()
                self.label1.grid_forget()
                self.label2.grid_forget()
                self.label3.grid_forget()
            except:
                pass
            
    def on_slider_change(self, value):
        # Update label when slider changes
        self.label.configure(text=f"Current value: {int(value)}")

    def gamma_change(self, value):
        self.label1.configure(text=f"Gamma: {value:.2f}")
        self.gamma = value

    def degree_change(self, value):
        self.label2.configure(text=f"Degree: {value:.2f}")
        self.degree = value

    def coef_change(self, value):
        self.label3.configure(text=f"Coef: {value:.2f}")
        self.coef = value

    def on_button_click(self, button, command):
        # Handle button click and change appearance
        if self.current_button is not None:
            self.current_button.configure(fg_color='transparent') #'#3B8ED0'
        else:
            pass
            
        button.configure(fg_color="grey")
        self.current_button = button
        command()

    def load_data(self):
        # Load data from selected file
        try:
            #self.file_path = 'Geochemistry Results-AGG reduced variables.xlsx'
            self.file_path = fd.askopenfilename()
            if not self.file_path:
                self.output_text.insert("end", "No file selected\n")
                return

            self.clear_all()
            
            self.button_0()
            self.sheet_names = pd.ExcelFile(self.file_path).sheet_names
            self.sheet_combobox.configure(values=self.sheet_names)
            self.output_text.insert("end", "File loaded successfully. Please select a sheet.\n")

        except:
            self.output_text.insert("end", f"Error loading file: \n")

    def load_selected_sheet(self):
        # Load selected sheet from the file
        try:
            selected_sheet = self.sheet_combobox.get()
            if selected_sheet not in self.sheet_names:
                self.output_text.insert("end", "Invalid sheet name\n")
                return
            self.sheet_name=selected_sheet
            self.df_0 = pd.read_excel(self.file_path, self.sheet_name)
            
            for widget in self.selection_frame.winfo_children():
                widget.destroy()
            self.process_sheet()
            
        except Exception as e:
            self.output_text.insert("end", f"Error loading sheet: {str(e)}\n")

    def process_sheet(self):
        # Process the sheet and populate lithology options
        if "Lithology" in self.df_0.columns:
            self.df_0['Lithology'] = self.df_0['Lithology'].apply(lambda x: x.strip().lower().title() if isinstance(x, str) and pd.notnull(x) else x)
            self.lithology_listbox = tk.Listbox(self.selection_frame, selectmode=tk.MULTIPLE)
            self.lithologies = self.df_0['Lithology'].unique()
            for lithology in self.lithologies:
                self.lithology_listbox.insert(tk.END, lithology)
            self.create_lithobuttons()

        else:
            self.output_text.insert("end", "Lithology column not found in dataframe\n")
            self.create_buttons()

    def clear_all(self):
        # Clear all widgets in frames
        for widget in self.bargraph_frame.winfo_children():
            widget.destroy()
        for widget in self.box_frame.winfo_children():
            widget.destroy()
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()
        # for widget in self.output_frame.winfo_children():
        #     widget.destroy()
        for widget in self.legend_frame.winfo_children():
            widget.destroy()
        for widget in self.selection_frame.winfo_children():
            widget.destroy()

    def clear(self):
        # Clear specific widgets in frames
        # for widget in self.bargraph_frame.winfo_children():
        #     widget.destroy()
        for widget in self.box_frame.winfo_children():
            widget.destroy()
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()
        self.output_text.delete("1.0", tk.END)
        for widget in self.legend_frame.winfo_children():
            widget.destroy()

    def filter_dataframe(self):   
        #Filter dataframe based on selected lithology and update UI
        if hasattr(self, 'scaler_label') and self.scaler_label is not None:
            self.scaler_label.destroy()
            del self.scaler_label
            
        if hasattr(self, 'pca_label') and self.pca_label is not None:
            self.pca_label.destroy()
            del self.pca_label

        if self.pca_type_combo.get() == "PCA":
            self.pca_label = ctk.CTkLabel(self.selection_frame, text="PCA", font=("Arial", 12))
            
        if self.pca_type_combo.get() == "Kernel PCA":
            if self.kernel == "rbf":
                self.pca_label = ctk.CTkLabel(self.selection_frame, text= f"Kernel PCA-{self.kernel} ({self.gamma:.2f})", font=("Arial", 12))  
            elif self.kernel == "poly":
                self.pca_label = ctk.CTkLabel(self.selection_frame, text= f"Kernel PCA-{self.kernel} ({self.gamma:.2f}, {self.degree:.1f}, {self.coef:.1f})", font=("Arial", 12))  
            elif self.kernel == "rbf":
                self.pca_label = ctk.CTkLabel(self.selection_frame, text= f"Kernel PCA-{self.kernel} ({self.gamma:.2f}, {self.coef:.1f})", font=("Arial", 12))  
            else:
                self.pca_label = ctk.CTkLabel(self.selection_frame, text= f"Kernel PCA-{self.kernel}", font=("Arial", 12)) 
                
        self.pca_label.grid(row=13, column=0, columnspan=4, pady=(0,5))
            
        if self.scaler_combo.get() == "Select PCA Scaler:":
            self.scaler_label = ctk.CTkLabel(self.selection_frame, text="Select a Scaler", font=("Arial", 12))
            self.scaler_label.grid(row=12, column=0, columnspan=4, padx=5, pady=0)
            return
        else:
            # self.scaler_label.grid_forget()
            self.scaler_label = ctk.CTkLabel(self.selection_frame, text= f"Selected scaler: {self.scaler_combo.get()}", font=("Arial", 12))
            self.scaler_label.grid(row=12, column=0, columnspan=4, padx=5, pady=0)
   
        self.clear()
        self.current_button = None
        if hasattr(self, 'lithologies_label') and self.lithologies_label is not None:
            self.lithologies_label.destroy()
            del self.lithologies_label

        # self.output_text = scrolledtext.ScrolledText(self.output_frame, wrap="word", bd=0)
        # self.output_text.pack(fill="both", expand=True)
        # TerminalApp(self.output_text)

        try:
            # Check selected lithologies and add check mark to them
            selected_indices = self.lithology_listbox.curselection()
            print (selected_indices)
            select_lithologies = [self.lithology_listbox.get(i) for i in selected_indices]
            print (select_lithologies)
            selected_lithologies = [item.replace('✔️ ', '') for item in select_lithologies]
            print (selected_lithologies)
            
            self.filtered_df = self.df_0[self.df_0['Lithology'].isin(selected_lithologies)]
            print(self.filtered_df)
            
            item = None
            self.lithology_listbox.delete(0, tk.END)
            self.selected_items = {lithology: (lithology in selected_lithologies) for lithology in self.lithologies}
        
            for lithology, selected in self.selected_items.items():
                if selected:
                    item = "✔️ " + lithology
                else:
                    item = lithology
                self.lithology_listbox.insert(tk.END, item)
    
            def deselect_all():
                self.lithology_listbox.selection_clear(0, tk.END)
    
            self.lithology_listbox.bind('<Command-d>', lambda event: deselect_all())
            
            if not selected_indices:
                self.lithologies_label = ctk.CTkLabel(self.selection_frame, text="No lithology selected:(", font=("Arial", 12))
                self.lithologies_label.grid(row=14, column=0, columnspan=4, padx=5, pady=(0,5))
                                 
        except:
            self.filtered_df = self.df_0

        # Cleaning Data Frame to contain only the datas/elements
        filtered_columns = [col for col in self.filtered_df.columns if '_ppm' in col or '_pct' in col]

        df_filtered = self.filtered_df.replace('<', '', regex=True)
        df_filtered[filtered_columns] = df_filtered[filtered_columns].apply(pd.to_numeric, errors='coerce')
        self.cleaned_df = df_filtered.dropna(subset=filtered_columns)
        
        df_filtered = self.cleaned_df[filtered_columns]
        self.cleaned_df.columns = self.cleaned_df.columns.str.lower()

        self.df = df_filtered

        # Creating data frame with Rare Earth Elements
        self.df_c = self.df.copy()
        self.df_c.columns = self.df_c.columns.str.replace('_ppm', '').str.replace('_pct', '').str.replace('_Howell', '')
        
        LREE_elements = ['La', 'Ce', 'Pr', 'Nd', 'Pm', 'Sm', 'Eu', 'Gd']
        HREE_elements = ['Tb', 'Dy', 'Ho', 'Er', 'Tm', 'Yb', 'Lu']
        
        LREE_elements_in_df = [element for element in LREE_elements if element in self.df_c.columns]
        HREE_elements_in_df = [element for element in HREE_elements if element in self.df_c.columns]
        
        if LREE_elements_in_df:
            self.df_c['LREE'] = self.df_c[LREE_elements_in_df].sum(axis=1)
        
        if HREE_elements_in_df:
            self.df_c['HREE'] = self.df_c[HREE_elements_in_df].sum(axis=1)
        
        if LREE_elements_in_df or HREE_elements_in_df:
            self.df_c['REE'] = self.df_c[LREE_elements_in_df + HREE_elements_in_df].sum(axis=1)

        # self.df_c['LREE'] = self.df_c[lree_elements].sum(axis=1)
        # self.df_c['HREE'] = self.df_c[hree_elements].sum(axis=1)
        
        # ree_elements = lree_elements + hree_elements
        # self.df_c['REE'] = self.df_c[ree_elements].sum(axis=1)

        # print(self.df_c)
        
        self.update_output()
        self.print_pca()

    def update_output(self):
        # Update output text with filtered and cleaned dataframe info
        # self.output_text.delete(1.0, "end")
        # self.output_text.insert("end", f"Filtered & Cleaned df:\n{self.df.head()}\n")
        
        self.output_text.insert("end", f"Filtered & Cleaned df:\n{self.df.shape}\n")
        
    def print_pca(self):
        # Perform PCA and update output
        try:
            self.pca_instance = PCA_class(self.df, self.scaler_combo, self.pca_type_combo, self.output_text, self.slider, self.kernel_combo, self.gamma, self.degree, self.coef)
            self.pca_instance.get_variance_ratio()
            self.output_text.insert("end", f"PCA performed successfully. Shape of transformed data: {self.pca_instance.x.shape}\n")
            #self.output_text.insert("end", f"PCA Components:\n{self.pca_instance.pca_df.head()}\n")
            self.color_function()
            self.shape_map()
            self.editmenu.entryconfig("Edit Data Point Color/Shape", state="normal")
            self.expand_buttons()
            self.loading_graph()   
            
        except ValueError as e:
            self.output_text.insert("end", f"{e}\n")

    def expand_buttons(self):
        # Create labels for PCA and Cluster Analysis sections
        self.pca_list = ctk.CTkLabel(self.box_frame, text="PCA Analysis Graphs")
        self.pca_list.grid(row=0, column=0, columnspan=6, sticky="w", pady=(10,0), padx=5)
        
        self.cluster_list = ctk.CTkLabel(self.box_frame, text="Cluster Analysis")
        self.cluster_list.grid(row=2, column=2, columnspan=3, sticky="w", pady=(10,0), padx=(25,0))

        self.supervised_list = ctk.CTkLabel(self.box_frame, text="Supervised Learning")
        self.supervised_list.grid(row=4, column=0, columnspan=6, sticky="w", pady=(5,0), padx=5)

    def color_function(self):
        global dc, color_map
        # Generate color mapping for lithologies
        try:
            dc = self.cleaned_df['lithology']
            dc = dc.to_frame()
            dc = dc.rename(columns={dc.columns[0]: 'Lithology'})
            
            # color_list = dc['Lithology'].unique()
            # default_colors = ["lime", "blue", "cyan", "deeppink", "pink", "black", "black", "black", "grey", "lightgrey"]
            # color_map = {}
            # # Populate the dictionary with symbols for each shape
            # for i, color in enumerate(color_list):
            #     if i < len(default_colors):
            #         color_map[color] = default_colors[i]
            color_map = {
                'Calcitic Marble': 'lime',
                'Dolomitic Marble': 'blue',
                'Di-Tr Dolomitic Marble': 'cyan',
                'Apatite Marble': 'deeppink',
                'Siliceous Calcitic Marble': 'steelblue',
                #'Carbonatite': 'deeppink',
                #'Carbonatite-Like': 'pink',
            
                'Syenite': 'black',
                'Altered Syenite': 'black',
                'Intrusion': 'black',
                'Skarn': 'sienna',
                'Syenite-Like': 'black',
                'Impure Siliciclastic': 'dimgrey',
                'Pure Siliciclastic': 'lightgrey',
                'Siliciclastic': 'grey'
                
            }
            dc['Color'] = dc['Lithology'].map(color_map)
            #self.output_text.insert("end", dc)
        except Exception as e:
            color_map = {}
            dc = pd.DataFrame()
            self.output_text.insert("end", f"Error generating lithology-color map: {e}\n")     
        
    def shape_map(self):
        global ds, color_map1
        # Generate shape mapping for rock units
        try:
            self.cleaned_df['rock unit'] = self.cleaned_df['rock unit'].apply(lambda x: x.strip().lower().title() if isinstance(x, str) and pd.notnull(x) else x)
            shape_list = self.cleaned_df['rock unit'].unique()
            # print(shape_list)
            
            ds = self.cleaned_df['rock unit']
            ds = ds.str.strip().to_frame()
            ds = ds.rename(columns={ds.columns[0]: 'Shape'})
            
            #shape_list = ds['Shape'].unique()
            # default_symbols = ["^", "*", "o", "s", "D"]
            # color_map1 = {}
            # # Populate the dictionary with symbols for each shape
            # for i, shape in enumerate(shape_list):
            #     if i < len(default_symbols):
            #         color_map1[shape] = default_symbols[i]
            
            color_map1 = {
                'Marble Units': "^",
                'Altered Intrusion': "*",
                'Siliciclastic': "o",
                'Intrusion': "s",
                'Anomalous Rock': "D",
            }
    
            ds["Shapes"] = ds["Shape"].map(color_map1)
            #self.output_text.insert("end", ds)Shape

        except Exception as e:
            color_map1 = {}
            ds = pd.DataFrame()
            self.output_text.insert("end", f"Error generating rock unit-shape map: {e}\n") 
            self.output_text.insert("end", f"Column Rock Unit does not exist, or check error in name?\n") 

    def loading(self):
        # Create loadings
        self.pca = self.pca_instance.pca

        if isinstance(self.pca, PCA):
            # print(f"pca: {self.pca}")
            components = self.pca.components_
            print(components.shape)
            # print(f"components: {components}")
            # print(f"componentsT: {components.T}") 
            # print(f"pca_df: {self.pca_instance.pca_df_scaled}")
            self.loadings = pd.DataFrame(
                components.T,
                columns=['PC' + str(i) for i in range(1, components.shape[0] + 1)],
                index=self.df.columns
            )

            print(self.loadings)
            # self.loadings.index = self.loadings.index.str.replace('_ppm', '').str.replace('_pct', '').str.replace('(int)', '').str.replace('_Howell', '')
        elif isinstance(self.pca, KernelPCA):
            # KernelPCA does not provide direct loadings like PCA
            Scaled_data = self.pca_instance.Scaled_data  # Access the scaled data
            K = self.pca._get_kernel(Scaled_data, self.pca.X_fit_)  # Compute the kernel matrix
            
            # Use the eigenvectors to approximate loadings
            eigenvectors = self.pca.eigenvectors_  # Shape (n_samples, n_components)
            loadings_approx = np.dot(K.T, eigenvectors)  # Shape (n_samples, n_components)

            # Reshape to ensure the shape matches with the original features
            loadings_approx = loadings_approx.T  # Shape (n_components, n_samples)
            
            # Compute the final loadings by taking the dot product with the original data
            loadings_final = np.dot(Scaled_data.T, loadings_approx.T)  # Shape (n_features, n_components)

            self.loadings = pd.DataFrame(
                loadings_final,
                columns=['PC' + str(i) for i in range(1, eigenvectors.shape[1] + 1)],
                index=self.df.columns
            )
            
            print(self.loadings)
            # components = self.pca_instance.kx
            
            # components_df = pd.DataFrame(components)
            # components_df.reset_index(drop=True, inplace=True)
            # components = components_df.to_numpy()

            # self.loadings = pd.DataFrame(
            #     components,
            #     columns=['PC' + str(i) for i in range(1, components.shape[1] + 1)],
            #     index=self.df.columns
            # # self.loadings.index = self.loadings.index.str.replace('_ppm', '').str.replace('_pct', '').str.replace('(int)', '').str.replace('_Howell', '')
            # )
        else:
            raise ValueError("self.pca must be an instance of PCA or KernelPCA")
    
    def selection(self):
        # Create a combo box for selecting clustering methods
        self.cluster_combo = ctk.CTkComboBox(
            self.box_frame, values=["K-mean", "Hierarchical", "DBSCAN", "Mean Shift", "Spectral", "GMM", "Affinity Propagation", "BIRCH"], width=100, height=20, command=self.update_cluster, state="readonly")
        # self.cluster_combo.bind("<<ComboboxSelected>>", self.on_combobox_selected)
        self.cluster_combo.grid(row=2, column=0, columnspan=3, sticky="w", pady=(10, 0), padx=5)
        self.cluster_combo.set("K-mean")
        self.cluster = self.cluster_combo.get()
        self.twod_instance = Cluster2DPlotClass(self.shared_container, self.cluster, self.df_c, self.cleaned_df, self.box_frame, self.box_frame_sub, self.on_button_click, self.legend_frame)
        self.threed_instance = Cluster3DPlotClass(self.shared_container, self.cluster, self.df_c, self.cleaned_df, self.box_frame, self.box_frame_sub, self.on_button_click, self.legend_frame)
        self.loading_cluster_instance = loading_cluster(self.shared_container, self.cluster, self.loadings, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button, self.legend_frame)
        if "sample id" in self.cleaned_df.columns:
            self.sample_cluster_instance = sample_cluster(self.shared_container, self.cluster, self.pca_df_scaled, self.df, self.cleaned_df, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button, self.legend_frame)
        else:
            print("Column Sample ID not in data frame, sample PC cluster plot plot not avaliable")


    def update_cluster(self, *arg):
        # update cluster
        self.cluster = self.cluster_combo.get()
        self.twod_instance.cluster_result = self.cluster
        self.threed_instance.cluster_result = self.cluster
        self.loading_cluster_instance.cluster_result = self.cluster
        self.sample_cluster_instance.cluster_result = self.cluster

            
        # rerun plot_2d_cluster_sub to refresh options in box_frame_sub
        print("clusterchanged1")
        for widget in self.box_frame_sub.winfo_children():
            # print(widget.winfo_name())
            if widget.winfo_name() == "size_combo":
                self.threed_instance.plot_3d_cluster_sub(self.cluster)  
                print("3d")
            elif widget.winfo_name() == "size_combo1":
                self.twod_instance.plot_2d_cluster_sub(self.cluster)
                print("2d")
            elif getattr(widget, "custom_name", "") == "loading_check":
                self.loading_cluster_instance.plot_cluster_sub(self.cluster)  
                print("LC")
            elif getattr(widget, "custom_name", "") == "sample_check":
                self.sample_cluster_instance.plot_cluster_sub(self.cluster)  
                print("SC")
            else:
                pass

    def loading_graph(self):
        # preparing data
        self.loading()
        self.pca_df_scaled = self.pca_instance.pca_df_scaled
        self.pca_df_scaled.reset_index(drop=True, inplace=True)
        #self.loadings.reset_index(drop=True, inplace=True)
        self.df.reset_index(drop=True, inplace=True)
        dc.reset_index(drop=True, inplace=True)
        ds.reset_index(drop=True, inplace=True)
        self.cleaned_df.reset_index(drop=True, inplace=True)
        self.df_c.reset_index(drop=True, inplace=True)

        # Load graph with PCA and clustering
        self.selection()
        loading_class(self.loadings, self.shared_container, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button)
        threed_class(self.shared_container, self.pca_df_scaled, self.df, self.cleaned_df, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button, self.legend_frame, self.loadings)
        twod_class(self.shared_container, self.pca_df_scaled, self.df, self.cleaned_df, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button, self.legend_frame, self.loadings)
        supervised_learning(self.df, self.cleaned_df, self.on_button_click, self.apply_button, self.box_frame)
        
        if "sample id" in self.cleaned_df.columns:
            sample_class(self.shared_container, self.pca_df_scaled, self.df, self.cleaned_df, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button)
            column_name = self.cleaned_df.filter(like='depth').columns[0]

            if "drillhole" in self.cleaned_df.columns and column_name in self.cleaned_df.columns:
                drill_class(self.shared_container, self.pca_df_scaled, self.cleaned_df, self.df, self.box_frame, self.box_frame_sub, self.on_button_click, self.apply_button)
            else:
                print("Column Drillhole or Depth not in data frame, drill hole plot not avaliable")
        else:
            print("Column Sample ID not in data frame, sample bar graph, and drill hole plot not avaliable")
            
            pil_image = Image.open("mineralAI_images/images_program/blank.png")
            self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))
    
            self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text = "", width=20)
            self.image_button1 = ctk.CTkLabel(self.box_frame, image=self.icon_image, text = "", width=20)
            self.image_button.grid(row=3, column=1, sticky="w", pady=0, padx=5)
            self.image_button1.grid(row=1, column=1, sticky="w", pady=0, padx=5)

            
        self.save_pc_menu.entryconfig("Save PC by element excel", state=tk.NORMAL)
        if "sample id" in self.cleaned_df.columns:
            self.save_pc_menu.entryconfig("Save PC by sample excel", state=tk.NORMAL)
        
    def export_pc_by_sample(self):
        # Export PCA results by sample
        self.pca_df_scaled = self.pca_instance.pca_df_scaled
        self.pca_df_scaled.shape
        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return

        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            self.cleaned_df["sample id"].to_excel(excel_writer, sheet_name='Sheet1', index=False, header=True)
            # self.cleaned_df["sample id"].shape
            self.pca_df_scaled.to_excel(excel_writer, sheet_name='Sheet1', startcol=1, index=False)
    
    def export_pc_by_element(self):
        # Export PCA loadings by element
        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return
        
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            columns_df = pd.DataFrame(self.df.columns, columns=['Column Names'])
            columns_df.to_excel(excel_writer, sheet_name='Sheet1', index=False, header=True)
            self.loadings.to_excel(excel_writer, sheet_name='Sheet1', startcol=1, index=False)






class supervised_learning():
    def __init__(self, df, cleaned_df, on_button_click, apply_button, box_frame):
        self.df = df  # The dataframe without 'lithology'
        self.cleaned_df = cleaned_df  # DataFrame with 'lithology' column
        self.on_button_click = on_button_click
        self.apply_button = apply_button
        self.box_frame = box_frame
        self.rows = []  # List to store rows (combobox, spinbox, delete button)
        self.prediction_df = None  # For the imported Excel prediction data

        self.y_column = None

        self.graph_icon()  # Setup the graph icon button

    def graph_icon(self):
        # Create button for displaying PC bar graph
        pil_image = Image.open("mineralAI_images/images_program/ml.png")
        # resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        # Create and place image button
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=5, column=0, sticky="w", pady=0, padx=5)

        # Bind button click event to PC graph function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.create_window))

        # Add tooltip to the button
        ToolTip(self.image_button, msg="Supervised Learning-Prediction with data")


    def create_window(self):
        # Create a new window for dynamic rows and predictions
        self.window = ctk.CTkToplevel()
        self.window.title("Dynamic Rows Example")
        self.window.geometry("800x600")
        
        # Use CTkScrollableFrame instead of canvas and scrollbar
        self.scrollable_frame = ctk.CTkScrollableFrame(self.window, width=780, height=580)
        self.scrollable_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Initialize UI elements in the new window
        self.initialize_ui()

    def initialize_ui(self):
        # Import Excel button for quick predictions
        self.import_button = ctk.CTkButton(self.scrollable_frame, text="Import Excel for Prediction", command=self.import_excel)
        self.import_button.grid(row=0, column=0, padx=10, pady=5, sticky="ew")

        # ComboBox for selecting the model
        self.combo1 = ctk.CTkComboBox(self.scrollable_frame, values=["SVM", "Random Forest"], command=self.update_model_parameters)
        self.combo1.grid(row=0, column=1, columnspan=2, padx=10, pady=5, sticky="ew")
        self.combo1.set("Select Model")
        
        # Call the method to create ComboBox for selecting y column
        self.create_y_column_combobox()

        # Plus button for adding rows manually
        self.plus_button = ctk.CTkButton(self.scrollable_frame, text="+", command=self.add_manual_row)
        self.plus_button.grid(row=6, column=0, columnspan=3, pady=10, sticky="ew")

        # Predict button (initially hidden, will be repositioned after rows are added)
        self.predict_button = ctk.CTkButton(self.scrollable_frame, text="Predict", command=self.predict_model)
        self.predict_button.grid(row=7, column=0, columnspan=3, pady=10, sticky="ew")


    def create_y_column_combobox(self):
        filtered_columns = [col for col in self.cleaned_df.columns if '_ppm' not in col and '_pct' not in col]

        ctk.CTkLabel(self.scrollable_frame, text="Select y-data").grid(row=5, column=0, padx=10, pady=5, sticky="ew")
        # Create ComboBox to select which column to use for prediction (y data)
        self.y_column_combo = ctk.CTkComboBox(self.scrollable_frame, values=filtered_columns, command = self.set_y_column)
        self.y_column_combo.set("Select Y Column")
        self.y_column_combo.grid(row=5, column=1, columnspan=2, padx=10, pady=5, sticky="ew")

    def set_y_column(self, event):
        # Store the selected y column
        self.y_column = self.y_column_combo.get()
        

    def update_model_parameters(self, selected_model):
        # Place the parameter widgets above the manual rows
        param_row_start = 1  # Starting row for parameter widgets

        try:
            self.kernel_combo.grid_forget()
            self.c_spinbox.grid_forget()
            self.coef0_spinbox.grid_forget()
            self.degree_spinbox.grid_forget()
            self.kernel_text.grid_forget()
            self.c_text.grid_forget()
            self.coef0_text.grid_forget()
            self.degree_text.grid_forget()
        except:
            pass

        try:
            self.n_estimators_spinbox.grid_forget()
            sself.max_depth_spinbox.grid_forget()
            self.min_samples_split_spinbox.grid_forget()
            self.n_estimators_text.grid_forget()
            self.max_depth_text.grid_forget()
            self.min_samples_split_text.grid_forget()
        except:
            pass

    
        
        if selected_model == "SVM":
            # Add SVM parameter widgets starting at row 2
            self.kernel_text = ctk.CTkLabel(self.scrollable_frame, text = "Select Kernel").grid(row=param_row_start, column=1, padx=10, pady=5, sticky="ew")
            self.kernel_combo = ctk.CTkComboBox(self.scrollable_frame, values=["linear", "poly", "rbf", "sigmoid"])
            self.kernel_combo.grid(row=param_row_start, column=2, padx=10, pady=5, sticky="ew")
            self.kernel_combo.set("linear")
            
            self.c_text = ctk.CTkLabel(self.scrollable_frame, text = "Select C").grid(row=param_row_start+1, column=1, padx=10, pady=5, sticky="ew")
            self.c_spinbox = ttk.Spinbox(self.scrollable_frame, from_=0.1, to=1000, increment=0.1)
            self.c_spinbox.grid(row=param_row_start+1, column=2, padx=10, pady=5, sticky="ew")
            self.c_spinbox.set("1.0")

            self.coef0_text = ctk.CTkLabel(self.scrollable_frame, text = "Select Coef0").grid(row=param_row_start+2, column=1, padx=10, pady=5, sticky="ew")
            self.coef0_spinbox = ttk.Spinbox(self.scrollable_frame, from_=0.0, to=100, increment=0.1)
            self.coef0_spinbox.grid(row=param_row_start+2, column=2, padx=10, pady=5, sticky="ew")
            self.coef0_spinbox.set("0.0")
            
            self.degree_text = ctk.CTkLabel(self.scrollable_frame, text = "Select Degree").grid(row=param_row_start+3, column=1, padx=10, pady=5, sticky="ew")
            self.degree_spinbox = ttk.Spinbox(self.scrollable_frame, from_=1.0, to=10)
            self.degree_spinbox.grid(row=param_row_start + 3, column=2, padx=10, pady=5, sticky="ew")
            self.degree_spinbox.set("3")

        elif selected_model == "Random Forest":
            # Add Random Forest parameter widgets starting at row 2
            self.n_estimators_text = ctk.CTkLabel(self.scrollable_frame, text = "Select n_estimators").grid(row=param_row_start, column=1, padx=10, pady=5, sticky="ew")
            self.n_estimators_spinbox = ttk.Spinbox(self.scrollable_frame, from_=10.0, to=1000)
            self.n_estimators_spinbox.grid(row=param_row_start, column=2, padx=10, pady=5, sticky="ew")
            self.n_estimators_spinbox.set("100")

            self.max_depth_text = ctk.CTkLabel(self.scrollable_frame, text = "Select max_depth").grid(row=param_row_start+1, column=1, padx=10, pady=5, sticky="ew")
            self.max_depth_spinbox = ttk.Spinbox(self.scrollable_frame, from_=1.0, to=100)
            self.max_depth_spinbox.grid(row=param_row_start+1, column=2, padx=10, pady=5, sticky="ew")
            self.max_depth_spinbox.set("None")

            self.min_samples_split_text = ctk.CTkLabel(self.scrollable_frame, text = "Select min_samples_split").grid(row=param_row_start+2, column=1, padx=10, pady=5, sticky="ew")
            self.min_samples_split_spinbox = ttk.Spinbox(self.scrollable_frame, from_=2.0, to=10)
            self.min_samples_split_spinbox.grid(row=param_row_start+2, column=2, padx=10, pady=5, sticky="ew")
            self.min_samples_split_spinbox.set("2")

    def add_manual_row(self, column_name=None, values=None):
        # Continue placing manual rows below parameter widgets
        param_row_end = 8  # End row of the parameter widgets
        current_row = len(self.rows) + param_row_end + 1  # Start adding rows below parameters

        # Your existing code for adding manual rows remains the same, except for updating the row number
        if len(self.rows) < len(self.df.columns):
            combobox = ctk.CTkComboBox(self.scrollable_frame, values=self.df.columns.tolist())
            if column_name:
                combobox.set(column_name)
            combobox.grid(row=current_row, column=0, padx=10, pady=5, sticky="ew")

            spinbox = ttk.Spinbox(self.scrollable_frame, from_=0.0, to=100)
            spinbox.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")
            if values:
                spinbox.set(values[0])  # Set the spinbox value to the first row of the column

            delete_button = ctk.CTkButton(self.scrollable_frame, text="🗑️", width=20, command=lambda: self.delete_row(combobox))
            delete_button.grid(row=current_row, column=2, padx=10, pady=5, sticky="ew")

            self.rows.append([combobox, spinbox, delete_button])

            self.predict_button.grid(row=current_row + 1, column=0, columnspan=3, pady=10, sticky="ew")

            if len(self.rows) >= len(self.df.columns):
                self.plus_button.configure(state="disabled")
        else:
            self.plus_button.configure(state="disabled")

    def import_excel(self):
        # Open a file dialog to select the Excel file for prediction
        file_path = fd.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            # Load the dataset for prediction
            self.prediction_df = pd.read_excel(file_path)

            # Clear existing rows before adding new ones
            self.clear_existing_rows()

            # Add rows based on the new Excel data
            self.add_prediction_rows()

    def clear_existing_rows(self):
        # Clear all existing rows (comboboxes, spinboxes, and buttons)
        for row in self.rows:
            for widget in row:
                widget.destroy()

        # Clear the rows list and reset the count
        self.rows.clear()

    def add_prediction_rows(self):
        # Populate comboboxes and spinboxes with values from the uploaded Excel sheet
        if self.prediction_df is not None:
            # Iterate over the columns of the Excel file
            for column in self.prediction_df.columns:
                # Create combobox and spinbox for each feature (column in Excel)
                self.add_manual_row(column_name=column, values=self.prediction_df[column].tolist())

    def delete_row(self, combobox):
        # Find the row index dynamically based on the combobox
        row_index = next((i for i, row in enumerate(self.rows) if row[0] == combobox), None)

        if row_index is not None:
            # Remove the widgets for the specified row index
            for widget in self.rows[row_index]:
                widget.destroy()

            # Remove the row from the list
            self.rows.pop(row_index)

            # Re-enable the plus button if rows are deleted and below the maximum
            if len(self.rows) < len(self.df.columns):
                self.plus_button.configure(state="normal")

            # Reposition the remaining rows
            for idx, row in enumerate(self.rows):
                for col_idx, widget in enumerate(row):
                    widget.grid(row=idx + 3, column=col_idx, padx=10, pady=5, sticky="ew")

    def predict_model(self):
        # Ensure that the model is trained on the cleaned_df
        selected_model = self.combo1.get()
                
        if not self.y_column:
            self.display_warning_message("Please select a Y column for prediction.")
            return
            
        input_data = {}
        missing_columns = []  # To store columns not found in df
    
        # Clear previous messages
        self.clear_warning_labels()
    
        # Collect data from the dynamically created ComboBoxes and Spinboxes for prediction
        for idx, row in enumerate(self.rows):
            feature_name = row[0].get().lower()
            feature_value = float(row[1].get())  # Spinbox for feature value
            
            # Check if the feature_name is present in the cleaned_df
            if feature_name in self.cleaned_df.columns.str.lower():
                input_data[feature_name] = feature_value
            else:
                missing_columns.append(feature_name)  # Add missing column name to the list
        
        # Display the warning message in the app frame
        if missing_columns:
            missing_columns_message = f"Warning: The following columns were not found in the dataset and were skipped: {', '.join(missing_columns)}"
            self.display_warning_message1(missing_columns_message)
    
        # Create a DataFrame for the input data for prediction
        if input_data:
            input_df = pd.DataFrame([input_data])
    
            # Access columns in cleaned_df using the stripped and lowercased feature names
            X_train_filtered = self.cleaned_df[input_df.columns.str.lower()]
            y_train_filtered = self.cleaned_df[self.y_column]
    
            # SVM model setup with selected parameters
            if selected_model == "SVM":
                kernel = self.kernel_combo.get()
                C = float(self.c_spinbox.get())
                coef0 = float(self.coef0_spinbox.get())
                degree = int(self.degree_spinbox.get())
    
                model = make_pipeline(StandardScaler(), SVC(kernel=kernel, C=C, coef0=coef0, degree=degree, random_state=42))
    
            # Random Forest model setup with selected parameters
            elif selected_model == "Random Forest":
                n_estimators = int(self.n_estimators_spinbox.get())
                max_depth = None if self.max_depth_spinbox.get() == "None" else int(self.max_depth_spinbox.get())
                min_samples_split = int(self.min_samples_split_spinbox.get())
    
                model = RandomForestClassifier(n_estimators=n_estimators, max_depth=max_depth, min_samples_split=min_samples_split, random_state=42)
    
            else:
                self.display_warning_message("Please select a model.")
                return
            
            # Fit the model and make predictions based on the values entered in the spinboxes
            model.fit(X_train_filtered, y_train_filtered)
            predictions = model.predict(input_df)
            
            # Display the prediction result
            result_text = f"Prediction: {predictions[0]}"
            self.display_warning_message(result_text)
        else:
            self.display_warning_message("No valid columns found for prediction.")
    
    def display_warning_message(self, message):
        # Create a label to display the warning message in the UI
        self.warning_label = ctk.CTkLabel(self.scrollable_frame, text=message, text_color="red", wraplength=600)
        self.warning_label.grid(row=len(self.rows) + 10, column=0, columnspan=3, padx=10, pady=5, sticky="ew")

    def display_warning_message1(self, message):
        # Create a label to display the warning message in the UI
        self.warning_label = ctk.CTkLabel(self.scrollable_frame, text=message, text_color="red", wraplength=600)
        self.warning_label.grid(row=len(self.rows) + 11, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
    
    def clear_warning_labels(self):
        # Clear the previous warning label if it exists
        try:
            self.warning_label.destroy()
        except AttributeError:
            pass
    
    





class SharedContainer:
    def __init__(self, parent):
        # Initialize the tab container and content container
        self.tab_container = ctk.CTkFrame(parent)
        self.tab_container.pack(fill='x')
        self.content_container = ctk.CTkFrame(parent)
        self.content_container.pack(fill='both', expand=True)
        self.tabs = []
        self.current_tab = None

        # Add a plus button for creating new tabs
        self.plus_button = ctk.CTkButton(self.tab_container, text="+", width=25, height=25, border_width=0, hover_color="darkgrey", text_color="black", fg_color="transparent", command=self.create_tab_button)
        #self.plus_button.pack(side=tk.RIGHT, padx=3, pady=3)

    def create_tab(self, title="New Tab"):
        # Create a new tab with a title
        tab_frame = tk.Frame(self.tab_container, bd=1, relief=tk.RAISED, height=30)  # Set height
        tab_label = ctk.CTkLabel(tab_frame, text=title)
        tab_label.bind("<Double-1>", lambda event: self.edit_tab_label(tab_label))
        custom_font = ctk.CTkFont(size=16)
        close_button = ctk.CTkButton(tab_frame, text="x", width=25, height=25, border_width=0, font=custom_font, hover_color="darkgrey", text_color="black", command=lambda: self.close_tab(tab_frame, content_frame))
        
        close_button.pack(side=tk.RIGHT, padx=3, pady=3, fill=tk.Y)  
        tab_label.pack(side=tk.LEFT, padx=3, pady=3, fill=tk.Y) 

        self.plus_button.pack_forget()
        tab_frame.pack(side=tk.LEFT, padx=2, pady=2)
        self.plus_button.pack(side=tk.LEFT, padx=3, pady=3)

        content_frame = ctk.CTkFrame(self.content_container)
        content_frame.pack(fill='both', expand=True)

        tab_frame.bind("<Button-1>", lambda event: self.select_tab(tab_frame, content_frame))
        tab_label.bind("<Button-1>", lambda event: self.select_tab(tab_frame, content_frame))
        close_button.bind("<Button-1>", lambda event: self.select_tab(tab_frame, content_frame))

        self.tabs.append((tab_frame, content_frame))
        self.select_tab(tab_frame, content_frame)
        self.adjust_tab_widths()
    
    def adjust_tab_widths(self):
        self.max_tab_width = 150
        
        # Adjust the width of each tab icon
        total_width = self.tab_container.winfo_width() - self.plus_button.winfo_width() - 10
        tab_width = min(self.max_tab_width, total_width // len(self.tabs) - 4 if self.tabs else total_width)
    
        for tab_frame, _ in self.tabs:
            tab_frame.config(width=tab_width)
            tab_frame.pack_propagate(False)  # Prevent the frame from resizing to fit its content


    def create_tab_button(self):
        # Create a new tab with default title "New Tab"
        self.create_tab()

    def select_tab(self, tab_frame, content_frame):
        # Highlight the selected tab and show its content
        if self.current_tab:
            self.current_tab[0].config(bg='lightgrey')
            for widget in self.current_tab[0].winfo_children():
                widget.configure(fg_color='lightgrey')

        tab_frame.config(bg='gray95')
        for widget in tab_frame.winfo_children():
            widget.configure(fg_color='gray95')

        for frame in self.content_container.winfo_children():
            frame.pack_forget()

        content_frame.pack(fill='both', expand=True)
        self.current_tab = (tab_frame, content_frame)

    def close_tab(self, tab_frame, content_frame):
        # Close the selected tab and select the previous tab if available
        if self.current_tab == (tab_frame, content_frame):
            self.current_tab = None
        tab_frame.destroy()
        content_frame.destroy()
        self.tabs.remove((tab_frame, content_frame))
        if self.tabs:
            self.select_tab(*self.tabs[-1])
        self.adjust_tab_widths()

    def edit_tab_label(self, tab_label):
        # Enable editing of the tab label
        current_text = tab_label.cget("text")
        tab_frame = tab_label.master

        entry = tk.Entry(tab_frame, width=10)
        entry.insert(0, current_text)
        entry.pack(side=tk.LEFT, padx=3, pady=3)
        entry.bind("<Return>", lambda event: self.update_tab_label(tab_label, entry))
        entry.bind("<FocusOut>", lambda event: self.update_tab_label(tab_label, entry))
        entry.focus()

        tab_label.pack_forget()

    def update_tab_label(self, tab_label, entry):
        # Update the tab label with new text
        new_text = entry.get()
        tab_label.configure(text=new_text)
        entry.destroy()
        tab_label.pack(side=tk.LEFT, padx=3, pady=3)













class loading_class:
    def __init__(self, loadings, shared_container, box_frame, box_frame_sub, on_button_click, apply_button):
        # Initialize the loading class with necessary parameters
        self.apply_button = apply_button
        self.shared_container = shared_container
        self.on_button_click = on_button_click
        self.loadings = loadings

        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.var = IntVar()
        self.group_list = []

        # Call method to create the PC graph button
        self.Graph_PC()

    def Graph_PC(self):
        # Create button for displaying PC bar graph
        pil_image = Image.open("mineralAI_images/images_program/element-svgrepo-com.png")
        # resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        # Debug prints for loaded images
        # print(pil_image)
        # print(self.icon_image)

        # Create and place image button
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=1, column=0, sticky="w", pady=0, padx=5)

        # Bind button click event to PC graph function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.Graph_PC_sub))

        # Add tooltip to the button
        ToolTip(self.image_button, msg="PC Bar Graph by Elements")

    def Graph_PC_sub(self):
        # Create elements and groups selection UI
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Create and pack checkbox for sorting loadings
        self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Sort Loadings", variable=self.var)
        self.checkbox.pack(side="top", anchor="w", padx=5, pady=3)

        # Create and pack listbox for element selection
        self.elements_listbox = tk.Listbox(self.box_frame_sub, selectmode=tk.MULTIPLE)
        for element in self.loadings.index:
            self.elements_listbox.insert(tk.END, element)
        self.elements_listbox.pack(side="top", padx=5, pady=5)

        # Function to deselect all elements
        def deselect_all():
            self.elements_listbox.selection_clear(0, tk.END)

        # Bind deselect function to the listbox
        self.elements_listbox.bind('<Command-d>', lambda event: deselect_all())

        # Create and pack button to add selected elements as a group
        self.add_group_button = ctk.CTkButton(self.box_frame_sub, text="Add Group", command=self.add_group)
        self.add_group_button.pack(side="top", padx=5, pady=5)

        # Create and pack listbox to display added groups
        self.group_listbox = tk.Listbox(self.box_frame_sub, selectmode=tk.MULTIPLE)
        self.group_listbox.pack(side="top", padx=5, pady=5)

        # Create and pack apply button
        self.apply_Graph_PC = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.update_plots)
        self.apply_Graph_PC.pack(side="top", padx=5, pady=5)

        # Create and pack button to clear all groups
        self.clear_groups_button = ctk.CTkButton(self.box_frame_sub, text="Clear Groups", command=self.clear_groups)
        self.clear_groups_button.pack(side="top", padx=5, pady=5)

    def clear_groups(self):
        # Clear all groups
        self.group_list.clear()
        self.group_listbox.delete(0, tk.END)

    def add_group(self):
        # Add selected elements as a group
        selected_elements = [self.elements_listbox.get(i) for i in self.elements_listbox.curselection()]
        self.group_list.append(selected_elements)
        self.group_listbox.insert(tk.END, f"Group {len(self.group_list)}: {', '.join(selected_elements)}")

    def normalize(self, values, vmin=None, vmax=None):
        # Normalize values for color scaling
        vmin = vmin if vmin is not None else np.min(values)
        vmax = vmax if vmax is not None else np.max(values)
        norm_values = (values - vmin) / (vmax - vmin)
        return norm_values

    def plot_pc_barchart(self, data, pc_name, ax, sort, group_indices):
        # Plot the PC bar chart
        if sort:
            data = data.sort_values(ascending=False)
        norm_values = self.normalize(data, vmin=-max(abs(data)), vmax=max(abs(data)))
        bar_colors = plt.cm.coolwarm(norm_values)
        data.plot(kind='barh', color=bar_colors, ax=ax, fontsize="x-small")
        ax.set_title(pc_name)
        ax.grid(True, linestyle='--', linewidth=0.5)
        ax.set_ylabel(' ')

        # Draw horizontal lines to separate groups
        for idx in group_indices:
            ax.axhline(y=idx - 0.5, color='black', linewidth=1)

        # Update y-axis labels
        labels = ax.get_yticklabels()
        new_labels = [label.get_text().replace('_ppm', '')
                                      .replace('_pct', '')
                                      .replace('_Howell', '') for label in labels]
        ax.set_yticklabels(new_labels)

    def update_plots(self):
        # Update plots with selected groups and options
        sort = self.var.get() == 1
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("Element BarGraph")

        # Close all existing plots and clear widgets
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()

        # Create a new figure and axes for the plots
        fig, axes = plt.subplots(ncols=len(self.loadings.columns), figsize=(15, 8))
        if len(self.loadings.columns) == 1:
            axes = [axes]

        # Iterate over each principal component
        for i in range(len(self.loadings.columns)):
            combined_data = pd.Series(dtype='float64')
            group_indices = []
            current_index = 0
            selected_indices = self.group_listbox.curselection()
            
            # Combine data for selected groups
            for index in selected_indices:
                group = self.group_list[index]
                group_data = self.loadings.loc[group, f'PC{i+1}']
                if sort:
                    group_data = group_data.sort_values(ascending=False)
                combined_data = pd.concat([combined_data, group_data])
                current_index += len(group_data)
                group_indices.append(current_index)

            # Plot the combined data
            if not combined_data.empty:
                self.plot_pc_barchart(combined_data, f'PC{i+1}', axes[i], sort=False, group_indices=group_indices)

        # Adjust layout and add colorbar
        plt.tight_layout()
        fig_c = plt.figure(figsize=(12, 1))
        min_val = self.loadings.min().min()
        max_val = self.loadings.max().max()
        norm = plt.Normalize(vmin=min_val, vmax=max_val)
        cmap = plt.cm.coolwarm
        cbar = fig_c.colorbar(plt.cm.ScalarMappable(norm=norm, cmap=cmap), ax=axes, orientation='horizontal', fraction=0.02)
        
        # Add the plot to the tkinter canvas
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()

        # Add navigation toolbar to the canvas
        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)







class loading_cluster:
    def __init__(self, shared_container, cluster, loadings, box_frame, box_frame_sub, on_button_click, apply_button, legend_frame):
        # Initialize the loading cluster class with necessary parameters
        self.loadings = loadings
        self.shared_container = shared_container
        self.cluster_result = cluster
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.legend_frame = legend_frame
        self.on_button_click = on_button_click
        self.apply_button = apply_button
        self.cluster_sort_var = IntVar()
        self.graph_data_df = None

        self.plot_cluster()

    def plot_cluster(self):
        # Create button for displaying cluster bar graph
        pil_image = Image.open("mineralAI_images/images_program/ele.png")
        # resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))
        
        # Debug prints for loaded images
        print(pil_image)
        print(self.icon_image)
        
        # Create and place image button
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=3, column=0, sticky="w", pady=0, padx=5)

        # Bind button click event to cluster function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_cluster_sub0))

        ToolTip(self.image_button, msg="PC Cluster Bar Graph by Elements")

    def plot_cluster_sub0(self):
        # Create elements and clusters selection UI
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Create and pack checkbox for sorting clusters
        self.cluster_sort_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Sort Clusters Together", variable=self.cluster_sort_var)
        # self.widget_dict["loading_check"] = self.cluster_sort_checkbox
        self.cluster_sort_checkbox.custom_name = "loading_check"
        self.cluster_sort_checkbox.grid(columnspan=2, row=0, column=0, pady=(5,0), padx=5)
            
        # Create and pack slider for number of clusters
        self.cluster_text = ctk.CTkLabel(self.box_frame_sub, text="Number of Clusters:")
        self.k_slider = tk.Scale(self.box_frame_sub, from_=2, to=10, orient=tk.HORIZONTAL)
            
        # Create and pack apply button
        self.apply_plot_3d_cluster = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.update_plots)
        self.apply_plot_3d_cluster.grid(columnspan=2, row=6, column=0, pady=10, padx=5)

        # Create and pack save button
        self.save_button = ctk.CTkButton(self.box_frame_sub, text="Save Clusters to Excel", command=self.save_loadings_to_excel)
        self.save_button.pack(side="bottom", padx=5, pady=5)

        # Define the values for the Spinbox
        linkage_box_values = ('ward', 'complete', 'average', 'single')
        self.linkage_box = ttk.Combobox(self.box_frame_sub, values=linkage_box_values, state="readonly")
        self.linkage_box.current(0)

        affinity_box_values = ('nearest_neighbors', 'rbf', 'precomputed', 'precomputed_nearest_neighbors')
        self.affinity_box = ttk.Combobox(self.box_frame_sub, values=affinity_box_values, state="readonly")
        self.affinity_box.current(0)
                
        self.min_sample_box = CTkSpinbox(self.box_frame_sub, step_size=1, min_value=2, max_value=100, width = 110) #, command=self.runeps)
        self.eps_box = CTkSpinbox(self.box_frame_sub, step_size=0.1, min_value=0.1, max_value=100, width = 110)
                    
        self.eps_text = ctk.CTkLabel(self.box_frame_sub, text="eps")
        self.min_sample_text = ctk.CTkLabel(self.box_frame_sub, text="min_sample")

        self.plot_cluster_sub(self.cluster_result)
    
        
    def plot_cluster_sub(self, cluster):
        self.box_frame_sub.grid_propagate(False)
        self.cluster_result = cluster
        
        self.linkage_box.grid_forget()
        self.affinity_box.grid_forget()
        self.min_sample_box.grid_forget()
        self.eps_box.grid_forget()
        self.eps_text.grid_forget()
        self.min_sample_text.grid_forget()
        
        try:
            self.param_text.grid_forget()
        except:
            pass

        if self.cluster_result == "Hierarchical":
            print("Hierarchical Clustering Selected")  # Debug print
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Linkage Type:")
            self.linkage_box.grid(columnspan=2, row=2, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
            
        elif self.cluster_result == "Spectral":
            print("Spectral Clustering Selected")  # Debug print
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Affinity Type:")
            self.affinity_box.grid(columnspan=2, row=2, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
    
        elif self.cluster_result == "DBSCAN":
            print("DBSCAN Clustering Selected")  # Debug print
            # self.runeps() 
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="eps & min_sample value:")

            self.eps_text.grid(row=3, column=0, pady=0, padx=(5,0), sticky="w")
            self.min_sample_text.grid(row=2, column=0, pady=5, padx=(5,0), sticky="w")
        
            self.eps_box.grid(row=3, column=1, pady=0, padx=(0,5), sticky="e")
            self.min_sample_box.grid(row=2, column=1, pady=10, padx=(0,5), sticky="e")
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
            
        else:
            pass
            
            
        if self.cluster_result in ["DBSCAN", "Mean Shift", "Affinity Propagation"]:
            self.cluster_text.grid_forget()
            self.k_slider.grid_forget()
        else:
            self.cluster_text.grid(columnspan=2,row=4, column=0, pady=0, padx=5)
            self.k_slider.grid(columnspan=2,row=5, column=0, pady=(3,5), padx=5)

        self.yellowbrick()


    def normalize(self, values, vmin=None, vmax=None):
        # Normalize values for color scaling
        vmin = vmin if vmin is not None else np.min(values)
        vmax = vmax if vmin is not None else np.max(values)
        norm_values = (values - vmin) / (vmax - vmin)
        return norm_values

    def plot_pca_barchart(self, data, entry_name, ax, cluster_colors, sorted_labels):
        # Plot the PCA bar chart with clusters
        cluster_palette = plt.get_cmap('tab20', len(np.unique(cluster_colors)))
        colors = cluster_palette(cluster_colors)
        data.plot(kind='barh', color=colors, ax=ax, fontsize="xx-small")
        ax.set_title(entry_name)
        ax.grid(True, linestyle='--', linewidth=0.5)
        ax.set_yticks(range(len(data)))
        ax.set_yticklabels(sorted_labels, fontsize="xx-small") 
        ax.set_ylabel('')

        # Update y-axis labels
        labels = ax.get_yticklabels()
        new_labels = [label.get_text().replace('_ppm', '')
                                      .replace('_pct', '')
                                      .replace('_Howell', '') for label in labels]
        ax.set_yticklabels(new_labels)

    # def estimate_eps(self, k):
    #     # Estimate the optimal eps value for DBSCAN clustering
    #     from sklearn.neighbors import NearestNeighbors
    #     for widget in self.legend_frame.winfo_children():
    #         widget.destroy()

    #     # Compute the k-nearest neighbors
    #     nearest_neighbors = NearestNeighbors(n_neighbors=k)
    #     nearest_neighbors.fit(self.loadings)
    #     distances, indices = nearest_neighbors.kneighbors(self.loadings)

    #     # Sort the distances (distances are ordered for each point)
    #     distances = np.sort(distances, axis=0)
    #     distances = distances[:, 1]  # Taking the distance to the k-th nearest neighbor

    #     # Plot the k-distance graph
    #     fig, ax = plt.subplots(figsize=(3, 2))
    #     ax.plot(distances)
    #     ax.set_title('k-Distance Graph')
    #     ax.set_xlabel('Points sorted by distance')
    #     ax.set_ylabel(f'{k}th Nearest Neighbor Distance')

    #     # Add the plot to the tkinter canvas
    #     self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
    #     self.canvas1.get_tk_widget().pack(fill="both", expand=True)
    #     plt.show()

    #     elbow_point_index = 20
    #     eps_value = distances[elbow_point_index]
    #     return eps_value

    def update_plots(self):
        # Update plots with selected clustering method and options
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("Cluster Bar Graph by Elements")

        # Close all existing plots and clear widgets
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        k = self.k_slider.get()
        
        if hasattr(self, 'cluster_print') and self.cluster_print is not None:
            self.cluster_print.destroy()
            del self.cluster_print

        # Get the selected clustering method
        cluster_result = self.cluster_result
        self.cluster_print = ctk.CTkLabel(self.box_frame_sub, text=f"Cluster: {cluster_result}")
        self.cluster_print.grid(columnspan=2, row=6, column=0, pady=5, padx=5)
        
        # Set up the pipeline for different clustering algorithms
        if cluster_result == "K-mean":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
            ])
            
        elif cluster_result == "DBSCAN":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", DBSCAN(eps=self.eps_box.get(), min_samples=self.min_samples_box.get()))
            ])

        elif cluster_result == "Mean Shift":  
            from sklearn.cluster import estimate_bandwidth
            bandwidth_value = estimate_bandwidth(self.loadings, quantile=0.2)
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", MeanShift(bandwidth=bandwidth_value))
            ])
            
        elif cluster_result == "Spectral":  
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", SpectralClustering(n_clusters=k, affinity='nearest_neighbors', random_state=0))
            ])

        elif cluster_result == "GMM":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", GaussianMixture(n_components=k, random_state=0))
            ])

        elif cluster_result == "Affinity Propagation":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AffinityPropagation(random_state=0))
            ])

        elif cluster_result == "Hierarchical":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AgglomerativeClustering(n_clusters=k, linkage='ward'))
            ])

        elif cluster_result == "BIRCH":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", Birch(n_clusters=k))
            ])
            
        # Fit the pipeline to the data
        self.pipe.fit(self.loadings)
        self.clustering_results = pd.DataFrame(index=self.loadings.index)
        
        # Get clustering results
        if cluster_result == "GMM":  
            self.clustering_results['cluster'] = self.pipe.named_steps["model"].predict(self.loadings)
        else:
            self.clustering_results['cluster'] = self.pipe.named_steps['model'].labels_

        fig, axes = plt.subplots(ncols=len(self.loadings.columns), figsize=(15, 8))
        
        cluster_colors = self.clustering_results['cluster'].values
        self.clustering_results['element'] = self.loadings.index
        
        graph_data = []
        sort_clusters = self.cluster_sort_var.get() == 1
        
        for i in range(len(self.loadings.columns)):
            pc_name = f'PC{i+1}'
            data = self.loadings[pc_name]

            if sort_clusters:
                # Sort data by clusters
                temp_df = pd.DataFrame({
                    'data': data,
                    'cluster': cluster_colors,
                    'element':  self.clustering_results['element'],
                    'PC': pc_name
                })

                sorted_df = temp_df.sort_values(by='cluster')
                sorted_data = sorted_df['data']
                sorted_cluster_colors = sorted_df['cluster'].values
                sorted_labels = sorted_df['element'].values
                self.plot_pca_barchart(sorted_data, pc_name, axes[i], sorted_cluster_colors, sorted_labels)
                graph_data.append(sorted_df)
                
            else:
                temp_df = pd.DataFrame({
                    'data': data,
                    'cluster': cluster_colors,
                    'element':  self.clustering_results['element'],
                    'PC': pc_name
                })
                graph_data.append(temp_df)
                labels = temp_df['element'].values
                self.plot_pca_barchart(data, pc_name, axes[i], cluster_colors, labels)

        self.graph_data_df = pd.concat(graph_data)

        plt.tight_layout()
        
        # Add the plot to the tkinter canvas
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()

        # Add navigation toolbar to the canvas
        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)


    def save_loadings_to_excel(self):
        # Save the loadings to an Excel file
        if self.graph_data_df is None:
            print("No data to save. Ensure clusters are calculated and graph is generated.")
            return
        
        # Prompt user to select save location and filename
        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return

        # Create a pivot table with the clustering results
        pivot_data = self.graph_data_df.pivot_table(index='element', columns='PC', values='data', sort=False)
        pivot_data.columns = [f'{col}' for col in pivot_data.columns]

        # Save the pivot table to an Excel file
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            pivot_data.to_excel(excel_writer, index=True)

        # Load the saved workbook and worksheet
        wb = load_workbook(file_name)
        ws = wb.active

        # Define color map for clusters
        colormap = plt.get_cmap('Set2')
        cluster_colors = ['FF{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)) for r, g, b, _ in [colormap(i) for i in range(colormap.N)]]
        cluster_fill = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in cluster_colors]

        # Apply cluster colors to the Excel file
        for i in range(len(self.graph_data_df)):
            element = self.graph_data_df.iloc[i]['element']
            cluster = self.graph_data_df.iloc[i]['cluster']
            row = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
            for cell in row:
                if cell[0].value == element:
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=cell[0].row, column=col).fill = cluster_fill[cluster]

        wb.save(file_name)
        print(f"Loadings saved to {file_name}")

    def yellowbrick(self):
        try:
            # Create and display the KElbowVisualizer for clustering
            plt.rcParams.update({'font.size': 10})
            
            fig = plt.figure(figsize=(3, 2))
            ax = fig.add_subplot(111)
            ax.axis('off')
            
            model = self.pipe.named_steps['model']
            if 'n_clusters' in model.get_params():
                del model.get_params()['n_clusters']
        
            # Create and fit the KElbowVisualizer
            visualizer = KElbowVisualizer(model, k=(1, 9), timing=False, title="KElbowVisualizer").fit(self.loadings)
            
            visualizer.ax.set_yticklabels([])
            visualizer.finalize()
            
            # Add the plot to the tkinter canvas
            self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
            toolbar = CustomToolbar(self.canvas1, self.legend_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            self.canvas1.get_tk_widget().pack(fill="both", expand=True)
            
            plt.close(fig)
            
        except:
            pass


        





class sample_class:
    def __init__(self, shared_container, pca_df_scaled, df, cleaned_df, box_frame, box_frame_sub, on_button_click, apply_button):
        # Initialize the sample class with necessary parameters
        self.apply_button = apply_button
        self.shared_container = shared_container
        self.on_button_click = on_button_click
        self.pca_df_scaled = pca_df_scaled
        self.df = df
        self.cleaned_df = cleaned_df
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.var = IntVar()  # Variable for checkbox state
        self.Graph_PCA()  # Initialize the PCA graph

    def Graph_PCA(self):
        # Create button for displaying PC bar graph by samples
        pil_image = Image.open("mineralAI_images/images_program/sample-screen-svgrepo-com.png")
        # resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        # Debug prints for loaded images
        print(pil_image)
        print(self.icon_image)

        # Create and place image button
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=1, column=1, sticky="w", pady=0, padx=5)

        # Bind button click event to PCA graph function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.Graph_PCA_sub))

        ToolTip(self.image_button, msg="PC Bar Graph by Samples")

    def Graph_PCA_sub(self):
        # Create UI for PCA graph options and actions
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Create and pack checkbox for sorting loadings
        self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Sort Loadings", variable=self.var)
        self.checkbox.pack(side="top", anchor="w", padx=5, pady=3)

        # Create and pack apply button
        self.apply_Graph_PCA = ctk.CTkButton(self.box_frame_sub, text="apply", command=self.update_plots)
        self.apply_Graph_PCA.pack(side="top", padx=5, pady=5)

    def normalize(self, values, vmin=None, vmax=None):
        # Normalize values for color scaling
        vmin = vmin if vmin is not None else np.min(values)
        vmax = vmax if vmin is not None else np.max(values)
        norm_values = (values - vmin) / (vmax - vmin)
        return norm_values

    def plot_pca_barchart(self, data, entry_name, ax, sort):
        # Plot the PCA bar chart
        if sort:
            data = data.sort_values(ascending=False)
        norm_values = self.normalize(data, vmin=-max(abs(data)), vmax=max(abs(data)))
        bar_colors = plt.cm.coolwarm(norm_values)
        data.plot(kind='barh', color=bar_colors, ax=ax, fontsize="xx-small")
        ax.set_title(entry_name)
        ax.grid(True, linestyle='--', linewidth=0.5)
        ax.set_ylabel('')

    def update_plots(self):
        # Update plots with selected options
        sort = self.var.get() == 1  # Get the state of the sort checkbox
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("Sample BarGraph")

        # Close all existing plots and clear widgets
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()

        # Determine the number of samples and PCs
        num_samples = len(self.pca_df_scaled)
        num_pcs = len(self.pca_df_scaled.columns)
        
        # Create subplots for each PC
        fig, axes = plt.subplots(ncols=num_pcs, figsize=(9, 10))

        for i in range(num_pcs):
            pc_name = f'PC{i+1}'
            data = self.pca_df_scaled.iloc[:, i]
            if sort:
                sorted_indices = data.sort_values(ascending=False).index
                data = data.loc[sorted_indices]
                y_tick_labels = self.cleaned_df.loc[sorted_indices, 'sample id']
            else:
                y_tick_labels = self.cleaned_df['sample id']

            # Plot PCA bar chart for each PC
            self.plot_pca_barchart(data, pc_name, axes[i], sort)
            axes[i].set_yticks(range(num_samples))
            axes[i].set_yticklabels(y_tick_labels, fontsize="xx-small")
            axes[i].set_ylabel('')

        plt.tight_layout()

        # Create and display the color map
        min_val = self.pca_df_scaled.min().min()
        max_val = self.pca_df_scaled.max().max()
        norm = plt.Normalize(vmin=min_val, vmax=max_val)
        cmap = plt.cm.coolwarm

        # Add the plot to the tkinter canvas
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()

        # Add navigation toolbar to the canvas
        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)







class sample_cluster:
    def __init__(self, shared_container, cluster, pca_df_scaled, df, cleaned_df, box_frame, box_frame_sub, on_button_click, apply_button, legend_frame):
        # Initialize the PCCluster class with necessary parameters
        self.pca_df_scaled = pca_df_scaled
        self.shared_container = shared_container
        self.cluster_result = cluster
        self.df = df.dropna()
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.legend_frame = legend_frame
        self.cleaned_df = cleaned_df
        self.on_button_click = on_button_click
        self.apply_button = apply_button

        self.var = IntVar()
        self.cluster_sort_var = IntVar()
        self.graph_data_df = None
        self.plot_cluster()
    
    def plot_cluster(self):
        # Create button for displaying cluster bar graph by samples
        pil_image = Image.open("mineralAI_images/images_program/id.png")
        # resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text = "", width=20)
        self.image_button.grid(row=3, column=1, sticky="w", pady=0, padx=5)

        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_cluster_sub0))

        ToolTip(self.image_button, msg="PC Cluster Bar Graph by Samples")
        
        # self.plot_cluster_button = ctk.CTkButton(self.box_frame, text="BG", command=lambda: self.on_button_click(self.plot_cluster_button, self.plot_cluster_sub), width=20)
        # self.plot_cluster_button.grid(row=3, column=3, sticky="w", pady=(5, 0), padx=0)

    def plot_cluster_sub0(self):
        # Create UI for cluster options and actions
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()
        
        self.cluster_sort_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Sort Clusters Together", variable=self.cluster_sort_var)
        # self.widget_dict["sample_check"] = self.cluster_sort_checkbox
        self.cluster_sort_checkbox.custom_name = "sample_check"
        self.cluster_sort_checkbox.grid(columnspan=2, row=0, column=0, pady=(5,0), padx=5)

        self.cluster_text = ctk.CTkLabel(self.box_frame_sub, text="Number of Clusters:")
        self.k_slider = tk.Scale(self.box_frame_sub, from_=2, to=10, orient=tk.HORIZONTAL)

        self.apply_plot_3d_cluster = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.show_cluster)
        self.apply_plot_3d_cluster.grid(columnspan=2, row=6, column=0, pady=10, padx=5)

        self.save_button = ctk.CTkButton(self.box_frame_sub, text="Save Clusters to Excel", command=self.save_clusters_to_excel)
        self.save_button.pack(side="bottom", padx=5, pady=5)

        # Define the values for the Spinbox
        linkage_box_values = ('ward', 'complete', 'average', 'single')
        self.linkage_box = ttk.Combobox(self.box_frame_sub, values=linkage_box_values, state="readonly")
        self.linkage_box.current(0)

        affinity_box_values = ('nearest_neighbors', 'rbf', 'precomputed', 'precomputed_nearest_neighbors')
        self.affinity_box = ttk.Combobox(self.box_frame_sub, values=affinity_box_values, state="readonly")
        self.affinity_box.current(0)
                
        self.min_sample_box = CTkSpinbox(self.box_frame_sub, step_size=1, min_value=2, max_value=100, width = 110) #, command=self.runeps)
        self.eps_box = CTkSpinbox(self.box_frame_sub, step_size=0.1, min_value=0.1, max_value=100, width = 110)
                    
        self.eps_text = ctk.CTkLabel(self.box_frame_sub, text="eps")
        self.min_sample_text = ctk.CTkLabel(self.box_frame_sub, text="min_sample")

        self.plot_cluster_sub(self.cluster_result)

    def plot_cluster_sub(self, cluster):
        self.box_frame_sub.grid_propagate(False)
        self.cluster_result = cluster
        
        self.linkage_box.grid_forget()
        self.affinity_box.grid_forget()
        self.min_sample_box.grid_forget()
        self.eps_box.grid_forget()
        self.eps_text.grid_forget()
        self.min_sample_text.grid_forget()
        
        try:
            self.param_text.grid_forget()
        except:
            pass

        if self.cluster_result == "Hierarchical":
            print("Hierarchical Clustering Selected")  # Debug print
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Linkage Type:")
            self.linkage_box.grid(columnspan=2, row=2, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
            
        elif self.cluster_result == "Spectral":
            print("Spectral Clustering Selected")  # Debug print
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Affinity Type:")
            self.affinity_box.grid(columnspan=2, row=2, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
    
        elif self.cluster_result == "DBSCAN":
            print("DBSCAN Clustering Selected")  # Debug print
            # self.runeps() 
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="eps & min_sample value:")

            self.eps_text.grid(row=3, column=0, pady=0, padx=(5,0), sticky="w")
            self.min_sample_text.grid(row=2, column=0, pady=5, padx=(5,0), sticky="w")
        
            self.eps_box.grid(row=3, column=1, pady=0, padx=(0,5))
            self.min_sample_box.grid(row=2, column=1, pady=10, padx=(0,5))
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
            
        else:
            pass
            
            
        if self.cluster_result in ["DBSCAN", "Mean Shift", "Affinity Propagation"]:
            self.cluster_text.grid_forget()
            self.k_slider.grid_forget()
        else:
            self.cluster_text.grid(columnspan=2,row=4, column=0, pady=0, padx=5)
            self.k_slider.grid(columnspan=2,row=5, column=0, pady=(3,5), padx=5)



    def estimate_eps(self, k):
        # Estimate the optimal eps value for DBSCAN clustering
        from sklearn.neighbors import NearestNeighbors
        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        # Compute the k-nearest neighbors
        nearest_neighbors = NearestNeighbors(n_neighbors=k)
        nearest_neighbors.fit(self.X)
        distances, indices = nearest_neighbors.kneighbors(self.X)

        # Sort the distances (distances are ordered for each point)
        distances = np.sort(distances, axis=0)
        distances = distances[:, 1]  # Taking the distance to the k-th nearest neighbor

        # Plot the k-distance graph
        fig, ax = plt.subplots(figsize=(3, 2))
        ax.plot(distances)
        ax.set_title('k-Distance Graph')
        ax.set_xlabel('Points sorted by distance')
        ax.set_ylabel(f'{k}th Nearest Neighbor Distance')

        self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
        self.canvas1.get_tk_widget().pack(fill="both", expand=True)
        plt.show()

        elbow_point_index = 80  
        eps_value = distances[elbow_point_index]
        return eps_value

    def show_cluster(self):
        # Display cluster analysis
        self.X = None
    
        # Create a new tab for displaying the PCA cluster by samples
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("PCA Cluster by Samples")
    
        # Close all existing plots and clear widgets
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
        for widget in self.legend_frame.winfo_children():
            widget.destroy()
    
        # Get the number of clusters from the slider
        k = self.k_slider.get()
    
        # Get the scaled PCA values
        self.X = self.pca_df_scaled.values
    
        # Clear any existing cluster print label
        if hasattr(self, 'cluster_print') and self.cluster_print is not None:
            self.cluster_print.destroy()
            del self.cluster_print
    
        # Get the selected cluster method
        cluster_result = self.cluster_result
        self.cluster_print = ctk.CTkLabel(self.box_frame_sub, text=f"Cluster: {cluster_result}")
        self.cluster_print.grid(columnspan=2, row=6, column=0, pady=5, padx=5)
    
        # Set up the pipeline for different clustering algorithms
        if cluster_result == "K-mean":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
            ])
            
        elif cluster_result == "DBSCAN":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", DBSCAN(eps=self.eps_box.get(), min_samples=self.min_samples_box.get()))
            ])

        elif cluster_result == "Mean Shift":
            from sklearn.cluster import estimate_bandwidth
            bandwidth_value = estimate_bandwidth(self.X, quantile=0.2)
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", MeanShift(bandwidth=bandwidth_value))
            ])

        elif cluster_result == "Spectral":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", SpectralClustering(n_clusters=k, affinity='nearest_neighbors', random_state=0))
            ])

        elif cluster_result == "GMM":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", GaussianMixture(n_components=k, random_state=0))
            ])

        elif cluster_result == "Affinity Propagation":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AffinityPropagation(random_state=0))
            ])
 
        elif cluster_result == "Hierarchical":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AgglomerativeClustering(n_clusters=k, linkage='ward'))
            ])

        elif cluster_result == "BIRCH":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", Birch(n_clusters=k))
            ])

        # Fit the pipeline to the data
        self.pipe.fit(self.X)
    
        # Assign cluster labels to the dataframe
        if cluster_result == "GMM":
            self.df['cluster'] = self.pipe.named_steps["model"].predict(self.X)
        else:
            self.df['cluster'] = self.pipe.named_steps['model'].labels_
    
        # Create the KElbowVisualizer
        self.yellowbrick()
    
        # Plot PCA components with clusters
        fig, axes = plt.subplots(nrows=1, ncols=self.X.shape[1], figsize=(15, 8))
        cluster_colors = self.df['cluster'].values
        sort_clusters = self.cluster_sort_var.get() == 1
        graph_data = []
    
        for i in range(self.X.shape[1]):
            pc_name = f'PC{i+1}'
            data = self.pca_df_scaled.iloc[:, i]
    
            if sort_clusters:
                temp_df = pd.DataFrame({
                    'data': data,
                    'cluster': cluster_colors,
                    'sample id': self.cleaned_df['sample id']
                })
                sorted_df = temp_df.sort_values(by='cluster')
                sorted_data = sorted_df['data']
                sorted_cluster_colors = sorted_df['cluster'].values
                sorted_labels = sorted_df['sample id'].values
                self.plot_pca_barchart(sorted_data, pc_name, axes[i], sorted_cluster_colors, sorted_labels)
                sorted_df['PC'] = pc_name
                graph_data.append(sorted_df)
            else:
                temp_df = pd.DataFrame({
                    'data': data,
                    'cluster': cluster_colors,
                    'sample id': self.cleaned_df['sample id'],
                    'PC': pc_name
                })
                graph_data.append(temp_df)
                labels = self.cleaned_df['sample id'].values
                self.plot_pca_barchart(data, pc_name, axes[i], cluster_colors, labels)
    
        # Combine graph data for saving to Excel
        self.graph_data_df = pd.concat(graph_data)
    
        # Adjust layout and display the plot
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()
        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)
    
    def plot_pca_barchart(self, data, entry_name, ax, cluster_colors, sorted_labels):
        # Plot the PCA bar chart with clusters
        cluster_palette = plt.get_cmap('tab20', len(np.unique(cluster_colors)))
        colors = cluster_palette(cluster_colors)
        data.plot(kind='barh', color=colors, ax=ax, fontsize="xx-small")
        ax.set_title(entry_name)
        ax.grid(True, linestyle='--', linewidth=0.5)
        ax.set_yticks(range(len(data)))
        ax.set_yticklabels(sorted_labels, fontsize="xx-small")
        ax.set_ylabel('')
    
    def yellowbrick(self):
        # Create and display the KElbowVisualizer for clustering
        plt.rcParams.update({'font.size': 10})
        fig = plt.figure(figsize=(3, 2))
        ax = fig.add_subplot(111)
        ax.axis('off')
        try:
            model = self.pipe.named_steps['model']
            if 'n_clusters' in model.get_params():
                del model.get_params()['n_clusters']
    
            X = self.pca_df_scaled.values
            visualizer = KElbowVisualizer(model, k=(1, 9), timing=False, title="KElbowVisualizer").fit(X)
    
            visualizer.ax.set_yticklabels([])
            visualizer.finalize()
    
            self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
            toolbar = CustomToolbar(self.canvas1, self.legend_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            self.canvas1.get_tk_widget().pack(fill="both", expand=True)
    
            plt.close(fig)
        except:
            pass
    
    def save_clusters_to_excel(self):
        # Save the clusters to an Excel file
        if self.graph_data_df is None:
            print("No data to save. Ensure clusters are calculated and graph is generated.")
            return
    
        # Prompt user to save the file
        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return
    
        # Pivot the data for saving
        pivot_data = self.graph_data_df.pivot_table(index='sample id', columns='PC', values='data', sort=False)
        pivot_data.columns = [f'{col}' for col in pivot_data.columns]
    
        # Save the pivoted data to an Excel file
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            pivot_data.to_excel(excel_writer, index=True)
    
        # Load the workbook and worksheet
        wb = load_workbook(file_name)
        ws = wb.active
    
        # Apply cluster colors to the cells
        colormap = plt.get_cmap('Set2')
        cluster_colors = ['FF{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)) for r, g, b, _ in [colormap(i) for i in range(colormap.N)]]
        cluster_fill = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in cluster_colors]
    
        for i in range(len(self.graph_data_df)):
            sample_id = self.graph_data_df.iloc[i]['sample id']
            cluster = self.graph_data_df.iloc[i]['cluster']
            row = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
            for cell in row:
                if cell[0].value == sample_id:
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=cell[0].row, column=col).fill = cluster_fill[cluster]
    
        # Save the workbook
        wb.save(file_name)
        print(f"Clusters saved to {file_name}")
    

                





class threed_class:
    def __init__(self, shared_container, pca_df_scaled, df, cleaned_df, box_frame, box_frame_sub, on_button_click, apply_button, legend_frame, loadings):
        # Initialize the 3D class with necessary parameters
        self.apply_button = apply_button
        self.shared_container = shared_container
        self.on_button_click = on_button_click
        self.pca_df_scaled = pca_df_scaled
        self.df = df
        self.cleaned_df = cleaned_df

        self.loadings = loadings
        # Remove '_ppm' and '_pct' from the index of self.loadings
        self.loadings.index = self.loadings.index.str.replace('_ppm', '').str.replace('_pct', '')

        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.legend_frame = legend_frame
        self.var = IntVar()
        self.var1 = IntVar()
        self.var2 = IntVar()
        self.plot_3d()

    def plot_3d(self):
        # Create button for displaying 3D PCA biplot
        pil_image = Image.open("mineralAI_images/images_program/cube-3d-svgrepo-com.png")
        # resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))
        
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=1, column=2, sticky="w", pady=0, padx=5)

        # Bind button click to plot_3d_sub function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_3d_sub))

        ToolTip(self.image_button, msg="PCA 3D Biplot")

    def plot_3d_sub(self):
        # Create UI for selecting PCA components and options for 3D plot
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        pc_options = [f'PC{i+1}' for i in range(len(self.pca_df_scaled.T))]

        # Dropdowns for selecting PCA components
        ctk.CTkLabel(self.box_frame_sub, text="Select X-Axis:").pack(side="top", padx=5, pady=(5, 0))
        self.pc1_combo = ttk.Combobox(self.box_frame_sub, values=pc_options, state="readonly")
        self.pc1_combo.current(0)
        self.pc1_combo.pack(side="top", padx=5, pady=(3, 0))

        ctk.CTkLabel(self.box_frame_sub, text="Select Y-Axis:").pack(side="top", padx=5, pady=(5, 0))
        self.pc2_combo = ttk.Combobox(self.box_frame_sub, values=pc_options, state="readonly")
        self.pc2_combo.current(1)
        self.pc2_combo.pack(side="top", padx=5, pady=(3, 0))

        ctk.CTkLabel(self.box_frame_sub, text="Select Z-Axis:").pack(side="top", padx=5, pady=(5, 0))
        self.pc3_combo = ttk.Combobox(self.box_frame_sub, values=pc_options, state="readonly")
        self.pc3_combo.current(2)
        self.pc3_combo.pack(side="top", padx=5, pady=(3, 0))

        #select element for data point size
        ctk.CTkLabel(self.box_frame_sub, text="Element for Data Point Size:").pack(side="top", padx=5, pady=(5, 0))
        columns_with_na = ["N/A"] + self.df.columns.tolist()
        self.size_combo = ttk.Combobox(self.box_frame_sub, values=columns_with_na)
        self.size_combo.set("N/A")
        self.size_combo.pack(side="top", padx=5, pady=(3, 5))
        
        # Checkbox for displaying shapes
        if not ds.empty:
            self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Display shape", variable=self.var)
            self.checkbox.pack(side="top", anchor="w", padx=5, pady=(3, 0))

        # Checkbox for displaying trendlines
        self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Display Trendlines", variable=self.var2)
        self.checkbox.pack(side="top", anchor="w", padx=5, pady=(3, 0))

        # Checkbox for displaying legend
        self.checkbox1 = ctk.CTkCheckBox(self.box_frame_sub, text="Display Legend", variable=self.var1)
        self.checkbox1.pack(side="top", anchor="w", padx=5, pady=(3, 0))

        # Listbox for multiselect options
        self.multiselect_3d = tk.Listbox(self.box_frame_sub, selectmode=tk.MULTIPLE)
        self.multiselect_3d.pack(side="top", padx=5, pady=5)
        for names in self.df.columns.tolist():
            self.multiselect_3d.insert(tk.END, names)

        # Deselect all items in the listbox
        def deselect_all():
            self.multiselect_3d.selection_clear(0, tk.END)

        self.multiselect_3d.bind('<Command-d>', lambda event: deselect_all())

        # Apply button to show the 3D plot
        self.apply_plot_3d = ctk.CTkButton(self.box_frame_sub, text="apply", command=self.show_shape)
        self.apply_plot_3d.pack(side="top", padx=5, pady=5)
        
        # Display message for clicking data point if 'sample id' column exists
        if "sample id" in self.cleaned_df.columns:
            ctk.CTkLabel(self.box_frame_sub, text="Click Data Point for\n Sample ID").pack(side="top", padx=5, pady=(5, 0))

    def show_shape(self):
        # Show 3D PCA plot with selected components and options         
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("3D Biplot")

        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        pc1 = self.pc1_combo.get()
        pc2 = self.pc2_combo.get()
        pc3 = self.pc3_combo.get()

        fig = plt.figure(figsize=(11, 11))
        self.ax = fig.add_subplot(111, projection='3d')
        xdata = self.pca_df_scaled[pc1]
        ydata = self.pca_df_scaled[pc2]
        zdata = self.pca_df_scaled[pc3]

        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()

        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)

        # # Plot the data points in 3D space
        # if not dc.empty and not ds.empty:
        #     self.shapes = []
        #     for i in range(len(dc)):
        #         shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], c=dc['Color'][i], marker=ds["Shapes"][i])
        #         self.shapes.append(shape)
    
        #     self.dots = self.ax.scatter3D(xdata, ydata, zdata, c=dc['Color'])
        # else:
        #     self.dots = self.ax.scatter3D(xdata, ydata, zdata)

        
        # Plot the points with shapes and colors
        element_size = self.size_combo.get()
        self.shapes = []
        if element_size == "N/A":
            if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], c=dc['Color'][i], marker=ds["Shapes"][i])
                    self.shapes.append(shape)

                self.dots = self.ax.scatter3D(xdata, ydata, zdata, c=dc['Color'])

            elif "lithology" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], c=dc['Color'][i])
                    self.shapes.append(shape)
                    
                self.dots = self.ax.scatter3D(xdata, ydata, zdata, c=dc['Color'])                        

            elif "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], marker=ds["Shapes"][i])
                    self.shapes.append(shape)
                    
                self.dots = self.ax.scatter3D(xdata, ydata, zdata)                        

            else:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i])
                    self.shapes.append(shape)
                    
                self.dots = self.ax.scatter3D(xdata, ydata, zdata)
        else:
            # Try to get element_size column values for size mapping
            self.max_ele = self.df[element_size].max()
            self.min_ele = self.df[element_size].min()
            
            # Function to map element_size values to point sizes
            def map_size(x, new_min, new_max):
                old_min = self.min_ele
                old_max = self.max_ele
                return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
                
            # New logic for zn
            def size_for_zn(x):
                if x < 1000:
                    return 20
                elif 1000 <= x <= 9999:
                    return 70
                elif x > 10000:
                    return 110
                # return 20  # Default size if none of the conditions match
            
            # Apply size mapping logic
            if element_size == 'Zn_ppm':
                sizes = self.df[element_size].apply(size_for_zn)
            else:
                print("?")
                sizes = self.df[element_size].apply(lambda x: map_size(x, 20, 100))

            if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], c=dc['Color'][i], marker=ds["Shapes"][i], s=sizes[i]
                    )
                    self.shapes.append(shape)   
                self.dots = self.ax.scatter3D(xdata, ydata, zdata, c=dc['Color'], s=sizes)

            elif "lithology" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], c=dc['Color'][i], s=sizes[i]
                    )
                    self.shapes.append(shape)
                self.dots = self.ax.scatter3D(xdata, ydata, zdata, c=dc['Color'], s=sizes)   

            elif "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], marker=ds["Shapes"][i], s=sizes[i]
                    )
                    self.shapes.append(shape)
                self.dots = self.ax.scatter3D(xdata, ydata, zdata, s=sizes)
                    
            else:
                for i in range(len(self.df)):
                    shape = self.ax.scatter3D(xdata[i], ydata[i], zdata[i], s=sizes[i])
                    
                    self.shapes.append(shape)
                self.dots = self.ax.scatter3D(xdata, ydata, zdata, s=sizes)

        # Plot trendlines if selected
        def plot_trendlines_3d():
            trendline = self.var2.get() == 1
            if trendline and 'lithology' in self.cleaned_df.columns:
                from mpl_toolkits.mplot3d import Axes3D
                from sklearn.linear_model import LinearRegression
                lithology_groups = self.cleaned_df.groupby('lithology')
                for name, group in lithology_groups:
                    x = self.pca_df_scaled.loc[group.index, pc1]
                    y = self.pca_df_scaled.loc[group.index, pc2]
                    z = self.pca_df_scaled.loc[group.index, pc3]
                    
                    X = np.vstack((x, y)).T
                    Y = z
                    
                    model = LinearRegression().fit(X, Y)
                    
                    trend_x = np.linspace(x.min(), x.max(), 100)
                    trend_y = np.linspace(y.min(), y.max(), 100)
                    trend_z = model.predict(np.vstack((trend_x, trend_y)).T)
                    color = color_map.get(name)
                    
                    self.ax.plot(trend_x, trend_y, trend_z, linestyle='--', label=f"Trendline {name}", color=color)
        
        plot_trendlines_3d()

        # Display sample IDs on click if selected
        if "sample id" in self.cleaned_df.columns:
            description = self.cleaned_df["sample id"]
            self.annotations = {}
            self.selected_points = set()
            
            from mpl_toolkits.mplot3d import proj3d
            
            def add_annotation(ind, source):
                if source == self.dots:
                    index = ind["ind"][0]
                    pos = source.get_offsets()[index]
                    text = f"{description.iloc[index]}"
                else:
                    index = ind
                    x, y, _ = proj3d.proj_transform(self.pca_df_scaled.iloc[index][pc1], self.pca_df_scaled.iloc[index][pc2], self.pca_df_scaled.iloc[index][pc3], self.ax.get_proj())
                    pos = (x, y)
                    text = f"{description.iloc[index]}"

                offsets = [(20, 20), (-20, 20), (20, -20), (-20, -20), (40, 40), (-40, 40), (40, -40), (-40, -40)]
                
                for offset in offsets:
                    annot = self.ax.annotate(text, xy=pos, xytext=offset,
                                        textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->",),
                                        zorder=100)
                    fig.canvas.draw()
                    bbox = annot.get_window_extent()
                    overlap = False
                    for existing_annot in self.annotations.values():
                        existing_bbox = existing_annot.get_window_extent()
                        if bbox.overlaps(existing_bbox):
                            overlap = True
                            annot.remove()
                            break
                    if not overlap:
                        annot.set_visible(True)
                        self.annotations[index] = annot
                        self.selected_points.add(index)
                        fig.canvas.draw_idle()
                        return 
            
                annot = self.ax.annotate(text, xy=pos, xytext=offsets[0],
                                    textcoords="offset points",
                                    bbox=dict(boxstyle="round", fc="w"),
                                    arrowprops=dict(arrowstyle="->"),
                                    zorder=100)
                annot.set_visible(True)
                self.annotations[index] = annot
                self.selected_points.add(index)
                fig.canvas.draw_idle()
            
            def remove_annotation(index):
                annot = self.annotations.pop(index, None)
                if annot:
                    annot.remove()
                self.selected_points.discard(index)
                fig.canvas.draw_idle()
            
            def update_annotations(event):
                for index, annot in self.annotations.items():
                    if index in self.selected_points:
                        x, y, _ = proj3d.proj_transform(self.pca_df_scaled.iloc[index][pc1], self.pca_df_scaled.iloc[index][pc2], self.pca_df_scaled.iloc[index][pc3], self.ax.get_proj())
                        annot.xy = (x, y)
                fig.canvas.draw_idle()
            
            def on_click(event):
                if event.inaxes == self.ax and event.button == MouseButton.LEFT:
                    contains_dots, ind_dots = self.dots.contains(event)
                    contains_shapes = False
                    shape_index = None
                    for i, shape in enumerate(self.shapes):
                        contains_shapes, ind_shapes = shape.contains(event)
                        if contains_shapes:
                            shape_index = i
                            break
                    
                    if contains_dots:
                        index = ind_dots["ind"][0]
                        if index in self.selected_points:
                            remove_annotation(index)
                        else:
                            add_annotation(ind_dots, self.dots)
                    elif contains_shapes:
                        if shape_index in self.selected_points:
                            remove_annotation(shape_index)
                        else:
                            add_annotation(shape_index, shape)
            
            fig.canvas.mpl_connect("button_press_event", on_click)
            fig.canvas.mpl_connect("motion_notify_event", update_annotations)

        self.setups(pc1, pc2, pc3)
        self.shape()
        self.legend()

    def shape(self):
        # Toggle visibility of shapes in the plot
        try:
            shape = self.var.get() == 1
            if shape:
                for shape in self.shapes:
                    shape.set_visible(True)
                self.dots.set_visible(False)
            else:
                self.dots.set_visible(True)
                for shape in self.shapes:
                    shape.set_visible(False)
        except:
            self.dots.set_visible(True)
            
    def legend(self):
        # Display legend for the plot
        if not dc.empty and not ds.empty:
            
            unique_lithologies = self.cleaned_df['lithology'].unique()
            unique_shapes = self.cleaned_df['rock unit'].unique()
    
            handle1 = []
            label1 = []
            for lithology in unique_lithologies:
                color = color_map[lithology]
                l1 = self.ax.scatter([], [], c=color, label=lithology)
                handle1.append(l1)
                label1.append(lithology)
            
            handle2 = []
            label2 = []
            for shape in unique_shapes:
                marker = color_map1[shape]
                l2 = self.ax.scatter([], [], c="black", marker=marker, label=shape)
                handle2.append(l2)
                label2.append(shape)
    
            total_entries = len(label1) + len(label2)
            height_per_entry = 0.225
            fig_height = total_entries * height_per_entry
    
            figx = plt.figure(figsize=(2.5, fig_height))
            axx = figx.add_subplot(111)
            axx.axis('off')
    
            self.legend1 = plt.legend(handle1, label1, bbox_to_anchor=(0.9, 1.15), title='Lithology', fontsize=10, labelspacing=0.3)
            self.legend2 = plt.legend(handle2, label2, bbox_to_anchor=(0.9, 0.35), title='Rock Type', fontsize=10, labelspacing=0.3)

            axx.add_artist(self.legend1)
            axx.add_artist(self.legend2)
            self.canvas1 = FigureCanvasTkAgg(figx, master=self.legend_frame)
            
            toolbar = CustomToolbar(self.canvas1, self.legend_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)

            self.canvas1.get_tk_widget().pack(fill="both", expand=True)
            
            self.legend1.set_visible(False)
            self.legend2.set_visible(False)
    
            Legend = self.var1.get() == 1
            if Legend:
                self.legend1.set_visible(True)
                self.legend2.set_visible(True)
            else:
                self.legend1.set_visible(False)
                self.legend2.set_visible(False)
        else:
            pass

    def setups(self, pc1, pc2, pc3):
        # Set up 3D biplot with loadings
        show_names = self.loadings.index.tolist()
        indx = self.loadings.index.get_indexer(show_names)
        scale = 1

        xs = scale * self.loadings[pc1]
        ys = scale * self.loadings[pc2]
        zs = scale * self.loadings[pc3]

        plt.title(f'3D Biplot')

        self.ax.set_xlabel(pc1)
        self.ax.set_ylabel(pc2)
        self.ax.set_zlabel(pc3)
        x_arr = np.zeros(len(self.loadings[pc1]))
        y_arr = z_arr = x_arr

        all_points = np.concatenate((xs, ys, zs))
        min_value = np.min(all_points)
        max_value = np.max(all_points)
        axis_range = [min_value, max_value]

        self.ax.plot([x_arr.any() + axis_range[0], x_arr.any() + axis_range[1]], [y_arr.any(), y_arr.any()], [z_arr.any(), z_arr.any()], color='k', linestyle='--')
        self.ax.plot([x_arr.any(), x_arr.any()], [y_arr.any() + axis_range[0], y_arr.any() + axis_range[1]], [z_arr.any(), z_arr.any()], color='k', linestyle='--')
        self.ax.plot([x_arr.any(), x_arr.any()], [y_arr.any(), y_arr.any()], [z_arr.any() + axis_range[0], z_arr.any() + axis_range[1]], color='k', linestyle='--')

        self.arrow_list = []
        for i, name in enumerate(self.df.columns):
            ip = self.df.columns.get_loc(name)
            arrow = self.ax.quiver(x_arr[ip], y_arr[ip], z_arr[ip], xs[ip], ys[ip], zs[ip], color='r', arrow_length_ratio=0.1, linewidth=0.5)
            self.arrow_list.append(arrow)

        self.namelist = []
        for i, names in enumerate(show_names):
            ip = indx[i]
            n = self.ax.text(xs[ip] + 0.02, ys[ip] + 0.02, zs[ip] + 0.02, names, fontsize='small')
            self.namelist.append(n)

        selected_indices = self.multiselect_3d.curselection()
        self.selected_items = [self.multiselect_3d.get(i) for i in selected_indices]
        self.findname_specific()
        self.ax.callbacks.connect('ylim_changed', lambda ax: self.update_quiver_3d(arrow_list_3d))
        self.ax.callbacks.connect('xlim_changed', lambda ax: self.update_quiver_3d(arrow_list_3d))
        self.ax.callbacks.connect('zlim_changed', lambda ax: self.update_quiver_3d(arrow_list_3d))
        
    # Add this function to handle quiver arrow updates when the axes change for 3D plots
    def update_quiver_3d(self, arrow_list_3d):
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        zlim = self.ax.get_zlim()
        
        arrow_scale = max(abs(xlim[1] - xlim[0]), abs(ylim[1] - ylim[0]), abs(zlim[1] - zlim[0])) / 0.5
        
        for arrow in arrow_list_3d:
            arrow.set_segments([arrow.U, arrow.V, arrow.W], scale=arrow_scale)


    def findname_specific(self):
        # Highlight selected names and arrows in the plot
        options = self.df.columns.tolist()
        for i in range(len(options[:])):
            if options[i] in self.selected_items:
                self.namelist[i].set_visible(True)
                self.arrow_list[i].set_visible(True)
            else:
                self.namelist[i].set_visible(False)
                self.arrow_list[i].set_visible(False)






    


class twod_class:
    def __init__(self, shared_container, pca_df_scaled, df, cleaned_df, box_frame, box_frame_sub, on_button_click, apply_button, legend_frame, loadings):
        # Initialize the 2D class with required parameters
        self.shared_container = shared_container
        self.pca_df_scaled = pca_df_scaled
        self.df = df
        self.cleaned_df = cleaned_df

        self.loadings = loadings
        # Remove '_ppm' and '_pct' from the index of self.loadings
        self.loadings.index = self.loadings.index.str.replace('_ppm', '').str.replace('_pct', '')


        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.legend_frame = legend_frame

        self.var = IntVar()
        self.var1 = IntVar()
        self.var2 = IntVar()
        self.on_button_click = on_button_click
        self.apply_button = apply_button
        
        self.plot_2d()

    def plot_2d(self):
        # Load and display the 2D plot icon button
        pil_image = Image.open("mineralAI_images/images_program/chart-scatterplot-svgrepo-com.png")
        resized_image = pil_image.resize((32, 32), Image.LANCZOS)
        self.icon_image = ImageTk.PhotoImage(resized_image)
        
        print(pil_image)
        print(self.icon_image)
        
        # Create a label with the image button
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=1, column=3, sticky="w", pady=0, padx=5)
        
        # Bind the button click to the sub plot function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_2d_sub))

        ToolTip(self.image_button, msg="PCA 2D Biplot")

    def plot_2d_sub(self):
        # Clear any existing widgets in the sub-frame
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Options for selecting principal components
        pc_options = [f'PC{i+1}' for i in range(len(self.pca_df_scaled.T))]
        
        # Dropdown for selecting the first principal component
        ctk.CTkLabel(self.box_frame_sub, text="Select PC1:").pack(side="top", padx=5, pady=(5, 0))
        self.pc1_combo = ttk.Combobox(self.box_frame_sub, values=pc_options, state="readonly")
        self.pc1_combo.current(0)
        self.pc1_combo.pack(side="top", padx=5, pady=(3, 0))
        
        # Dropdown for selecting the second principal component
        ctk.CTkLabel(self.box_frame_sub, text="Select PC2:").pack(side="top", padx=5, pady=(5, 0))
        self.pc2_combo = ttk.Combobox(self.box_frame_sub, values=pc_options, state="readonly")
        self.pc2_combo.current(1)
        self.pc2_combo.pack(side="top", padx=5, pady=(3, 0))
        
        #select element for data point size
        ctk.CTkLabel(self.box_frame_sub, text="Element for Data Point Size:").pack(side="top", padx=5, pady=(5, 0))
        columns_with_na = ["N/A"] + self.df.columns.tolist()
        self.size_combo = ttk.Combobox(self.box_frame_sub, values=columns_with_na, state="readonly")
        self.size_combo.set("N/A")
        self.size_combo.pack(side="top", padx=5, pady=(3, 5))
        
        # Checkbox for displaying shapes
        if not ds.empty:
            self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Display shape", variable=self.var)
            self.checkbox.pack(side="top", anchor="w", padx=5, pady=(3, 0))

        # Checkbox for displaying trendlines
        self.checkbox2 = ctk.CTkCheckBox(self.box_frame_sub, text="Display Trendlines", variable=self.var2)
        self.checkbox2.pack(side="top", anchor="w", padx=5, pady=(3, 0))
        
        # Checkbox for displaying legend
        self.checkbox1 = ctk.CTkCheckBox(self.box_frame_sub, text="Display Legend", variable=self.var1)
        self.checkbox1.pack(side="top", anchor="w", padx=5, pady=(3, 0))

        # Multiselect listbox for selecting columns
        self.multiselect_2d = tk.Listbox(self.box_frame_sub, selectmode=tk.MULTIPLE)
        self.multiselect_2d.pack(side="top", padx=5, pady=5)
        for names in self.df.columns.tolist():
            self.multiselect_2d.insert(tk.END, names)
            
        # Function to deselect all items in the listbox
        def deselect_all():
            self.multiselect_2d.selection_clear(0, tk.END)

        # Bind Command-d (deselect all) to the listbox
        self.multiselect_2d.bind('<Command-d>', lambda event: deselect_all())

        # Apply button to show the 2D plot
        self.apply_plot_2d = ctk.CTkButton(self.box_frame_sub, text="apply", command=self.show_shape)
        self.apply_plot_2d.pack(side="top", padx=5, pady=5)
        
        # Display message for clicking data point if 'sample id' column exists
        if "sample id" in self.cleaned_df.columns:
            ctk.CTkLabel(self.box_frame_sub, text="Click Data Point for \n Sample ID").pack(side="top", padx=5, pady=(5, 0))

    def show_shape(self):            
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("2D Biplot")

        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
        for widget in self.legend_frame.winfo_children():
            widget.destroy()
    
        pc1 = self.pc1_combo.get()
        pc2 = self.pc2_combo.get()
    
        xdata = self.pca_df_scaled[pc1]
        ydata = self.pca_df_scaled[pc2]
        
        fig = plt.figure(figsize=(10, 10))
        
        # Create a subplot grid if 'lithology' column exists
        if 'lithology' in self.cleaned_df.columns:
            gs = fig.add_gridspec(5, 5, hspace=0.4, wspace=0.4)
            self.ax = fig.add_subplot(gs[1:5, 0:4])
            ax_box_x = fig.add_subplot(gs[0, 0:4], sharex=self.ax)
            ax_box_y = fig.add_subplot(gs[1:5, 4], sharey=self.ax)
        else:
            self.ax = fig.add_subplot(111)
        
        # # Plot data points with shapes and colors if available
        # if not dc.empty and not ds.empty:
        #     self.shapes = []
        #     for i in range(len(self.df)):
        #         shape = self.ax.scatter(xdata[i], ydata[i], c=dc['Color'][i], marker=ds["Shapes"][i])
        #         self.shapes.append(shape)
                
        #     self.dots = self.ax.scatter(xdata, ydata, c=dc['Color'])
        # else:
        #     self.dots = self.ax.scatter(xdata, ydata)

        # Plot the points with shapes and colors
        element_size = self.size_combo.get()
        self.shapes = []
        if element_size == "N/A":
            if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], c=dc['Color'][i], marker=ds["Shapes"][i])
                    self.shapes.append(shape)

                self.dots = self.ax.scatter(xdata, ydata, c=dc['Color'])

            elif "lithology" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], c=dc['Color'][i])
                    self.shapes.append(shape)
                    
                self.dots = self.ax.scatter(xdata, ydata, c=dc['Color'])                        

            elif "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], marker=ds["Shapes"][i])
                    self.shapes.append(shape)
                    
                self.dots = self.ax.scatter(xdata, ydata)                        

            else:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i])
                    self.shapes.append(shape)
                    
                self.dots = self.ax.scatter(xdata, ydata)
        else:
            # Try to get element_size column values for size mapping
            self.max_ele = self.df[element_size].max()
            self.min_ele = self.df[element_size].min()
            
            # Function to map element_size values to point sizes
            def map_size(x, new_min, new_max):
                old_min = self.min_ele
                old_max = self.max_ele
                return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
                
            # New logic for zn
            def size_for_zn(x):
                if x < 1000:
                    return 20
                elif 1000 <= x <= 9999:
                    return 70
                elif x > 10000:
                    return 110
                else:
                    raise Exception ("size zn")
                # return 20  # Default size if none of the conditions match
            
            # Apply size mapping logic
            if element_size == 'Zn_ppm':
                sizes = self.df[element_size].apply(size_for_zn)
            else:
                print("?")
                sizes = self.df[element_size].apply(lambda x: map_size(x, 20, 100))

            if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], c=dc['Color'][i], marker=ds["Shapes"][i], s=sizes[i]
                    )
                    self.shapes.append(shape)   
                self.dots = self.ax.scatter(xdata, ydata, c=dc['Color'], s=sizes)

            elif "lithology" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], c=dc['Color'][i], s=sizes[i]
                    )
                    self.shapes.append(shape)
                self.dots = self.ax.scatter(xdata, ydata, c=dc['Color'], s=sizes)   

            elif "rock unit" in self.cleaned_df.columns:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], marker=ds["Shapes"][i], s=sizes[i]
                    )
                    self.shapes.append(shape)
                self.dots = self.ax.scatter(xdata, ydata, s=sizes)
                    
            else:
                for i in range(len(self.df)):
                    shape = self.ax.scatter(xdata[i], ydata[i], s=sizes[i])
                    
                    self.shapes.append(shape)
                self.dots = self.ax.scatter(xdata, ydata, s=sizes)

        def plot_trendlines():
            # Plot trendlines for lithology groups
            trendline = self.var2.get() == 1
            if trendline and 'lithology' in self.cleaned_df.columns:
                import statsmodels.api as sm
                lithology_groups = self.cleaned_df.groupby('lithology')
                for name, group in lithology_groups:
                    x = self.pca_df_scaled.loc[group.index, pc1]
                    y = self.pca_df_scaled.loc[group.index, pc2]
                    
                    color = color_map.get(name)
                    
                    sns.regplot(x=x, y=y, ax=self.ax, scatter=False, label=f"Trendline {name}", color=color, ci=25, line_kws={'linewidth': 1})
        
        # Call the plot_trendlines function if enabled
        plot_trendlines()

        if "sample id" in self.cleaned_df.columns:
            # Annotate sample IDs on click
            description = self.cleaned_df["sample id"]
            self.annotations = {}
            self.selected_points = set()
            
            def add_annotation(ind, source):
                if source == self.dots:
                    index = ind["ind"][0]
                    pos = source.get_offsets()[index]
                    text = f"{description.iloc[index]}"
                else:
                    index = ind
                    pos = (self.pca_df_scaled.iloc[index][pc1], self.pca_df_scaled.iloc[index][pc2])
                    text = f"{description.iloc[index]}"
            
                offsets = [(20, 20), (-20, 20), (20, -20), (-20, -20), (40, 40), (-40, 40), (40, -40), (-40, -40)]
                
                # Try different annotation positions to avoid overlap
                for offset in offsets:
                    annot = self.ax.annotate(text, xy=pos, xytext=offset,
                                        textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"),
                                        zorder=100)
                    fig.canvas.draw()
                    bbox = annot.get_window_extent()
                    overlap = False
                    for existing_annot in self.annotations.values():
                        existing_bbox = existing_annot.get_window_extent()
                        if bbox.overlaps(existing_bbox):
                            overlap = True
                            annot.remove()
                            break
                    if not overlap:
                        annot.set_visible(True)
                        self.annotations[index] = annot
                        self.selected_points.add(index)
                        fig.canvas.draw_idle()
                        return 
            
                # If all positions overlap, place annotation in the default position
                annot = self.ax.annotate(text, xy=pos, xytext=offsets[0],
                                    textcoords="offset points",
                                    bbox=dict(boxstyle="round", fc="w"),
                                    arrowprops=dict(arrowstyle="->"),
                                    zorder=100)
                annot.set_visible(True)
                self.annotations[index] = annot
                self.selected_points.add(index)
                fig.canvas.draw_idle()
            
            def remove_annotation(index):
                # Remove annotation from the plot
                annot = self.annotations.pop(index, None)
                if annot:
                    annot.remove()
                self.selected_points.discard(index)
                fig.canvas.draw_idle()
            
            def update_annotations(event):
                # Update annotation positions on mouse move
                for index, annot in self.annotations.items():
                    if index in self.selected_points:
                        annot.xy = (self.pca_df_scaled.iloc[index][pc1], self.pca_df_scaled.iloc[index][pc2])
                fig.canvas.draw_idle()
            
            def on_click(event):
                # Handle mouse click to add or remove annotations
                if event.inaxes == self.ax and event.button == MouseButton.LEFT:
                    contains_dots, ind_dots = self.dots.contains(event)
                    contains_shapes = False
                    shape_index = None
                    for i, shape in enumerate(self.shapes):
                        contains_shapes, ind_shapes = shape.contains(event)
                        if contains_shapes:
                            shape_index = i
                            break
                    
                    if contains_dots:
                        index = ind_dots["ind"][0]
                        if index in self.selected_points:
                            remove_annotation(index)
                        else:
                            add_annotation(ind_dots, self.dots)
                    elif contains_shapes:
                        if shape_index in self.selected_points:
                            remove_annotation(shape_index)
                        else:
                            add_annotation(shape_index, shape)
            
            # Connect event handlers for clicking and moving the mouse
            fig.canvas.mpl_connect("button_press_event", on_click)
            fig.canvas.mpl_connect("motion_notify_event", update_annotations)

        if 'lithology' in self.cleaned_df.columns:
            # Create and display box plots for lithology
            filtered_df = self.cleaned_df.copy()
            
            xdata_filtered = self.pca_df_scaled.loc[filtered_df.index, pc1]
            ydata_filtered = self.pca_df_scaled.loc[filtered_df.index, pc2]
    
            filtered_colors = filtered_df['lithology'].map(color_map)
            
            unique_lithologies = filtered_df['lithology'].nunique()
            palette = [color_map[lithology] for lithology in filtered_df['lithology'].unique()]
            
            sns.boxplot(x=xdata_filtered, y=filtered_df['lithology'], ax=ax_box_x, palette=palette, orient='h')
            sns.boxplot(y=ydata_filtered, x=filtered_df['lithology'], ax=ax_box_y, palette=palette, orient='v')
    
            # Customize tick parameters for the box plots
            ax_box_x.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)
            ax_box_x.tick_params(axis='y', labelsize=8, left=False, right=True, labelleft=False, labelright=True)
    
            ax_box_y.tick_params(axis='y', which='both', left=False, right=False, labelleft=False)
            ax_box_y.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)
    
            ax_box_x.set_ylabel(None)
            ax_box_x.set_xlabel(None)
            
            ax_box_y.set_ylabel(None)
            ax_box_y.set_xlabel(None)

            self.ax.axvline(x=0, color='black', linewidth=0.5)
            self.ax.axhline(y=0, color='black', linewidth=0.5)

        # Create and display the plot canvas and toolbar
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()

        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)
        
        # Setup additional plot features
        self.setups(pc1, pc2)
        self.shape()
        self.legend()

    def shape(self):
        # Show or hide shapes based on checkbox selection
        try:
            shape = self.var.get() == 1
            if shape:
                for shape in self.shapes:
                    shape.set_visible(True)
                self.dots.set_visible(False)
            else:
                self.dots.set_visible(True)
                for shape in self.shapes:
                    shape.set_visible(False)
        except:
            self.dots.set_visible(True)
            
    def legend(self):  
        # Create and display legends for lithology and rock type
        if not dc.empty and not ds.empty:
            
            unique_lithologies = self.cleaned_df['lithology'].unique()
            unique_shapes = self.cleaned_df['rock unit'].unique()
    
            handle1 = []
            label1 = []
            for lithology in unique_lithologies:
                color = color_map[lithology]
                l1 = self.ax.scatter([], [], c=color, label=lithology)
                handle1.append(l1)
                label1.append(lithology)
            
            handle2 = []
            label2 = []
            for shape in unique_shapes:
                marker = color_map1[shape]
                l2 = self.ax.scatter([], [], c="black", marker=marker, label=shape)
                handle2.append(l2)
                label2.append(shape)
            
            total_entries = len(label1) + len(label2)
            height_per_entry = 0.225
            fig_height = total_entries * height_per_entry
                
            figx = plt.figure(figsize=(2.5, fig_height))
            axx = figx.add_subplot(111)
            axx.axis('off')
        
            self.legend1 = plt.legend(handle1, label1, bbox_to_anchor=(0.9, 1.15), title='Lithology', fontsize=10, labelspacing=0.3)
            self.legend2 = plt.legend(handle2, label2, bbox_to_anchor=(0.9, 0.35), title='Rock Type', fontsize=10, labelspacing=0.3)
            axx.add_artist(self.legend1)
            axx.add_artist(self.legend2)
            self.canvas1 = FigureCanvasTkAgg(figx, master=self.legend_frame)
            toolbar = CustomToolbar(self.canvas1, self.legend_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)

            self.canvas1.get_tk_widget().pack(fill="both", expand=True)
    
            self.legend1.set_visible(False)
            self.legend2.set_visible(False)
            self.legenda()
        else:
            pass
            
    def legenda(self):
        # Show or hide legends based on checkbox selection
        Legend = self.var1.get() == 1
        if Legend:
            self.legend1.set_visible(True)
            self.legend2.set_visible(True)
        else:
            self.legend1.set_visible(False)
            self.legend2.set_visible(False)

    def setups(self, pc1, pc2):
        # Setup for 2D biplot with arrows and labels for loadings
        show_names = self.loadings.index.tolist()
        indx = self.loadings.index.get_indexer(show_names)
        
        pc1_index = int(pc1[2:]) - 1
        pc2_index = int(pc2[2:]) - 1
        self.xs = self.loadings[pc1]
        self.ys = self.loadings[pc2]

        self.arrow_list_2d = []
        
        self.x_arr = np.zeros(len(self.loadings[pc1]))
        self.y_arr = self.x_arr

        max_x = max(abs(self.xs))
        max_y = max(abs(self.ys))
        arrow_scale = max(max_x, max_y) / 0.5
        
        for i, name in enumerate(self.df.columns):
            ip = self.df.columns.get_loc(name)
            arrow_2d = self.ax.quiver(self.x_arr[ip], self.y_arr[ip], self.xs[ip], self.ys[ip], color='r', scale=arrow_scale, width=0.002)
            self.arrow_list_2d.append(arrow_2d)
        
        namelist_2d = []
        for i, names in enumerate(show_names):
            ip = indx[i]
            n = self.ax.text((1 / arrow_scale) * self.xs[ip], (1 / arrow_scale) * self.ys[ip], names, fontsize='small')
            namelist_2d.append(n)

        selected_indices = self.multiselect_2d.curselection()
        self.selected_items = [self.multiselect_2d.get(i) for i in selected_indices]
        self.findname_specific_2d(namelist_2d)
        
        self.ax.set_xlabel(pc1)
        self.ax.set_ylabel(pc2)
        self.ax.set_title(f"{pc1} vs {pc2}")
        self.prev_xlim = self.ax.get_xlim()
        self.prev_ylim = self.ax.get_ylim()
        
        # Connect to 'draw_event' to track canvas redraws
        self.ax.figure.canvas.mpl_connect('draw_event', self.on_draw_event)
        
    def on_draw_event(self, event):
        # Get the current axis limits
        curr_xlim = self.ax.get_xlim()
        curr_ylim = self.ax.get_ylim()
    
        # Check if the axis limits have changed
        if curr_xlim != self.prev_xlim or curr_ylim != self.prev_ylim:
            # Update the quiver arrows if the limits have changed
            self.update_quiver()
    
            # Save the current limits as the previous limits
            self.prev_xlim = curr_xlim
            self.prev_ylim = curr_ylim
    
    def update_quiver(self):
        print("lol")
        # Store the current visibility state of the arrows
        arrow_visibility = [arrow.get_visible() for arrow in self.arrow_list_2d]
    
        # Clear old quivers
        for arrow in self.arrow_list_2d:
            arrow.remove()
    
        # Get the current axis limits
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
    
        # Calculate the appropriate scale for arrows
        x_range = xlim[1] - xlim[0]
        y_range = ylim[1] - ylim[0]
    
        arrow_scale = max(x_range, y_range)
    
        # Recreate the quivers with updated scale
        self.arrow_list_2d.clear()
        for i, name in enumerate(self.df.columns):
            ip = self.df.columns.get_loc(name)
            arrow_2d = self.ax.quiver(self.x_arr[ip], self.y_arr[ip], self.xs[ip], self.ys[ip], color='r', scale=arrow_scale, width=0.002)
            arrow_2d.set_visible(arrow_visibility[i])
            self.arrow_list_2d.append(arrow_2d)
    
        # Redraw the canvas to reflect changes
        self.ax.figure.canvas.draw_idle()


    def findname_specific_2d(self, namelist_2d):
        # Show or hide specific names and arrows based on user selection
        options = self.df.columns.tolist()
        for i in range(len(options[:])):
            if options[i] in self.selected_items:
                namelist_2d[i].set_visible(True)
                self.arrow_list_2d[i].set_visible(True)
            else:
                namelist_2d[i].set_visible(False)
                self.arrow_list_2d[i].set_visible(False)









class Cluster2DPlotClass:
    def __init__(self, shared_container, cluster, df, cleaned_df, box_frame, box_frame_sub, on_button_click, legend_frame):
        # Initialize the 2D cluster plot class with required parameters
        self.df = df.dropna()
        self.shared_container = shared_container
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub

        self.on_button_click = on_button_click
        self.legend_frame = legend_frame
        self.cleaned_df = cleaned_df
        self.var = IntVar()
        self.var1 = IntVar()
        self.cluster_result = cluster
        # Bind the cluster change event to update the plot
        # self.kmean_instance.scaler_combo.bind("<<ClusterChanged>>", self.on_cluster_change)
    
        self.plot_2d_cluster()

    # def on_cluster_change(self):
    #     # Handle cluster change event and update the plot
    #     self.cluster_result = self.kmean_instance.get_cluster()
    #     self.plot_2d_cluster_sub()

    def plot_2d_cluster(self):
        # Load and display the 2D cluster plot icon button
        pil_image = Image.open("mineralAI_images/images_program/2d.png")
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        print(pil_image)
        print(self.icon_image)
        
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text = "", width=20)
        self.image_button.grid(row=3, column=3, sticky="w", pady=0, padx=5)
        
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_2d_cluster_sub0))

        ToolTip(self.image_button, msg="PCA 2D Biplot with Clusters")
        
        # self.plot_2d_cluster_button = ctk.CTkButton(self.box_frame, text="2D", command=lambda: self.on_button_click(self.plot_2d_cluster_button, self.plot_2d_cluster_sub), width=20)
        # self.plot_2d_cluster_button.grid(row=3, column=1, sticky="w", pady=(5,0), padx=0)
    
    def plot_2d_cluster_sub0(self):
        # clear column cluster if exist
        if 'cluster' in self.df.columns:
            self.df.drop(columns=['cluster'], inplace=True)
        else:
            pass

        # Display options for selecting axes and cluster settings
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()
            
        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 1:").grid(columnspan=2, row=0, column=0, pady=(5,0), padx=5)
        self.axis1_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis1_combo.current(0)
        self.axis1_combo.grid(columnspan=2, row=1, column=0, pady=(3,0), padx=5)
        
        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 2:").grid(columnspan=2, row=2, column=0, pady=(5,0), padx=5)
        self.axis2_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis2_combo.current(1)
        self.axis2_combo.grid(columnspan=2, row=3, column=0, pady=(3,0), padx=5)

        ctk.CTkLabel(self.box_frame_sub, text="Element for Data Point Size").grid(columnspan=2, row=4, column=0, pady=(5,0), padx=5)
        self.size_combo1 = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), name="size_combo1", state="readonly")
        self.size_combo1.set("N/A")
        self.size_combo1.grid(columnspan=2, row=5, column=0, pady=(3,5), padx=5)

        if not ds.empty:
            self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Display shape", variable=self.var)
            self.checkbox.grid(columnspan=2, row=6, column=0, sticky=W, pady=(5,0), padx=5)

        self.legend_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Show Legend", variable=self.var1)
        self.legend_checkbox.grid(columnspan=2, row=7, column=0, sticky=W, pady=(5,0), padx=5)
            
        self.cluster_text = ctk.CTkLabel(self.box_frame_sub, text="Number of Clusters:")
        self.k_slider = tk.Scale(self.box_frame_sub, from_=2, to=10, orient=tk.HORIZONTAL)

        self.apply_plot_2d_cluster = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.show_cluster)
        self.apply_plot_2d_cluster.grid(columnspan=2, row=13, column=0, pady=(5,0), padx=5)

        self.save_button = ctk.CTkButton(self.box_frame_sub, text="Save Clusters to Excel", command=self.save_clusters_to_excel)
        self.save_button.pack(side="bottom", padx=5, pady=5)

        # Define the values for the Spinbox
        linkage_box_values = ('ward', 'complete', 'average', 'single')
        self.linkage_box = ttk.Combobox(self.box_frame_sub, values=linkage_box_values, state="readonly")
        self.linkage_box.current(0)

        affinity_box_values = ('nearest_neighbors', 'rbf', 'precomputed', 'precomputed_nearest_neighbors')
        self.affinity_box = ttk.Combobox(self.box_frame_sub, values=affinity_box_values, state="readonly")
        self.affinity_box.current(0)
                
        self.min_sample_box = CTkSpinbox(self.box_frame_sub, step_size=1, min_value=2, max_value=100, width = 110) #, command=self.runeps)
        self.eps_box = CTkSpinbox(self.box_frame_sub, step_size=0.1, min_value=0.1, max_value=100, width = 110)
        
        self.eps_text = ctk.CTkLabel(self.box_frame_sub, text="eps")
        self.min_sample_text = ctk.CTkLabel(self.box_frame_sub, text="min_sample")

        self.plot_2d_cluster_sub(self.cluster_result)
        
    def plot_2d_cluster_sub(self, cluster):
        self.box_frame_sub.grid_propagate(True)
        self.cluster_result = cluster

        self.linkage_box.grid_forget()
        self.affinity_box.grid_forget()
        self.min_sample_box.grid_forget()
        self.eps_box.grid_forget()
        self.eps_text.grid_forget()
        self.min_sample_text.grid_forget()
        
        try:
            self.param_text.grid_forget()
        except:
            pass

        
        if self.cluster_result == "Hierarchical":
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Linkage Type:")
            self.linkage_box.grid(columnspan=2, row=9, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=8, column=0, pady=(5,0), padx=5)
            
        elif self.cluster_result == "Spectral":
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Affinity Type:")
            self.affinity_box.grid(columnspan=2, row=9, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=8, column=0, pady=(5,0), padx=5)

        elif self.cluster_result == "DBSCAN":
            # self.runeps() 
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="eps & min_sample value:")
            
            self.eps_text.grid(row=10, column=0, pady=0, padx=(5,0), sticky="w")
            self.min_sample_text.grid(row=9, column=0, pady=5, padx=(5,0), sticky="w")
        
            self.eps_box.grid(row=10, column=1, pady=0, padx=(0,5))
            self.min_sample_box.grid(row=9, column=1, pady=10, padx=(0,5))
            self.param_text.grid(columnspan=2,row=8, column=0, pady=(5,0), padx=5)
            
        else:
            pass
        
        if self.cluster_result in ["DBSCAN", "Mean Shift", "Affinity Propagation"]:
            self.cluster_text.grid_forget()
            self.k_slider.grid_forget()
        else:
            self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
            self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)


    def yellowbrick(self):
        # Visualize the elbow method for cluster selection
        try:
            plt.rcParams.update({'font.size': 10})
            
            fig = plt.figure(figsize=(3, 2))
            ax = fig.add_subplot(111)
            ax.axis('off')
            
            model = self.pipe.named_steps['model']
            if 'n_clusters' in model.get_params():
                del model.get_params()['n_clusters']
            visualizer = KElbowVisualizer(model, k=(1, 12), timing=False, title="KElbowVisualizer").fit(self.X)
            
            visualizer.ax.set_xlabel("")
            visualizer.ax.set_ylabel("")
            visualizer.ax.set_yticklabels([])
            visualizer.ax.set_xticklabels([])
            
            visualizer.finalize()
            
            self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
            toolbar = CustomToolbar(self.canvas1, self.legend_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            self.canvas1.get_tk_widget().pack(fill="both", expand=True)
            
            plt.close(fig)
        except:
            pass

    def estimate_eps(self, k):
        # Estimate epsilon for DBSCAN using k-distance graph
        from sklearn.neighbors import NearestNeighbors
        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        # Compute the k-nearest neighbors
        nearest_neighbors = NearestNeighbors(n_neighbors=k)
        nearest_neighbors.fit(self.X)
        distances, indices = nearest_neighbors.kneighbors(self.X)

        # Sort the distances (distances are ordered for each point)
        distances = np.sort(distances, axis=0)
        distances = distances[:, 1]  # Taking the distance to the k-th nearest neighbor

        # Plot the k-distance graph
        fig, ax = plt.subplots(figsize=(3, 2))
        ax.plot(distances)
        ax.set_title('k-Distance Graph')
        ax.set_xlabel('Points sorted by distance')
        ax.set_ylabel(f'{k}th Nearest Neighbor Distance')


        self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
        self.canvas1.get_tk_widget().pack(fill="both", expand=True)
        plt.show()

        elbow_point_index = 80  
        eps_value = distances[elbow_point_index]
        return eps_value

    def show_cluster(self):         
        # Check if there is an active tab in the shared container
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("2D Cluster")

        # Clear the content frame and legend frame
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
            
        for widget in self.legend_frame.winfo_children():
            widget.destroy()
        
        axis1 = self.axis1_combo.get()
        axis2 = self.axis2_combo.get()
        k = self.k_slider.get()

        # Ensure the selected axes are different
        if axis1 != axis2:
            self.df[axis1] = self.df[axis1].apply(lambda x: x.replace('<', '') if isinstance(x, str) and '<' in x else x).astype(float)
            self.df[axis2] = self.df[axis2].apply(lambda x: x.replace('<', '') if isinstance(x, str) and '<' in x else x).astype(float)
            
            self.X = self.df[[axis1, axis2]].values
            
            if hasattr(self, 'cluster_print') and self.cluster_print is not None:
                self.cluster_print.destroy()
                del self.cluster_print

            self.cluster_print = ctk.CTkLabel(self.box_frame_sub, text=f"Cluster: {self.cluster_result}")
            self.cluster_print.grid(columnspan=2, row=14, column=0, pady=(5,0), padx=5)

            # Set up the clustering pipeline based on selected method
            if self.cluster_result == "K-mean":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "DBSCAN":

                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", DBSCAN(eps=self.eps_box.get(), min_samples= int(self.min_sample_box.get())))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif self.cluster_result == "Mean Shift":  
                from sklearn.cluster import estimate_bandwidth
                bandwidth_value = estimate_bandwidth(self.X, quantile=0.2)
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", MeanShift(bandwidth=bandwidth_value))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif self.cluster_result == "Spectral":  
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", SpectralClustering(n_clusters=k, affinity='nearest_neighbors', random_state=0))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "GMM":   
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", GaussianMixture(n_components=k, random_state=0))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "Affinity Propagation":   
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", AffinityPropagation(random_state=0))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif self.cluster_result == "Hierarchical":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", AgglomerativeClustering(n_clusters=k, linkage='ward'))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "BIRCH":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", Birch(n_clusters=k))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
    
            self.pipe.fit(self.X)
            
            if self.cluster_result == "GMM":  
                self.df['cluster'] = self.pipe.named_steps["model"].predict(self.X)
            else:
                self.df['cluster'] = self.pipe.named_steps['model'].labels_
            
            # Create a meshgrid for the plot
            try:
                fig, ax = plt.subplots(figsize=(9, 7))
                
                rangex = (self.X[:, 0].max()-self.X[:, 0].min())/10
                rangey = (self.X[:, 1].max()-self.X[:, 1].min())/10
                x_min, x_max = self.X[:, 0].min() - rangex, self.X[:, 0].max() + rangex
                y_min, y_max = self.X[:, 1].min() - rangey, self.X[:, 1].max() + rangey
                # print (rangex)
                # print (rangey)

                # # Get the width of the axis in pixels
                # bbox = ax.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
                # width_in_inches = bbox.width
                # height_in_inches = bbox.height
                # stepsize_x = width_in_inches / 100
                # stepsize_y = height_in_inches / 100
                
                stepsize_x = rangex/75
                stepsize_y = rangey/75
                
                print (stepsize_x)
                print (stepsize_y)
                xx, yy = np.meshgrid(np.arange(x_min, x_max, stepsize_x), np.arange(y_min, y_max, stepsize_y))
                
                if self.cluster_result == "Spectral" or self.cluster_result == "DBSCAN" or self.cluster_result == "Hierarchical":
                    self.pipe.named_steps['scale'].fit(self.X)
                    X_scaled = self.pipe.named_steps['scale'].transform(self.X)
                    self.pipe.named_steps['model'].fit(X_scaled)
            
                    def predict_new_data(new_data, pipe, train_data, train_labels):
                        new_data_scaled = pipe.named_steps['scale'].transform(new_data)
                        distances = np.linalg.norm(new_data_scaled[:, np.newaxis] - train_data[np.newaxis, :], axis=2)
                        nearest_indices = np.argmin(distances, axis=1)
                        return train_labels[nearest_indices]
            
                    labels = self.pipe.named_steps['model'].labels_
    
                    new_data = np.c_[xx.ravel(), yy.ravel()]
                    new_data_labels = predict_new_data(new_data, self.pipe, X_scaled, labels)
            
                    Z = new_data_labels.reshape(xx.shape)
    
                else:
                    Z = self.pipe.named_steps['model'].predict(self.pipe.named_steps['scale'].transform(np.c_[xx.ravel(), yy.ravel()]))
                    Z = Z.reshape(xx.shape)
    
                ax.contourf(xx, yy, Z, cmap=plt.cm.Pastel1, alpha=0.2)
                
                def map_size(x, new_min, new_max):
                    old_min = self.min_ele
                    old_max = self.max_ele
                    return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
            except:
                fig, ax = plt.subplots(figsize=(9, 7))

            self.shapes = []
            colors = dc['Color']

             # Plot the points with shapes and colors
            element_size = self.size_combo1.get()
            if element_size == "N/A":
                if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i], marker=ds["Shapes"][i]
                        )
                        self.shapes.append(shape)

                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors)

                elif "lithology" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors)                        

                elif "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            marker=ds["Shapes"][i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2])                        

                else:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df.iloc[i][axis1], self.df.iloc[i][axis2])
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2])
            else:
                # Try to get element_size column values for size mapping
                self.max_ele = self.df[element_size].max()
                self.min_ele = self.df[element_size].min()
                
                # Function to map element_size values to point sizes
                def map_size(x, new_min, new_max):
                    old_min = self.min_ele
                    old_max = self.max_ele
                    return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
                    
                # New logic for zn
                def size_for_zn(x):
                    if x < 1000:
                        return 20
                    elif 1000 <= x <= 9999:
                        return 70
                    elif x > 10000:
                        return 110
                    return 20  # Default size if none of the conditions match
                
                # Apply size mapping logic
                if element_size == 'Zn':
                    sizes = self.df[element_size].apply(size_for_zn)
                else:
                    sizes = self.df[element_size].apply(lambda x: map_size(x, 20, 100))

                if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i], marker=ds["Shapes"][i], s=sizes[i]
                        )
                        self.shapes.append(shape)   
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors, s=sizes)

                elif "lithology" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i], s=sizes[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors, s=sizes)   

                elif "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            marker=ds["Shapes"][i], s=sizes[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], s=sizes)
                        
                else:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df.iloc[i][axis1], self.df.iloc[i][axis2], self.df.iloc[i][axis3], s=sizes[i])
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], s=sizes)

            if "sample id" in self.cleaned_df.columns:
                description = self.cleaned_df["sample id"]
                self.annotations = {}
                self.selected_points = set()
                
                # Functions to handle annotations on the plot
                def add_annotation(ind, source):
                    if source == self.dots:
                        index = ind["ind"][0]
                        pos = source.get_offsets()[index]
                        text = f"{description.iloc[index]}"
                    else:
                        index = ind
                        pos = (self.df.iloc[index][axis1], self.df.iloc[index][axis2])
                        text = f"{description.iloc[index]}"
                
                    offsets = [(20, 20), (-20, 20), (20, -20), (-20, -20), (40, 40), (-40, 40), (40, -40), (-40, -40)]
                    
                    for offset in offsets:
                        annot = ax.annotate(text, xy=pos, xytext=offset,
                                            textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"),
                                            zorder=100)
                        fig.canvas.draw()
                        bbox = annot.get_window_extent()
                        overlap = False
                        for existing_annot in self.annotations.values():
                            existing_bbox = existing_annot.get_window_extent()
                            if bbox.overlaps(existing_bbox):
                                overlap = True
                                annot.remove()
                                break
                        if not overlap:
                            annot.set_visible(True)
                            self.annotations[index] = annot
                            self.selected_points.add(index)
                            fig.canvas.draw_idle()
                            return 
                
                    annot = ax.annotate(text, xy=pos, xytext=offsets[0],
                                        textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"),
                                        zorder=100)
                    annot.set_visible(True)
                    self.annotations[index] = annot
                    self.selected_points.add(index)
                    fig.canvas.draw_idle()
                
                def remove_annotation(index):
                    annot = self.annotations.pop(index, None)
                    if annot:
                        annot.remove()
                    self.selected_points.discard(index)
                    fig.canvas.draw_idle()
                
                def update_annotations(event):
                    for index, annot in self.annotations.items():
                        if index in self.selected_points:
                            annot.xy = (self.df.iloc[index][axis1], self.df.iloc[index][axis2])
                    fig.canvas.draw_idle()
                
                def on_click(event):
                    if event.inaxes == ax and event.button == MouseButton.LEFT:
                        contains_dots, ind_dots = self.dots.contains(event)
                        contains_shapes = False
                        shape_index = None
                        for i, shape in enumerate(self.shapes):
                            contains_shapes, ind_shapes = shape.contains(event)
                            if contains_shapes:
                                shape_index = i
                                break
                        
                        if contains_dots:
                            index = ind_dots["ind"][0]
                            if index in self.selected_points:
                                remove_annotation(index)
                            else:
                                add_annotation(ind_dots, self.dots)
                        elif contains_shapes:
                            if shape_index in self.selected_points:
                                remove_annotation(shape_index)
                            else:
                                add_annotation(shape_index, shape)
                
                fig.canvas.mpl_connect("button_press_event", on_click)
                fig.canvas.mpl_connect("motion_notify_event", update_annotations)

            # Plot centroids
            scaler = self.pipe.named_steps['scale']

            if self.cluster_result == "BIRCH":
                centroids = self.pipe.named_steps['model'].subcluster_centers_
            elif self.cluster_result == "GMM":
                centroids = self.pipe.named_steps['model'].means_
            elif self.cluster_result == "Spectral" or self.cluster_result == "DBSCAN" or self.cluster_result == "Hierarchical":
                model = self.pipe.named_steps['model']
                labels = model.labels_
                labels = np.array(labels)
                
                unique_labels = np.unique(labels)
                centroids = []
                
                for i in unique_labels:
                    cluster_points = self.pipe.named_steps['scale'].transform(self.X[labels == i])
                    print(f"Cluster {i} points:", cluster_points)
                    centroid = cluster_points.mean(axis=0)
                    centroids.append(centroid)
                        
                centroids = np.array(centroids)

            else:
                centroids = self.pipe.named_steps['model'].cluster_centers_
                
            centroids_original = scaler.inverse_transform(centroids)
            
            ax.scatter(centroids_original[:, 0], centroids_original[:, 1], 
                       marker="X", c="red", s=30, label="centroids")
            ax.legend()
            ax.set_title("K-Means Clustering")
            ax.set_xlabel(axis1)
            ax.set_ylabel(axis2)
            
            canvas = FigureCanvasTkAgg(fig, master=content_frame)
            canvas.draw()

            toolbar = NavigationToolbar2Tk(canvas, content_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)

            self.shape()
            Legend = self.var1.get() == 1
            print(Legend)
            if Legend:
                pass
            else:
                self.yellowbrick()
            self.legend()
            
        else:
            self.output_text.insert("end", 'Select two different elements')

    def shape(self):
        # Display or hide shapes in the plot
        shape = self.var.get() == 1
        if shape:
            for shape in self.shapes:
                shape.set_visible(True)
            self.dots.set_visible(False)
        else:
            self.dots.set_visible(True)
            for shape in self.shapes:
                shape.set_visible(False)

    def save_clusters_to_excel(self):
        # Save the clustered data to an Excel file with colored cells based on clusters
        axis1 = self.axis1_combo.get().lower()
        axis2 = self.axis2_combo.get().lower()
        
        if axis1 == axis2:
            print("Select two different elements to save to Excel.")
            return

        if 'cluster' not in self.df.columns:
            print("Clusters have not been calculated yet.")
            return

        if axis1 not in self.cleaned_df.columns or axis2 not in self.cleaned_df.columns or 'sample id' not in self.cleaned_df.columns:
            print(f"One or more of the columns '{axis1}', '{axis2}', or 'sample id' are missing in the data.")
            return

        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return

        self.cleaned_df['cluster'] = self.df['cluster']
        
        pivot_data = self.cleaned_df.pivot_table(index='sample id', values=[axis1, axis2, 'cluster'], sort=False)
        
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            pivot_data.to_excel(excel_writer, index=True)
        
        wb = load_workbook(file_name)
        ws = wb.active

        colormap = plt.get_cmap('Set2')
        cluster_colors = ['FF{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)) for r, g, b, _ in [colormap(i) for i in range(colormap.N)]]
        cluster_fill = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in cluster_colors]

        for i in range(len(self.cleaned_df)):
            sample_id = self.cleaned_df.iloc[i]['sample id']
            cluster = self.cleaned_df.iloc[i]['cluster']
            row = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
            for cell in row:
                if cell[0].value == sample_id:
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=cell[0].row, column=col).fill = cluster_fill[cluster]

        wb.save(file_name)
        print(f"Clusters saved to {file_name}")

    def legend(self):
        # Create and display legends for lithology and rock type
        Legend = self.var1.get() == 1
        if Legend:
            print ("yes")
            for widget in self.legend_frame.winfo_children():
                widget.destroy()
            if not dc.empty and not ds.empty:
                unique_lithologies = self.cleaned_df['lithology'].unique()
                unique_shapes = self.cleaned_df['rock unit'].unique()
        
                # For lithologies
                handle1 = []
                label1 = []
                for lithology in unique_lithologies:
                    color = color_map[lithology]
                    l1 = Line2D([], [], color=color, marker='o', linestyle='None', label=lithology)
                    handle1.append(l1)
                    label1.append(lithology)
        
                # For shapes
                handle2 = []
                label2 = []
                for shape in unique_shapes:
                    marker = color_map1[shape]
                    l2 = Line2D([], [], color='black', marker=marker, linestyle='None', label=shape)
                    handle2.append(l2)
                    label2.append(shape)
        
                total_entries = len(label1) + len(label2)
                height_per_entry = 0.225
                fig_height = total_entries * height_per_entry
        
                figx = plt.figure(figsize=(2.5, fig_height))
                axx = figx.add_subplot(111)
                axx.axis('off')
        
                self.legend1 = plt.legend(handle1, label1, bbox_to_anchor=(0.9, 1.15), title='Lithology', fontsize=10, labelspacing=0.3)
                self.legend2 = plt.legend(handle2, label2, bbox_to_anchor=(0.9, 0.35), title='Rock Type', fontsize=10, labelspacing=0.3)
                axx.add_artist(self.legend1)
                axx.add_artist(self.legend2)
                self.canvas1 = FigureCanvasTkAgg(figx, master=self.legend_frame)
                toolbar = CustomToolbar(self.canvas1, self.legend_frame)
                toolbar.update()
                toolbar.pack(side=tk.TOP, fill=tk.X)
        
                self.canvas1.get_tk_widget().pack(fill="both", expand=True)
        
                self.legend1.set_visible(self.var1.get() == 1)
                self.legend2.set_visible(self.var1.get() == 1)
            else:
                pass
        else:
            pass







class Cluster3DPlotClass:
    def __init__(self, shared_container, cluster, df, cleaned_df, box_frame, box_frame_sub, on_button_click, legend_frame):
        # Initialize class variables
        self.df = df.dropna()
        self.shared_container = shared_container
        self.cluster_result = cluster
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.legend_frame = legend_frame
        self.cleaned_df = cleaned_df
        self.on_button_click = on_button_click

        self.var = IntVar()
        self.var1 = IntVar()

        self.plot_3d_cluster()

    def plot_3d_cluster(self):
        # Load and resize the image for the 3D plot button
        pil_image = Image.open("mineralAI_images/images_program/3d.png")
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))
        
        # Create a label with the icon and bind a click event to it
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=3, column=2, sticky="w", pady=0, padx=5)
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_3d_cluster_sub0))

        # Add a tooltip to the label
        ToolTip(self.image_button, msg="PCA 3D Biplot with Clusters")

    def plot_3d_cluster_sub0(self):
        # clear column cluster if exist
        if 'cluster' in self.df.columns:
            self.df.drop(columns=['cluster'], inplace=True)
        else:
            pass
        self.box_frame_sub.grid_propagate(True)

        # Clear any existing widgets in the sub-frame
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Create UI elements for selecting axes and other options
        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 1:").grid(columnspan=2, row=0, column=0, pady=(5,0), padx=5)
        self.axis1_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis1_combo.current(0)
        self.axis1_combo.grid(columnspan=2, row=1, column=0, pady=(3,0), padx=5)
        
        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 2:").grid(columnspan=2, row=2, column=0, pady=(5,0), padx=5)
        self.axis2_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis2_combo.current(2)
        self.axis2_combo.grid(columnspan=2, row=3, column=0, pady=(3,0), padx=5)

        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 3:").grid(columnspan=2,row=4, column=0, pady=(5,0), padx=5)
        self.axis3_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis3_combo.current(4)
        self.axis3_combo.grid(columnspan=2,row=5, column=0, pady=(3,0), padx=5)
        
        #select element for data point size
        ctk.CTkLabel(self.box_frame_sub, text="Element for Data Point Size").grid(columnspan=2,row=6, column=0, pady=(5,0), padx=5)
        self.size_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), name="size_combo", state="readonly")
        self.size_combo.set("N/A")
        self.size_combo.grid(columnspan=2,row=7, column=0, pady=(3,5), padx=5)

        # Clean and prepare data for clustering
        axis1 = self.axis1_combo.get()
        axis2 = self.axis2_combo.get()
        axis3 = self.axis3_combo.get()
        
        for col in [axis1, axis2, axis3]:
            self.df[col] = self.df[col].apply(lambda x: x.replace('<', '') if isinstance(x, str) and '<' in x else x)
        self.df = self.df.astype({axis1: 'float', axis2: 'float', axis3: 'float'})

        self.X = self.df[[axis1, axis2, axis3]].values
        
        
        # Add checkbox for displaying shapes if ds is not empty
        if not ds.empty:
            self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Display shape", variable=self.var)
            self.checkbox.grid(columnspan=2,row=8, column=0, sticky=W, pady=(5,0), padx=5)

        self.legend_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Show Legend", variable=self.var1)
        self.legend_checkbox.grid(columnspan=2,row=9, column=0, sticky=W, pady=(5,0), padx=5)
        
        # Add checkbox for coloring by cluster if dc is not empty
        if not dc.empty:
            self.color_checkbox_var = IntVar()
            self.color_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Color by Cluster", variable=self.color_checkbox_var)
            self.color_checkbox.grid(columnspan=2,row=10, column=0, sticky=W, pady=(5,0), padx=5)

        # Add slider for selecting the number of clusters
        self.cluster_text = ctk.CTkLabel(self.box_frame_sub, text="Number of Clusters:")
        self.k_slider = tk.Scale(self.box_frame_sub, from_=2, to=10, orient=tk.HORIZONTAL)
        
        # Add apply button
        self.apply_plot_3d_cluster = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.show_cluster)
        self.apply_plot_3d_cluster.grid(columnspan=2, row=16, column=0, pady=(5,0), padx=5)

        # Add save button
        self.save_button = ctk.CTkButton(self.box_frame_sub, text="Save Clusters to Excel", command=self.save_clusters_to_excel)
        self.save_button.pack(side="bottom", padx=5, pady=5)
                
        # Define the values for the Spinbox
        linkage_box_values = ('ward', 'complete', 'average', 'single')
        self.linkage_box = ttk.Combobox(self.box_frame_sub, values=linkage_box_values, state="readonly")
        self.linkage_box.current(0)

        affinity_box_values = ('nearest_neighbors', 'rbf', 'precomputed', 'precomputed_nearest_neighbors')
        self.affinity_box = ttk.Combobox(self.box_frame_sub, values=affinity_box_values, state="readonly")
        self.affinity_box.current(0)
                
        self.min_sample_box = CTkSpinbox(self.box_frame_sub, step_size=1, min_value=2, max_value=100, width = 110) #, command=self.runeps)
        self.eps_box = CTkSpinbox(self.box_frame_sub, step_size=0.1, min_value=0.1, max_value=100, width = 110)
                    
        self.eps_text = ctk.CTkLabel(self.box_frame_sub, text="eps")
        self.min_sample_text = ctk.CTkLabel(self.box_frame_sub, text="min_sample")
        
        self.plot_3d_cluster_sub(self.cluster_result)

    # self, *arg):
    #     self.eps_value = self.estimate_eps(self.min_sample_box.get())
    # self.eps_box = CTkSpinbox(self.box_frame_sub, step_value=1, scroll_value=1, start_value=self.eps_value, min_value=1, max_value=10000)    
    # self.eps_box.grid(columnspan=2, row=13, column=0, pady=0, padx=5)    def runeps(
        
    def plot_3d_cluster_sub(self, cluster):
        self.cluster_result = cluster
        
        self.linkage_box.grid_forget()
        self.affinity_box.grid_forget()
        self.min_sample_box.grid_forget()
        self.eps_box.grid_forget()
        self.eps_text.grid_forget()
        self.min_sample_text.grid_forget()
        
        try:
            self.param_text.grid_forget()
        except:
            pass

        if self.cluster_result == "Hierarchical":
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Linkage Type:")
            self.linkage_box.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
            
        elif self.cluster_result == "Spectral":
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Affinity Type:")
            self.affinity_box.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)

        elif self.cluster_result == "DBSCAN":
            # self.runeps() 
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="eps & min_sample value:")

            self.eps_text.grid(row=13, column=0, pady=0, padx=(5,0), sticky="w")
            self.min_sample_text.grid(row=12, column=0, pady=5, padx=(5,0), sticky="w")
        
            self.eps_box.grid(row=13, column=1, pady=0, padx=(0,5))
            self.min_sample_box.grid(row=12, column=1, pady=10, padx=(0,5))
            self.param_text.grid(columnspan=2,row=11, column=0, pady=(5,0), padx=5)
            
        else:
            pass
            
            
        if self.cluster_result in ["DBSCAN", "Mean Shift", "Affinity Propagation"]:
            self.cluster_text.grid_forget()
            self.k_slider.grid_forget()
        else:
            self.cluster_text.grid(columnspan=2,row=14, column=0, pady=0, padx=5)
            self.k_slider.grid(columnspan=2,row=15, column=0, pady=(3,5), padx=5)

        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        self.yellowbrick()
        

    def yellowbrick(self): 
        # Use Yellowbrick's KElbowVisualizer to help determine the optimal number of clusters
        try:
            plt.rcParams.update({'font.size': 10})
            
            fig = plt.figure(figsize=(3, 2))
            ax = fig.add_subplot(111)
            ax.axis('off')
            
            model = self.pipe.named_steps['model']
            if 'n_clusters' in model.get_params():
                del model.get_params()['n_clusters']
                
            visualizer = KElbowVisualizer(model, k=(1, 12), timing=False, title="KElbowVisualizer").fit(self.X)
            
            visualizer.ax.set_yticklabels([])
            visualizer.finalize()
            
            self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
            toolbar = CustomToolbar(self.canvas1, self.legend_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            self.canvas1.get_tk_widget().pack(fill="both", expand=True)
            
            plt.close(fig)
        except:
            pass

    # def estimate_eps(self, k):
    #     from kneed import KneeLocator
    # from sklearn.neighbors import NearestNeighbors    
    #     for widget in self.legend_frame.winfo_children():
    #         widget.destroy()
    
    #     # Compute the k-nearest neighbors
    #     nearest_neighbors = NearestNeighbors(n_neighbors=k)
    #     nearest_neighbors.fit(self.X)
    #     distances, indices = nearest_neighbors.kneighbors(self.X)
    
    #     # Sort the distances (distances are ordered for each point)
    #     distances = np.sort(distances, axis=0)
    #     distances = distances[:, k-1]  # Taking the distance to the k-th nearest neighbor
    
    #     # Automatically detect the elbow point
    #     knee_locator = KneeLocator(range(len(distances)), distances, curve='convex', direction='increasing')
    #     elbow_point_index = knee_locator.elbow
    
    #     # Plot the k-distance graph
    #     fig, ax = plt.subplots(figsize=(3, 2))
    #     ax.plot(distances)
    #     ax.axvline(x=elbow_point_index, color='red', linestyle='--', label=f'Elbow at {elbow_point_index}')
    #     ax.set_title('k-Distance Graph')
    #     ax.set_xlabel('Points sorted by distance')
    #     ax.set_ylabel(f'{k}th Nearest Neighbor Distance')
    #     ax.legend()
    
    #     # Display the plot in the legend frame
    #     self.canvas1 = FigureCanvasTkAgg(fig, master=self.legend_frame)
    #     self.canvas1.get_tk_widget().pack(fill="both", expand=True)
    #     plt.show()
    
    #     # Estimate epsilon from the k-distance graph
    #     eps_value = distances[elbow_point_index]
    #     return eps_value
        
    def show_cluster(self):
        # Display the 3D cluster plot based on the selected axes and clustering algorithm
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("3D Cluster")

        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        # get combobox values
        axis1 = self.axis1_combo.get()
        axis2 = self.axis2_combo.get()
        axis3 = self.axis3_combo.get()
        linkage = self.linkage_box.get()
        affinity = self.affinity_box.get()
        
        k = self.k_slider.get()

        # Ensure the selected axes are different
        if axis1 != axis2 and axis1 != axis3 and axis2 != axis3:
            # Remove previous cluster result label if it exists
            if hasattr(self, 'cluster_print') and self.cluster_print is not None:
                self.cluster_print.destroy()
                del self.cluster_print

            # Get the selected clustering algorithm
            cluster_result = self.cluster_result
            self.cluster_print = ctk.CTkLabel(self.box_frame_sub, text=f"Cluster: {cluster_result}")
            self.cluster_print.grid(columnspan=2,row=17, column=0, pady=(5,0), padx=5)
            
            # Set up the clustering pipeline based on the selected algorithm
            if cluster_result == "K-mean":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
                ])
                self.cluster_text.grid(columnspan=2,row=14, column=0, pady=0, padx=5)
                self.k_slider.grid(columnspan=2,row=15, column=0, pady=(3,5), padx=5)
                
            elif cluster_result == "DBSCAN":
                print(self.eps_box.get())
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", DBSCAN(eps=self.eps_box.get(), min_samples= int(self.min_sample_box.get())))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif cluster_result == "Mean Shift":  
                from sklearn.cluster import estimate_bandwidth
                # bandwidth_value = estimate_bandwidth(self.X, quantile=0.2)
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", MeanShift())
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                    
            elif cluster_result == "Spectral":  
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", SpectralClustering(n_clusters=k, affinity=affinity, random_state=0))
                ])
                self.cluster_text.grid(columnspan=2,row=14, column=0, pady=0, padx=5)
                self.k_slider.grid(columnspan=2,row=15, column=0, pady=(3,5), padx=5)
                
            elif cluster_result == "GMM":   
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", GaussianMixture(n_components=k, random_state=0))
                ])
                self.cluster_text.grid(columnspan=2,row=14, column=0, pady=0, padx=5)
                self.k_slider.grid(columnspan=2,row=15, column=0, pady=(3,5), padx=5)
                
            elif cluster_result == "Affinity Propagation":   
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", AffinityPropagation(random_state=0))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif cluster_result == "Hierarchical":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", AgglomerativeClustering(n_clusters=k, linkage=linkage))
                ])
                self.cluster_text.grid(columnspan=2,row=14, column=0, pady=0, padx=5)
                self.k_slider.grid(columnspan=2,row=15, column=0, pady=(3,5), padx=5)
                
            elif cluster_result == "BIRCH":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", Birch(n_clusters=k))
                ])
                self.cluster_text.grid(columnspan=2,row=14, column=0, pady=0, padx=5)
                self.k_slider.grid(columnspan=2,row=15, column=0, pady=(3,5), padx=5)
    
            # Fit the clustering pipeline to the data
            self.pipe.fit(self.X)
            
            # Get cluster labels
            if cluster_result == "GMM":  
                self.df['cluster'] = self.pipe.named_steps["model"].predict(self.X)
            else:
                self.df['cluster'] = self.pipe.named_steps['model'].labels_

            # Create a 3D plot
            fig = plt.figure(figsize=(10, 8))
            ax = fig.add_subplot(111, projection='3d')
                
            self.shapes = []

            # Color points by cluster if the checkbox is selected
            color_by_cluster = self.color_checkbox_var.get() == 1
            if color_by_cluster:
                unique_clusters = self.df['cluster'].unique()
                colormap = plt.get_cmap('tab10')
                cluster_colors = {cluster: colormap(i / len(unique_clusters)) for i, cluster in enumerate(unique_clusters)}
                colors = self.df['cluster'].map(cluster_colors)
            else:
                colors = dc['Color']
            
            # Plot the points with shapes and colors
            element_size = self.size_combo.get()
            if element_size == "N/A":
                if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i], self.df[axis3][i],
                            color=colors[i], marker=ds["Shapes"][i]
                        )
                        self.shapes.append(shape)

                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], c=colors)

                elif "lithology" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i], self.df[axis3][i],
                            color=colors[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], c=colors)                        

                elif "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i], self.df[axis3][i],
                            marker=ds["Shapes"][i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3])                        

                else:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df.iloc[i][axis1], self.df.iloc[i][axis2], self.df.iloc[i][axis3])
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3])
            else:
                # Try to get element_size column values for size mapping
                self.max_ele = self.df[element_size].max()
                self.min_ele = self.df[element_size].min()
                
                # Function to map element_size values to point sizes
                def map_size(x, new_min, new_max):
                    old_min = self.min_ele
                    old_max = self.max_ele
                    return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
                    
                # New logic for zn
                def size_for_zn(x):
                    if x < 1000:
                        return 20
                    elif 1000 <= x <= 9999:
                        return 70
                    elif x > 10000:
                        return 110
                    return 20  # Default size if none of the conditions match
                
                # Apply size mapping logic
                if element_size == 'Zn':
                    sizes = self.df[element_size].apply(size_for_zn)
                else:
                    sizes = self.df[element_size].apply(lambda x: map_size(x, 20, 100))

                if "lithology" in self.cleaned_df.columns and "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i], self.df[axis3][i],
                            color=colors[i], marker=ds["Shapes"][i], s=sizes[i]
                        )
                        self.shapes.append(shape)   
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], c=colors, s=sizes)

                elif "lithology" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i], self.df[axis3][i],
                            color=colors[i], s=sizes[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], c=colors, s=sizes)   

                elif "rock unit" in self.cleaned_df.columns:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i], self.df[axis3][i],
                            marker=ds["Shapes"][i], s=sizes[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], s=sizes)
                        
                else:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df.iloc[i][axis1], self.df.iloc[i][axis2], self.df.iloc[i][axis3], s=sizes[i])
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], self.df[axis3], s=sizes)


            # Annotate points with sample IDs if available
            if "sample id" in self.cleaned_df.columns:
                description = self.cleaned_df["sample id"]
                self.annotations = {}
                self.selected_points = set()
            
                from mpl_toolkits.mplot3d import proj3d
                
                def add_annotation(ind, source):
                    if source == self.dots:
                        index = ind["ind"][0]
                        pos = source.get_offsets()[index]
                        text = f"{description.iloc[index]}"
                    else:
                        index = ind
                        x, y, _ = proj3d.proj_transform(self.df.iloc[index][axis1], self.df.iloc[index][axis2], self.df.iloc[index][axis3], ax.get_proj())
                        pos = (x, y)
                        text = f"{description.iloc[index]}"

                    offsets = [(20, 20), (-20, 20), (20, -20), (-20, -20), (40, 40), (-40, 40), (40, -40), (-40, -40)]
                    
                    for offset in offsets:
                        annot = ax.annotate(text, xy=pos, xytext=offset,
                                            textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->",),
                                            zorder=100)
                        fig.canvas.draw()
                        bbox = annot.get_window_extent()
                        overlap = False
                        for existing_annot in self.annotations.values():
                            existing_bbox = existing_annot.get_window_extent()
                            if bbox.overlaps(existing_bbox):
                                overlap = True
                                annot.remove()
                                break
                        if not overlap:
                            annot.set_visible(True)
                            self.annotations[index] = annot
                            self.selected_points.add(index)
                            fig.canvas.draw_idle()
                            return 
                
                    annot = ax.annotate(text, xy=pos, xytext=offsets[0],
                                        textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"),
                                        zorder=100)
                    annot.set_visible(True)
                    self.annotations[index] = annot
                    self.selected_points.add(index)
                    fig.canvas.draw_idle()
                
                def remove_annotation(index):
                    annot = self.annotations.pop(index, None)
                    if annot:
                        annot.remove()
                    self.selected_points.discard(index)
                    fig.canvas.draw_idle()
                
                def update_annotations(event):
                    for index, annot in self.annotations.items():
                        if index in self.selected_points:
                            x, y, _ = proj3d.proj_transform(self.df.iloc[index][axis1], self.df.iloc[index][axis2], self.df.iloc[index][axis3], ax.get_proj())
                            annot.xy = (x, y)
                    fig.canvas.draw_idle()
                
                def on_click(event):
                    if event.inaxes == ax and event.button == MouseButton.LEFT:
                        contains_dots, ind_dots = self.dots.contains(event)
                        contains_shapes = False
                        shape_index = None
                        for i, shape in enumerate(self.shapes):
                            contains_shapes, ind_shapes = shape.contains(event)
                            if contains_shapes:
                                shape_index = i
                                break
                        
                        if contains_dots:
                            index = ind_dots["ind"][0]
                            if index in self.selected_points:
                                remove_annotation(index)
                            else:
                                add_annotation(ind_dots, self.dots)
                        elif contains_shapes:
                            if shape_index in self.selected_points:
                                remove_annotation(shape_index)
                            else:
                                add_annotation(shape_index, shape)
                
                fig.canvas.mpl_connect("button_press_event", on_click)
                fig.canvas.mpl_connect("motion_notify_event", update_annotations)

            # Get centroids of the clusters
            scaler = self.pipe.named_steps['scale']
            if cluster_result == "BIRCH":
                centroids = self.pipe.named_steps['model'].subcluster_centers_
            elif cluster_result == "GMM":
                centroids = self.pipe.named_steps['model'].means_
            elif cluster_result == "Spectral" or cluster_result == "DBSCAN" or cluster_result == "Hierarchical":
                model = self.pipe.named_steps['model']
                labels = model.labels_
                labels = np.array(labels)
                
                unique_labels = np.unique(labels)
                centroids = []
                
                for i in unique_labels:
                    cluster_points = self.pipe.named_steps['scale'].transform(self.X[labels == i])
                    centroid = cluster_points.mean(axis=0)
                    centroids.append(centroid)
                        
                centroids = np.array(centroids)
            else:
                centroids = self.pipe.named_steps['model'].cluster_centers_
                
            # Transform centroids back to original scale
            centroids_original = scaler.inverse_transform(centroids)
            
            # Plot centroids in red with X marker
            ax.scatter(centroids_original[:, 0], centroids_original[:, 1], centroids_original[:, 2], 
                       marker="X", c="red", s=100, label="centroids")
        
            ax.set_xlabel(axis1)
            ax.set_ylabel(axis2)
            ax.set_zlabel(axis3)
            plt.legend()
            plt.title("3D K-Means Clustering")

            # Create a canvas for the plot and add toolbar
            canvas = FigureCanvasTkAgg(fig, master=content_frame)
            canvas.draw()
            toolbar = NavigationToolbar2Tk(canvas, content_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)

            # Toggle shape visibility and call Yellowbrick visualizer
            self.shape()
            Legend = self.var1.get() == 1
            print(Legend)
            if Legend:
                pass
            else:
                self.yellowbrick()
            self.legend()
            
        else:
            self.output_text.insert("end", 'Select three different elements')

    def shape(self):
        # Toggle the visibility of shapes in the plot
        shape = self.var.get() == 1
        if shape:
            for shape in self.shapes:
                shape.set_visible(True)
            self.dots.set_visible(False)
        else:
            self.dots.set_visible(True)
            for shape in self.shapes:
                shape.set_visible(False)

    def save_clusters_to_excel(self):
        # Save the clustered data to an Excel file with colored cells based on clusters
        axis1 = self.axis1_combo.get().lower()
        axis2 = self.axis2_combo.get().lower()
        axis3 = self.axis3_combo.get().lower()
        
        # Ensure that the selected axes are different
        if axis1 == axis2 == axis3:
            print("Select three different elements to save to Excel.")
            return

        if 'cluster' not in self.df.columns:
            print("Clusters have not been calculated yet.")
            return

        # Ensure the required columns exist in the cleaned DataFrame
        if axis1 not in self.cleaned_df.columns or axis2 not in self.cleaned_df.columns or axis3 not in self.cleaned_df.columns or 'sample id' not in self.cleaned_df.columns:
            print(f"One or more of the columns '{axis1}', '{axis2}', '{axis3}', or 'sample id' are missing in the data.")
            returnv

        # Prompt the user to select a file name for saving the Excel file
        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return
        
        # Add the cluster labels to the cleaned DataFrame
        self.cleaned_df['cluster'] = self.df['cluster']
        
        # Pivot the data to have sample IDs as index and selected elements as columns
        pivot_data = self.cleaned_df.pivot_table(index='sample id', values=[axis1, axis2, axis3, 'cluster'], sort=False)
        
        # Save the pivoted data to an Excel file
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            pivot_data.to_excel(excel_writer, index=True)
        
        wb = load_workbook(file_name)
        ws = wb.active

        # Generate colors for each cluster
        colormap = plt.get_cmap('Set2')
        cluster_colors = ['FF{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)) for r, g, b, _ in [colormap(i) for i in range(colormap.N)]]
        cluster_fill = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in cluster_colors]

        # Apply color to each cell based on cluster
        for i in range(len(self.cleaned_df)):
            sample_id = self.cleaned_df.iloc[i]['sample id']
            cluster = self.cleaned_df.iloc[i]['cluster']
            row = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
            for cell in row:
                if cell[0].value == sample_id:
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=cell[0].row, column=col).fill = cluster_fill[cluster]

        wb.save(file_name)
        print(f"Clusters saved to {file_name}")

    def legend(self):
        # Create and display legends for lithology and rock type
        Legend = self.var1.get() == 1
        if Legend:
            print ("yes")
            for widget in self.legend_frame.winfo_children():
                widget.destroy()
            if not dc.empty and not ds.empty:
                unique_lithologies = self.cleaned_df['lithology'].unique()
                unique_shapes = self.cleaned_df['rock unit'].unique()
        
                # For lithologies
                handle1 = []
                label1 = []
                for lithology in unique_lithologies:
                    color = color_map[lithology]
                    l1 = Line2D([], [], color=color, marker='o', linestyle='None', label=lithology)
                    handle1.append(l1)
                    label1.append(lithology)
        
                # For shapes
                handle2 = []
                label2 = []
                for shape in unique_shapes:
                    marker = color_map1[shape]
                    l2 = Line2D([], [], color='black', marker=marker, linestyle='None', label=shape)
                    handle2.append(l2)
                    label2.append(shape)
        
                total_entries = len(label1) + len(label2)
                height_per_entry = 0.225
                fig_height = total_entries * height_per_entry
        
                figx = plt.figure(figsize=(2.5, fig_height))
                axx = figx.add_subplot(111)
                axx.axis('off')
        
                self.legend1 = plt.legend(handle1, label1, bbox_to_anchor=(0.9, 1.15), title='Lithology', fontsize=10, labelspacing=0.3)
                self.legend2 = plt.legend(handle2, label2, bbox_to_anchor=(0.9, 0.35), title='Rock Type', fontsize=10, labelspacing=0.3)
                axx.add_artist(self.legend1)
                axx.add_artist(self.legend2)
                self.canvas1 = FigureCanvasTkAgg(figx, master=self.legend_frame)
                toolbar = CustomToolbar(self.canvas1, self.legend_frame)
                toolbar.update()
                toolbar.pack(side=tk.TOP, fill=tk.X)
        
                self.canvas1.get_tk_widget().pack(fill="both", expand=True)
        
                self.legend1.set_visible(self.var1.get() == 1)
                self.legend2.set_visible(self.var1.get() == 1)
            else:
                pass
        else:
            pass






class drill_class:
    def __init__(self, shared_container, pca_df_scaled, cleaned_df, df, box_frame, box_frame_sub, on_button_click, apply_button):
        self.shared_container = shared_container
        self.apply_button = apply_button
        self.on_button_click = on_button_click
        self.pca_df_scaled = pca_df_scaled
        self.cleaned_df = cleaned_df
        self.df = df
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.var = IntVar()

        # Initialize the drill hole graph
        self.Graph_PCA()

    def Graph_PCA(self):
        # Load and resize the image for the button icon
        pil_image = Image.open("mineralAI_images/images_program/drill-svgrepo-com.png")
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        # Create a label with the icon and bind a click event to it
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=1, column=4, sticky="w", pady=0, padx=5)
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.Graph_PCA_sub))

        # Add a tooltip to the label
        ToolTip(self.image_button, msg="Drill Hole Chart")

    def Graph_PCA_sub(self):
        # Clean and process sample IDs
        self.cleaned_df['Sample_ID_Clean'] = self.cleaned_df['sample id'].apply(lambda x: x.replace('-', ''))
        self.cleaned_df['Drill_Hole'] = self.cleaned_df['Sample_ID_Clean'].apply(lambda x: x[:3])
        
        # Extract depth unit from column name
        column_name = self.cleaned_df.filter(like='depth').columns[0]
        match = re.search(r'\((.*?)\)', column_name)
        self.unit = match.group(1) if match else None
        
        self.cleaned_df['Depth'] = self.cleaned_df[column_name]
        self.drill_holes = sorted(self.cleaned_df['Drill_Hole'].unique().tolist())

        # Clear previous widgets
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Create a dropdown for selecting drill hole
        ctk.CTkLabel(self.box_frame_sub, text="Select Drill Hole:").pack(side="top", padx=5, pady=(5, 0))
        self.drill_hole_combo = ttk.Combobox(self.box_frame_sub, values=self.drill_holes, state="readonly")
        self.drill_hole_combo.current(0)
        self.drill_hole_combo.pack(side="top", padx=5, pady=(5, 0))

        # Add an apply button
        self.apply_Graph_PCA = ctk.CTkButton(self.box_frame_sub, text="apply", command=self.update_plots)
        self.apply_Graph_PCA.pack(side="top", padx=5, pady=5)

    def normalize(self, values, vmin=None, vmax=None):
        # Normalize values to range [0, 1]
        vmin = vmin if vmin is not None else np.min(values)
        vmax = vmax if vmin is not None else np.max(values)
        norm_values = (values - vmin) / (vmax - vmin)
        return norm_values

    def plot_pca_scatter(self, data, y_positions, ax):
        # Normalize data and create scatter plot with color mapping
        norm_values = self.normalize(data, vmin=-max(abs(data)), vmax=max(abs(data)))
        scatter_colors = plt.cm.coolwarm(norm_values)
        ax.scatter(data, y_positions, color=scatter_colors, s=30)
        ax.set_yticks(y_positions)
        ax.set_yticklabels([])
        ax.invert_yaxis()
        ax.grid(True, linestyle='--', linewidth=0.5)
        ax.axvline(x=0, color='black', linewidth=0.5)
        max_val = max(abs(data.min()), data.max())
        ax.set_xlim(-max_val, max_val)
        ax.tick_params(rotation=-35, labelsize="medium")

        # Customize plot spines
        for spine in ax.spines.values():
            spine.set_edgecolor('black')
            spine.set_linewidth(1)

    def plot_depth_chart(self, depths, y_positions, ax):
        # Plot depth chart with horizontal lines representing depths
        ax.hlines(y=y_positions, xmin=0, xmax=0.2, linestyles='-', linewidth=0.8)
        ax.set_yticks(y_positions)
        ax.set_yticklabels(depths)
        ax.invert_yaxis()
        ax.set_xticklabels([])
        ax.set_title(f'Depth ({self.unit})')
        ax.axis('on')
        ax.set_ylabel(self.drill_hole)
        ax.tick_params(labelsize="small")

    def update_plots(self):
        # Create a new tab if not already present
        if not self.shared_container.current_tab:
            self.shared_container.create_tab("Drill Hole")

        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()
        
        # Filter and sort data for the selected drill hole
        self.drill_hole = self.drill_hole_combo.get()
        filtered_cleaned_df = self.cleaned_df[self.cleaned_df['Drill_Hole'] == self.drill_hole]
        if filtered_cleaned_df.empty:
            self.output_text.insert("end", "No samples found for the selected drill hole.")
            return

        filtered_cleaned_df = filtered_cleaned_df.sort_values(by='Depth')
        sample_indices = filtered_cleaned_df.index
        depths = filtered_cleaned_df['Depth'].tolist()
        if len(sample_indices) == 0:
            self.output_text.insert("end", "No samples found for the selected drill hole.")
            return

        # Calculate y positions for depth markers
        print(depths)
        max_depth = max(depths)
        min_depth = min(depths)
        if max_depth == min_depth:
            y_positions = [0.5 for _ in depths]
        else:
            y_positions = [((depth - min_depth) / (max_depth - min_depth)) * len(depths) for depth in depths]

        plt.close('all')

        # Create subplots for depth chart and PCA scatter plots
        fig, axes = plt.subplots(nrows=1, ncols=len(self.pca_df_scaled.columns) + 1, figsize=(15, 10))
        pos1 = axes[0].get_position()
        axes[0].set_position([pos1.x0, pos1.y0, pos1.width * 0.5, pos1.height])
        self.plot_depth_chart(depths, y_positions, axes[0])

        # Plot PCA scatter plots for each principal component
        for col, pc in enumerate(self.pca_df_scaled.columns):
            data = self.pca_df_scaled.loc[sample_indices, pc]
            ax = axes[col + 1]
            self.plot_pca_scatter(data, y_positions, ax)
            ax.set_title(f'PC{col + 1}')

        # Embed the plot in the Tkinter GUI
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()
        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)


    # def Graph_PCA(self):
    #     pil_image = Image.open("mineralAI_images/images_program/drill-svgrepo-com.png")
    #     #resized_image = pil_image.resize((32, 32), Image.LANCZOS)
    #     self.icon_image = ctk.CTkImage(light_image=pil_image, size=(32, 32))
    #     #self.icon_image = ImageTk.PhotoImage(resized_image)
        
    #     self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text = "", width=20)
    #     self.image_button.grid(row=4, column=0, sticky="w", pady=(5,0), padx=(5,0))

    #     self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.Graph_PCA_sub))

    #     ToolTip(self.image_button, msg="Drill Hole Chart")
             
    #     # #self.graph_bar_button = ctk.CTkButton(self.box_frame, text="Drill PC graph", command=lambda: self.on_button_click(self.graph_bar_button, self.Graph_PCA_sub))

        
    # #     self.icon_image = Image.open("mineralAI_images/images_program/chart-scatter-3d-svgrepo-com.png")
    # #     print(self.icon_image)
    # #     self.ctk_icon_image = ctk.CTkImage(light_image=self.icon_image, size=(24, 24))
    # #     print(self.ctk_icon_image)
        
    # #     self.graph_bar_button = ctk.CTkButton(
    # #         self.box_frame, 
    # #         image=self.ctk_icon_image, 
    # #         text="",
    # #         command=lambda: self.on_button_click(self.graph_bar_button, self.Graph_PCA_sub)
    # #     )
      
    # #     self.graph_bar_button.pack(side="top", padx=5, pady=(5, 0))






class CTkSpinbox(ctk.CTkFrame):
         
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: int = 1.0,
                 min_value: int = 0.0,
                 max_value: int = 1000.0,
                 button_color: str = None,
                 button_hover_color: str = None,
                 entry_color: str = None,
                 text_color: str = None,
                 border_color: str = None,
                 entry_border: str = 2,
                 command: Callable = None,
                 **kwargs):
        
        super().__init__(*args, width=width, height=height, **kwargs)
        

        self.step_size = step_size
        self.max_value = max_value
        self.min_value = min_value
        self.command = command
        self.validation = self.register(self.only_numbers)
        
        self.grid_columnconfigure((0, 2), weight=0) 
        self.grid_columnconfigure(1, weight=1)
        
        self.button_color = ctk.ThemeManager.theme["CTkButton"]["fg_color"] if button_color is None else button_color
        self.button_hover = ctk.ThemeManager.theme["CTkButton"]["hover_color"] if button_hover_color is None else button_hover_color
        self.entry_color = ctk.ThemeManager.theme["CTkEntry"]["fg_color"] if entry_color is None else entry_color
        self.border_color = ctk.ThemeManager.theme["CTkEntry"]["border_color"] if border_color is None else border_color
        self.text_color = ctk.ThemeManager.theme["CTkEntry"]["text_color"] if text_color is None else text_color
        super().configure(border_color=self.border_color)
        
        self.subtract_button = ctk.CTkButton(self, text="-", width=height-6, height=height-6, border_color=self.border_color, text_color=self.text_color,
                                                       command=self.subtract_button_callback, fg_color=self.button_color, hover_color=self.button_hover)
        self.subtract_button.grid(row=0, column=0, padx=(3, 0), pady=3)
    
        self.text = tk.DoubleVar()
        self.text.set(self.min_value)

        self.entry = ctk.CTkEntry(self, width=width-(2*height), height=height-6, textvariable=self.text, fg_color=self.entry_color, border_width=entry_border,
                                            placeholder_text=str(self.min_value), justify="right", validate='key', border_color=self.border_color,
                                            validatecommand=(self.validation, '%P'), text_color=self.text_color)       
        self.entry.grid(row=0, column=1, columnspan=1, padx=3, pady=3, sticky="ew")

        self.add_button = ctk.CTkButton(self, text="+", width=height-6, height=height-6, border_color=self.border_color, text_color=self.text_color,
                                                  command=self.add_button_callback, fg_color=self.button_color, hover_color=self.button_hover)
        self.add_button.grid(row=0, column=2, padx=(0, 3), pady=3)
        self.entry.bind("<MouseWheel>", self.on_mouse_wheel)
        
    def on_mouse_wheel(self, event):
        if event.delta > 0:
            self.add_button_callback()
        else:
            self.subtract_button_callback()
            
    def add_button_callback(self):
        if self.entry.get() == "":
            self.text.set(0.0)
        try:
            current_value = float(self.entry.get())
            if current_value < self.max_value:
                self.text.set(round(current_value + self.step_size, 2))
        except ValueError:
            return
        if self.command is not None:
            self.command()
            
    def configure(self, state="enabled", **kwargs):
        if state == "disabled":
            self.entry.configure(state="disabled")
            self.add_button.configure(state="disabled", **kwargs)
            self.subtract_button.configure(state="disabled", **kwargs)
            self.entry.unbind("<MouseWheel>")
        else:
            self.entry.configure(state="normal")
            self.add_button.configure(state="normal", **kwargs)
            self.subtract_button.configure(state="normal", **kwargs)
            self.entry.bind("<MouseWheel>", self.on_mouse_wheel)
            
    def subtract_button_callback(self):
        if self.entry.get() == "":
            self.text.set(0.0)
        try:
            current_value = float(self.entry.get())
            if self.min_value < current_value <= self.max_value:
                self.text.set(round(current_value - self.step_size, 2))
        except ValueError:
            return
        if self.command is not None:
            self.command()
        
    def only_numbers(self, char):
        if char == "" or (char.replace('.', '', 1).isdigit() and char.count('.') < 2):
            try:
                float(char)  # Ensure it's a valid float
                return True
            except ValueError:
                return False
        else:
            return False
        
    def get(self) -> Union[float, None]:
        if self.entry.get() == "":
            return 0.0
        try:
            return round(float(self.text.get()), 2)
        except ValueError:
            return None

    def set(self, value: float):
        try:
            float(value)
            self.text.set(round(float(value), 2))
        except ValueError:
            pass






class CustomToolbar(NavigationToolbar2Tk):
    #Toolbar with only the save button
    def __init__(self, canvas, parent):
        super().__init__(canvas, parent)

        for widget in self.winfo_children():
            try:
                if isinstance(widget, ttk.Separator):
                    widget.pack_forget()
                elif str(widget['text']) != 'Save':
                    widget.pack_forget()
            except (tk.TclError, KeyError):
                pass



if __name__ == "__main__":
    app = main()
    app.mainloop()






