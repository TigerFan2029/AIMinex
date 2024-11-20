import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, IntVar
from tkinter import *
from tkinter import filedialog as fd
from customtkinter import CTkImage
import os

import pandas as pd
import numpy as np

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import MouseButton

from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.pipeline import Pipeline

from sklearn.cluster import AgglomerativeClustering, DBSCAN, SpectralClustering, MeanShift, AffinityPropagation, AgglomerativeClustering, Birch
from sklearn.mixture import GaussianMixture

from PIL import Image
from tktooltip import ToolTip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .custom_spinbox import CTkSpinbox
from . import color_change
from .legend import legend
from .plot_yellowbrick import yellowbrick

class Cluster2DPlotClass:
    def __init__(self, shared_container, cluster, df, cleaned_df, box_frame, box_frame_sub, on_button_click, legend_frame, selected_column):
        # Initialize the 2D cluster plot class with required parameters
        self.df = df.dropna()
        self.shared_container = shared_container
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub

        self.on_button_click = on_button_click
        self.legend_frame = legend_frame
        self.cleaned_df = cleaned_df
        self.var = IntVar()
        self.var1 = IntVar()
        self.cluster_result = cluster
        self.selected_column = selected_column
    
        self.plot_2d_cluster()


    def plot_2d_cluster(self):
        # Load and display the 2D cluster plot icon button
        script_dir = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(script_dir, "images", "images_program", "2d.png")
        pil_image = Image.open(image_path)

        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))

        print(pil_image)
        print(self.icon_image)
        
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text = "", width=20)
        self.image_button.grid(row=3, column=3, sticky="w", pady=0, padx=5)
        
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_2d_cluster_sub0))

        ToolTip(self.image_button, msg="PCA 2D Biplot with Clusters")

    def plot_2d_cluster_sub0(self):
        # clear column cluster if exist
        if 'cluster' in self.df.columns:
            self.df.drop(columns=['cluster'], inplace=True)
        else:
            pass

        # Display options for selecting axes and cluster settings
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()
            
        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 1:").grid(columnspan=2, row=0, column=0, pady=(5,0), padx=5)
        self.axis1_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis1_combo.current(0)
        self.axis1_combo.grid(columnspan=2, row=1, column=0, pady=(3,0), padx=5)
        
        ctk.CTkLabel(self.box_frame_sub, text="Select Axis 2:").grid(columnspan=2, row=2, column=0, pady=(5,0), padx=5)
        self.axis2_combo = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), state="readonly")
        self.axis2_combo.current(1)
        self.axis2_combo.grid(columnspan=2, row=3, column=0, pady=(3,0), padx=5)

        ctk.CTkLabel(self.box_frame_sub, text="Element for Data Point Size").grid(columnspan=2, row=4, column=0, pady=(5,0), padx=5)
        self.size_combo1 = ttk.Combobox(self.box_frame_sub, values=self.df.columns.tolist(), name="size_combo1", state="readonly")
        self.size_combo1.set("N/A")
        self.size_combo1.grid(columnspan=2, row=5, column=0, pady=(3,5), padx=5)

        axis1 = self.axis1_combo.get()
        axis2 = self.axis2_combo.get()

        self.df[axis1] = self.df[axis1].apply(lambda x: x.replace('<', '') if isinstance(x, str) and '<' in x else x).astype(float)
        self.df[axis2] = self.df[axis2].apply(lambda x: x.replace('<', '') if isinstance(x, str) and '<' in x else x).astype(float)
        
        self.X = self.df[[axis1, axis2]].values

        if not color_change.ds.empty:
            self.checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Display shape", variable=self.var)
            self.checkbox.grid(columnspan=2, row=6, column=0, sticky=W, pady=(5,0), padx=5)

        self.legend_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Show Legend", variable=self.var1)
        self.legend_checkbox.grid(columnspan=2, row=7, column=0, sticky=W, pady=(5,0), padx=5)
            
        self.cluster_text = ctk.CTkLabel(self.box_frame_sub, text="Number of Clusters:")
        self.k_slider = tk.Scale(self.box_frame_sub, from_=2, to=10, orient=tk.HORIZONTAL, command = self.pipenplot)

        self.apply_plot_2d_cluster = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.show_cluster)
        self.apply_plot_2d_cluster.grid(columnspan=2, row=13, column=0, pady=(10,0), padx=5)

        self.save_button = ctk.CTkButton(self.box_frame_sub, text="Save Clusters to Excel", command=self.save_clusters_to_excel)
        self.save_button.pack(side="bottom", padx=5, pady=5)

        # Define the values for the Spinbox
        linkage_box_values = ('ward', 'complete', 'average', 'single')
        self.linkage_box = ttk.Combobox(self.box_frame_sub, values=linkage_box_values, state="readonly")
        self.linkage_box.current(0)

        affinity_box_values = ('nearest_neighbors', 'rbf', 'precomputed', 'precomputed_nearest_neighbors')
        self.affinity_box = ttk.Combobox(self.box_frame_sub, values=affinity_box_values, state="readonly")
        self.affinity_box.current(0)
                
        self.min_sample_box = CTkSpinbox(self.box_frame_sub, step_size=1, min_value=2, max_value=100, width = 110) #, command=self.runeps)
        self.eps_box = CTkSpinbox(self.box_frame_sub, step_size=0.1, min_value=0.1, max_value=100, width = 110)
        
        self.eps_text = ctk.CTkLabel(self.box_frame_sub, text="eps")
        self.min_sample_text = ctk.CTkLabel(self.box_frame_sub, text="min_sample")

        self.plot_2d_cluster_sub(self.cluster_result)
        
    def plot_2d_cluster_sub(self, cluster):
        self.box_frame_sub.grid_propagate(True)
        self.cluster_result = cluster

        self.linkage_box.grid_forget()
        self.affinity_box.grid_forget()
        self.min_sample_box.grid_forget()
        self.eps_box.grid_forget()
        self.eps_text.grid_forget()
        self.min_sample_text.grid_forget()
        
        try:
            self.param_text.grid_forget()
        except:
            pass

        
        if self.cluster_result == "Hierarchical":
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Linkage Type:")
            self.linkage_box.grid(columnspan=2, row=9, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=8, column=0, pady=(5,0), padx=5)
            
        elif self.cluster_result == "Spectral":
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Affinity Type:")
            self.affinity_box.grid(columnspan=2, row=9, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=8, column=0, pady=(5,0), padx=5)

        elif self.cluster_result == "DBSCAN":
            # self.runeps() 
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="eps & min_sample value:")
            
            self.eps_text.grid(row=10, column=0, pady=0, padx=(5,0), sticky="w")
            self.min_sample_text.grid(row=9, column=0, pady=5, padx=(5,0), sticky="w")
        
            self.eps_box.grid(row=10, column=1, pady=0, padx=(0,5))
            self.min_sample_box.grid(row=9, column=1, pady=10, padx=(0,5))
            self.param_text.grid(columnspan=2,row=8, column=0, pady=(5,0), padx=5)
            
        else:
            pass
        
        if self.cluster_result in ["DBSCAN", "Mean Shift", "Affinity Propagation"]:
            self.cluster_text.grid_forget()
            self.k_slider.grid_forget()
        else:
            self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
            self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
        
        self.pipenplot()

    def pipenplot(self, *arg):
        k = self.k_slider.get()
        linkage = self.linkage_box.get()
        affinity = self.affinity_box.get()

        # Set up the clustering pipeline based on selected method
        if self.cluster_result == "K-mean":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
            ])
            
        elif self.cluster_result == "DBSCAN":

            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", DBSCAN(eps=self.eps_box.get(), min_samples= int(self.min_sample_box.get())))
            ])
            
        elif self.cluster_result == "Mean Shift":  
            from sklearn.cluster import estimate_bandwidth
            bandwidth_value = estimate_bandwidth(self.X, quantile=0.2)
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", MeanShift(bandwidth=bandwidth_value))
            ])
            
        elif self.cluster_result == "Spectral":  
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", SpectralClustering(n_clusters=k, affinity=affinity, random_state=0))
            ])
            
        elif self.cluster_result == "GMM":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", GaussianMixture(n_components=k, random_state=0))
            ])
            
        elif self.cluster_result == "Affinity Propagation":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AffinityPropagation(random_state=0))
            ])
            
        elif self.cluster_result == "Hierarchical":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AgglomerativeClustering(n_clusters=k, linkage=linkage))
            ])

        elif self.cluster_result == "BIRCH":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", Birch(n_clusters=k))
            ])

        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        yellowbrick(self, self.X)


    def show_cluster(self):         
        # Check if there is an active tab in the shared container
        if not self.shared_container.current_tab:
            self.shared_container.create_tab() #("2D Cluster")

        # Clear the content frame and legend frame
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()

        axis1 = self.axis1_combo.get()
        axis2 = self.axis2_combo.get()
        linkage = self.linkage_box.get()
        affinity = self.affinity_box.get()

        # Ensure the selected axes are different
        if axis1 != axis2:
            k = self.k_slider.get()

            # Set up the clustering pipeline based on selected method
            if self.cluster_result == "K-mean":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "DBSCAN":

                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", DBSCAN(eps=self.eps_box.get(), min_samples= int(self.min_sample_box.get())))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif self.cluster_result == "Mean Shift":  
                from sklearn.cluster import estimate_bandwidth
                bandwidth_value = estimate_bandwidth(self.X, quantile=0.2)
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", MeanShift(bandwidth=bandwidth_value))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif self.cluster_result == "Spectral":  
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", SpectralClustering(n_clusters=k, affinity=affinity, random_state=0))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "GMM":   
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", GaussianMixture(n_components=k, random_state=0))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "Affinity Propagation":   
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", AffinityPropagation(random_state=0))
                ])
                self.cluster_text.grid_forget()
                self.k_slider.grid_forget()
                
            elif self.cluster_result == "Hierarchical":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", AgglomerativeClustering(n_clusters=k, linkage=linkage))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)
                
            elif self.cluster_result == "BIRCH":
                self.pipe = Pipeline([
                    ("scale", StandardScaler()),
                    ("model", Birch(n_clusters=k))
                ])
                self.cluster_text.grid(columnspan=2, row=11, column=0, pady=(5,0), padx=5)
                self.k_slider.grid(columnspan=2, row=12, column=0, pady=(3,0), padx=5)

 
            self.pipe.fit(self.X)
            
            if self.cluster_result == "GMM":  
                self.df['cluster'] = self.pipe.named_steps["model"].predict(self.X)
            else:
                self.df['cluster'] = self.pipe.named_steps['model'].labels_
            
            # Create a meshgrid for the plot
            try:
                fig, ax = plt.subplots(figsize=(9, 7))
                
                rangex = (self.X[:, 0].max()-self.X[:, 0].min())/10
                rangey = (self.X[:, 1].max()-self.X[:, 1].min())/10
                x_min, x_max = self.X[:, 0].min() - rangex, self.X[:, 0].max() + rangex
                y_min, y_max = self.X[:, 1].min() - rangey, self.X[:, 1].max() + rangey
                
                stepsize_x = rangex/75
                stepsize_y = rangey/75

                xx, yy = np.meshgrid(np.arange(x_min, x_max, stepsize_x), np.arange(y_min, y_max, stepsize_y))
                
                if self.cluster_result == "Spectral" or self.cluster_result == "DBSCAN" or self.cluster_result == "Hierarchical":
                    self.pipe.named_steps['scale'].fit(self.X)
                    X_scaled = self.pipe.named_steps['scale'].transform(self.X)
                    self.pipe.named_steps['model'].fit(X_scaled)
            
                    def predict_new_data(new_data, pipe, train_data, train_labels):
                        new_data_scaled = pipe.named_steps['scale'].transform(new_data)
                        distances = np.linalg.norm(new_data_scaled[:, np.newaxis] - train_data[np.newaxis, :], axis=2)
                        nearest_indices = np.argmin(distances, axis=1)
                        return train_labels[nearest_indices]
            
                    labels = self.pipe.named_steps['model'].labels_
    
                    new_data = np.c_[xx.ravel(), yy.ravel()]
                    new_data_labels = predict_new_data(new_data, self.pipe, X_scaled, labels)
            
                    Z = new_data_labels.reshape(xx.shape)
    
                else:
                    Z = self.pipe.named_steps['model'].predict(self.pipe.named_steps['scale'].transform(np.c_[xx.ravel(), yy.ravel()]))
                    Z = Z.reshape(xx.shape)
    
                ax.contourf(xx, yy, Z, cmap=plt.cm.Pastel1, alpha=0.2)
                
                def map_size(x, new_min, new_max):
                    old_min = self.min_ele
                    old_max = self.max_ele
                    return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
            except:
                fig, ax = plt.subplots(figsize=(9, 7))

            self.shapes = []
            try:
                colors = color_change.dc['Color']
            except:
                pass

            # Plot the points with shapes and colors
            element_size = self.size_combo1.get()
            from .color_change import column_to_use
            if element_size == "N/A":
                if column_to_use is not None:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i], marker=color_change.ds["Shapes"][i]
                        )
                        self.shapes.append(shape)

                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors)

                else:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors)                        

            else:
                # Try to get element_size column values for size mapping
                self.max_ele = self.df[element_size].max()
                self.min_ele = self.df[element_size].min()
                
                # Function to map element_size values to point sizes
                def map_size(x, new_min, new_max):
                    old_min = self.min_ele
                    old_max = self.max_ele
                    return ((np.log(x + 1) - np.log(old_min + 1)) / (np.log(old_max + 1) - np.log(old_min + 1))) * (new_max - new_min) + new_min
                    
                # New logic for zn
                def size_for_zn(x):
                    if x < 1000:
                        return 20
                    elif 1000 <= x <= 9999:
                        return 70
                    elif x > 10000:
                        return 110
                    return 20  # Default size if none of the conditions match
                
                # Apply size mapping logic
                if element_size == 'Zn':
                    sizes = self.df[element_size].apply(size_for_zn)
                else:
                    sizes = self.df[element_size].apply(lambda x: map_size(x, 20, 100))

                if column_to_use is not None:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i], marker=color_change.ds["Shapes"][i], s=sizes[i]
                        )
                        self.shapes.append(shape)   
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors, s=sizes)

                else:
                    for i in range(len(self.df)):
                        shape = ax.scatter(self.df[axis1][i], self.df[axis2][i],
                            color=colors[i], s=sizes[i]
                        )
                        self.shapes.append(shape)
                    self.dots = ax.scatter(self.df[axis1], self.df[axis2], c=colors, s=sizes)   

            if "sample id" in self.cleaned_df.columns:
                description = self.cleaned_df["sample id"]
                self.annotations = {}
                self.selected_points = set()
                
                # Functions to handle annotations on the plot
                def add_annotation(ind, source):
                    if source == self.dots:
                        index = ind["ind"][0]
                        pos = source.get_offsets()[index]
                        text = f"{description.iloc[index]}"
                    else:
                        index = ind
                        pos = (self.df.iloc[index][axis1], self.df.iloc[index][axis2])
                        text = f"{description.iloc[index]}"
                
                    offsets = [(20, 20), (-20, 20), (20, -20), (-20, -20), (40, 40), (-40, 40), (40, -40), (-40, -40)]
                    
                    for offset in offsets:
                        annot = ax.annotate(text, xy=pos, xytext=offset,
                                            textcoords="offset points",
                                            bbox=dict(boxstyle="round", fc="w"),
                                            arrowprops=dict(arrowstyle="->"),
                                            zorder=100)
                        fig.canvas.draw()
                        bbox = annot.get_window_extent()
                        overlap = False
                        for existing_annot in self.annotations.values():
                            existing_bbox = existing_annot.get_window_extent()
                            if bbox.overlaps(existing_bbox):
                                overlap = True
                                annot.remove()
                                break
                        if not overlap:
                            annot.set_visible(True)
                            self.annotations[index] = annot
                            self.selected_points.add(index)
                            fig.canvas.draw_idle()
                            return 
                
                    annot = ax.annotate(text, xy=pos, xytext=offsets[0],
                                        textcoords="offset points",
                                        bbox=dict(boxstyle="round", fc="w"),
                                        arrowprops=dict(arrowstyle="->"),
                                        zorder=100)
                    annot.set_visible(True)
                    self.annotations[index] = annot
                    self.selected_points.add(index)
                    fig.canvas.draw_idle()
                
                def remove_annotation(index):
                    annot = self.annotations.pop(index, None)
                    if annot:
                        annot.remove()
                    self.selected_points.discard(index)
                    fig.canvas.draw_idle()
                
                def update_annotations(event):
                    for index, annot in self.annotations.items():
                        if index in self.selected_points:
                            annot.xy = (self.df.iloc[index][axis1], self.df.iloc[index][axis2])
                    fig.canvas.draw_idle()
                
                def on_click(event):
                    if event.inaxes == ax and event.button == MouseButton.LEFT:
                        contains_dots, ind_dots = self.dots.contains(event)
                        contains_shapes = False
                        shape_index = None
                        for i, shape in enumerate(self.shapes):
                            contains_shapes, ind_shapes = shape.contains(event)
                            if contains_shapes:
                                shape_index = i
                                break
                        
                        if contains_dots:
                            index = ind_dots["ind"][0]
                            if index in self.selected_points:
                                remove_annotation(index)
                            else:
                                add_annotation(ind_dots, self.dots)
                        elif contains_shapes:
                            if shape_index in self.selected_points:
                                remove_annotation(shape_index)
                            else:
                                add_annotation(shape_index, shape)
                
                fig.canvas.mpl_connect("button_press_event", on_click)
                fig.canvas.mpl_connect("motion_notify_event", update_annotations)

            # Plot centroids
            scaler = self.pipe.named_steps['scale']

            if self.cluster_result == "BIRCH":
                centroids = self.pipe.named_steps['model'].subcluster_centers_
            elif self.cluster_result == "GMM":
                centroids = self.pipe.named_steps['model'].means_
            elif self.cluster_result == "Spectral" or self.cluster_result == "DBSCAN" or self.cluster_result == "Hierarchical":
                model = self.pipe.named_steps['model']
                labels = model.labels_
                labels = np.array(labels)
                
                unique_labels = np.unique(labels)
                centroids = []
                
                for i in unique_labels:
                    cluster_points = self.pipe.named_steps['scale'].transform(self.X[labels == i])
                    print(f"Cluster {i} points:", cluster_points)
                    centroid = cluster_points.mean(axis=0)
                    centroids.append(centroid)
                        
                centroids = np.array(centroids)

            else:
                centroids = self.pipe.named_steps['model'].cluster_centers_
                
            centroids_original = scaler.inverse_transform(centroids)
            
            ax.scatter(centroids_original[:, 0], centroids_original[:, 1], 
                       marker="X", c="red", s=30, label="centroids")
            ax.legend()
            ax.set_title(f"2D {self.cluster_result} Clustering")
            ax.set_xlabel(axis1)
            ax.set_ylabel(axis2)
            
            canvas = FigureCanvasTkAgg(fig, master=content_frame)
            canvas.draw()

            toolbar = NavigationToolbar2Tk(canvas, content_frame)
            toolbar.update()
            toolbar.pack(side=tk.TOP, fill=tk.X)
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)

            self.shape()
            Legend = self.var1.get() == 1
            print(Legend)
                        
            for widget in self.legend_frame.winfo_children():
                widget.destroy()

            if Legend:
                pass
            else:
                yellowbrick(self, self.X)
            legend(self)
            
        else:
            self.output_text.insert("end", 'Select two different elements')

    def shape(self):
        # Display or hide shapes in the plot
        shape = self.var.get() == 1
        if shape:
            for shape in self.shapes:
                shape.set_visible(True)
            self.dots.set_visible(False)
        else:
            self.dots.set_visible(True)
            for shape in self.shapes:
                shape.set_visible(False)

    def save_clusters_to_excel(self):
        # Save the clustered data to an Excel file with colored cells based on clusters
        axis1 = self.axis1_combo.get().lower()
        axis2 = self.axis2_combo.get().lower()

        cleaned_df_copy = self.cleaned_df.copy()
        cleaned_df_copy.columns = cleaned_df_copy.columns.str.replace('_ppm', '', regex=False).str.replace('_pct', '', regex=False)
        
        if axis1 == axis2:
            print("Select two different elements to save to Excel.")
            return

        if 'cluster' not in self.df.columns:
            print("Clusters have not been calculated yet.")
            return

        if axis1 not in cleaned_df_copy.columns or axis2 not in cleaned_df_copy.columns or 'sample id' not in cleaned_df_copy.columns:
            print(f"One or more of the columns '{axis1}', '{axis2}', or 'sample id' are missing in the data.")
            return

        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return

        cleaned_df_copy['cluster'] = self.df['cluster']
        
        pivot_data = cleaned_df_copy.pivot_table(index='sample id', values=[axis1, axis2, 'cluster'], sort=False)
        
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            pivot_data.to_excel(excel_writer, index=True)
        
        wb = load_workbook(file_name)
        ws = wb.active

        colormap = plt.get_cmap('Set2')
        cluster_colors = ['FF{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)) for r, g, b, _ in [colormap(i) for i in range(colormap.N)]]
        cluster_fill = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in cluster_colors]

        for i in range(len(cleaned_df_copy)):
            sample_id = cleaned_df_copy.iloc[i]['sample id']
            cluster = cleaned_df_copy.iloc[i]['cluster']
            row = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
            for cell in row:
                if cell[0].value == sample_id:
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=cell[0].row, column=col).fill = cluster_fill[cluster]

        wb.save(file_name)
        print(f"Clusters saved to {file_name}")
        del cleaned_df_copy