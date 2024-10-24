import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, IntVar
from tkinter import *
from tkinter import filedialog as fd
from customtkinter import CTkImage

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.pipeline import Pipeline

from sklearn.cluster import AgglomerativeClustering, DBSCAN, SpectralClustering, MeanShift, AffinityPropagation, AgglomerativeClustering, Birch
from sklearn.mixture import GaussianMixture 

from PIL import Image
from tktooltip import ToolTip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .custom_spinbox import CTkSpinbox
from .plot_yellowbrick import yellowbrick


class loading_cluster:
    def __init__(self, shared_container, cluster, loadings, box_frame, box_frame_sub, on_button_click, apply_button, legend_frame):
        # Initialize the loading cluster class with necessary parameters
        self.loadings = loadings
        self.shared_container = shared_container
        self.cluster_result = cluster
        self.box_frame = box_frame
        self.box_frame_sub = box_frame_sub
        self.legend_frame = legend_frame
        self.on_button_click = on_button_click
        self.apply_button = apply_button
        self.cluster_sort_var = IntVar()
        self.graph_data_df = None

        self.plot_cluster()

    def plot_cluster(self):
        # Create button for displaying cluster bar graph
        pil_image = Image.open("src/minexai/images/images_program/ele.png")
        self.icon_image = CTkImage(light_image=pil_image, dark_image=pil_image, size=(32, 32))
        
        # Create and place image button
        self.image_button = ctk.CTkLabel(self.box_frame, image=self.icon_image, text="", width=20)
        self.image_button.grid(row=3, column=0, sticky="w", pady=0, padx=5)

        # Bind button click event to cluster function
        self.image_button.bind("<Button-1>", lambda event: self.on_button_click(self.image_button, self.plot_cluster_sub0))

        ToolTip(self.image_button, msg="PC Cluster Bar Graph by Elements")

    def plot_cluster_sub0(self):
        # Create elements and clusters selection UI
        for widget in self.box_frame_sub.winfo_children():
            widget.destroy()

        # Create and pack checkbox for sorting clusters
        self.cluster_sort_checkbox = ctk.CTkCheckBox(self.box_frame_sub, text="Sort Clusters Together", variable=self.cluster_sort_var)
        self.cluster_sort_checkbox.custom_name = "loading_check"
        self.cluster_sort_checkbox.grid(columnspan=2, row=0, column=0, pady=(5,0), padx=5)
            
        # Create and pack slider for number of clusters
        self.cluster_text = ctk.CTkLabel(self.box_frame_sub, text="Number of Clusters:")
        self.k_slider = tk.Scale(self.box_frame_sub, from_=2, to=10, orient=tk.HORIZONTAL, command = self.pipenplot)
            
        # Create and pack apply button
        self.apply_plot_3d_cluster = ctk.CTkButton(self.box_frame_sub, text="Apply", command=self.update_plots)
        self.apply_plot_3d_cluster.grid(columnspan=2, row=6, column=0, pady=10, padx=5)

        # Create and pack save button
        self.save_button = ctk.CTkButton(self.box_frame_sub, text="Save Clusters to Excel", command=self.save_loadings_to_excel)
        self.save_button.pack(side="bottom", padx=5, pady=5)

        # Define the values for the Spinbox
        linkage_box_values = ('ward', 'complete', 'average', 'single')
        self.linkage_box = ttk.Combobox(self.box_frame_sub, values=linkage_box_values, state="readonly")
        self.linkage_box.current(0)

        affinity_box_values = ('nearest_neighbors', 'rbf', 'precomputed', 'precomputed_nearest_neighbors')
        self.affinity_box = ttk.Combobox(self.box_frame_sub, values=affinity_box_values, state="readonly")
        self.affinity_box.current(0)
                
        self.min_sample_box = CTkSpinbox(self.box_frame_sub, step_size=1, min_value=2, max_value=100, width = 110) #, command=self.runeps)
        self.eps_box = CTkSpinbox(self.box_frame_sub, step_size=0.1, min_value=0.1, max_value=100, width = 110)
                    
        self.eps_text = ctk.CTkLabel(self.box_frame_sub, text="eps")
        self.min_sample_text = ctk.CTkLabel(self.box_frame_sub, text="min_sample")

        self.plot_cluster_sub(self.cluster_result)
    
        
    def plot_cluster_sub(self, cluster):
        self.box_frame_sub.grid_propagate(False)
        self.cluster_result = cluster
        
        self.linkage_box.grid_forget()
        self.affinity_box.grid_forget()
        self.min_sample_box.grid_forget()
        self.eps_box.grid_forget()
        self.eps_text.grid_forget()
        self.min_sample_text.grid_forget()
        
        try:
            self.param_text.grid_forget()
        except:
            pass

        if self.cluster_result == "Hierarchical":
            print("Hierarchical Clustering Selected")  # Debug print
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Linkage Type:")
            self.linkage_box.grid(columnspan=2, row=2, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
            
        elif self.cluster_result == "Spectral":
            print("Spectral Clustering Selected")  # Debug print
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="Affinity Type:")
            self.affinity_box.grid(columnspan=2, row=2, column=0, pady=(3,0), padx=5)
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
    
        elif self.cluster_result == "DBSCAN":
            print("DBSCAN Clustering Selected")  # Debug print
            # self.runeps() 
            self.param_text = ctk.CTkLabel(self.box_frame_sub, text="eps & min_sample value:")

            self.eps_text.grid(row=3, column=0, pady=0, padx=(5,0), sticky="w")
            self.min_sample_text.grid(row=2, column=0, pady=5, padx=(5,0), sticky="w")
        
            self.eps_box.grid(row=3, column=1, pady=0, padx=(0,5), sticky="e")
            self.min_sample_box.grid(row=2, column=1, pady=10, padx=(0,5), sticky="e")
            self.param_text.grid(columnspan=2, row=1, column=0, pady=(5,0), padx=5)
            
        else:
            pass
            
        if self.cluster_result in ["DBSCAN", "Mean Shift", "Affinity Propagation"]:
            self.cluster_text.grid_forget()
            self.k_slider.grid_forget()
        else:
            self.cluster_text.grid(columnspan=2,row=4, column=0, pady=0, padx=5)
            self.k_slider.grid(columnspan=2,row=5, column=0, pady=(3,5), padx=5)

        self.pipenplot()


    def pipenplot(self, *arg):
        for widget in self.legend_frame.winfo_children():
            widget.destroy()
        k = self.k_slider.get()
        linkage = self.linkage_box.get()
        affinity = self.affinity_box.get()     

        if self.cluster_result == "K-mean":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
            ])
            
        elif self.cluster_result == "DBSCAN":
            print(self.eps_box.get())
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", DBSCAN(eps=self.eps_box.get(), min_samples= int(self.min_sample_box.get())))
            ])
            
        elif self.cluster_result == "Mean Shift":  
            from sklearn.cluster import estimate_bandwidth
            bandwidth_value = estimate_bandwidth(self.X, quantile=0.2)
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", MeanShift(bandwidth=bandwidth_value))
            ])
                
        elif self.cluster_result == "Spectral":  
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", SpectralClustering(n_clusters=k, affinity=affinity, random_state=0))
            ])
            
        elif self.cluster_result == "GMM":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", GaussianMixture(n_components=k, random_state=0))
            ])
            
        elif self.cluster_result == "Affinity Propagation":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AffinityPropagation(random_state=0))
            ])
            
        elif self.cluster_result == "Hierarchical":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AgglomerativeClustering(n_clusters=k, linkage=linkage))
            ])
            
        elif self.cluster_result == "BIRCH":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", Birch(n_clusters=k))
            ])

        for widget in self.legend_frame.winfo_children():
            widget.destroy()

        yellowbrick(self, self.loadings)
        

    def normalize(self, values, vmin=None, vmax=None):
        # Normalize values for color scaling
        vmin = vmin if vmin is not None else np.min(values)
        vmax = vmax if vmin is not None else np.max(values)
        norm_values = (values - vmin) / (vmax - vmin)
        return norm_values

    def plot_pca_barchart(self, data, entry_name, ax, cluster_colors, sorted_labels):
        # Plot the PCA bar chart with clusters
        cluster_palette = plt.get_cmap('tab20', len(np.unique(cluster_colors)))
        colors = cluster_palette(cluster_colors)
        data.plot(kind='barh', color=colors, ax=ax, fontsize="xx-small")
        ax.set_title(entry_name)
        ax.grid(True, linestyle='--', linewidth=0.5)
        ax.set_yticks(range(len(data)))
        ax.set_yticklabels(sorted_labels, fontsize="xx-small") 
        ax.set_ylabel('')

        # Update y-axis labels
        labels = ax.get_yticklabels()
        new_labels = [label.get_text().replace('_ppm', '')
                                      .replace('_pct', '')
                                      .replace('_Howell', '') for label in labels]
        ax.set_yticklabels(new_labels)

    def update_plots(self):
        # Update plots with selected clustering method and options
        if not self.shared_container.current_tab:
            self.shared_container.create_tab() #("Cluster Bar Graph by Elements")

        # Close all existing plots and clear widgets
        plt.close('all')
        content_frame = self.shared_container.current_tab[1]
        for widget in content_frame.winfo_children():
            widget.destroy()

        k = self.k_slider.get()

        # Get the selected clustering method
        cluster_result = self.cluster_result
        
        # Set up the pipeline for different clustering algorithms
        if cluster_result == "K-mean":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", KMeans(n_clusters=k, random_state=0, n_init='auto'))
            ])
            
        elif cluster_result == "DBSCAN":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", DBSCAN(eps=self.eps_box.get(), min_samples=self.min_sample_box.get()))
            ])

        elif cluster_result == "Mean Shift":  
            from sklearn.cluster import estimate_bandwidth
            bandwidth_value = estimate_bandwidth(self.loadings, quantile=0.2)
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", MeanShift(bandwidth=bandwidth_value))
            ])
            
        elif cluster_result == "Spectral":  
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", SpectralClustering(n_clusters=k, affinity='nearest_neighbors', random_state=0))
            ])

        elif cluster_result == "GMM":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", GaussianMixture(n_components=k, random_state=0))
            ])

        elif cluster_result == "Affinity Propagation":   
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AffinityPropagation(random_state=0))
            ])

        elif cluster_result == "Hierarchical":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", AgglomerativeClustering(n_clusters=k, linkage='ward'))
            ])

        elif cluster_result == "BIRCH":
            self.pipe = Pipeline([
                ("scale", StandardScaler()),
                ("model", Birch(n_clusters=k))
            ])
            
        # Fit the pipeline to the data
        self.pipe.fit(self.loadings)
        self.clustering_results = pd.DataFrame(index=self.loadings.index)
        
        # Get clustering results
        if cluster_result == "GMM":  
            self.clustering_results['cluster'] = self.pipe.named_steps["model"].predict(self.loadings)
        else:
            self.clustering_results['cluster'] = self.pipe.named_steps['model'].labels_

        fig, axes = plt.subplots(ncols=len(self.loadings.columns), figsize=(15, 8))
        
        cluster_colors = self.clustering_results['cluster'].values
        self.clustering_results['element'] = self.loadings.index
        
        graph_data = []
        sort_clusters = self.cluster_sort_var.get() == 1
        
        for i in range(len(self.loadings.columns)):
            pc_name = f'PC{i+1}'
            data = self.loadings[pc_name]

            if sort_clusters:
                # Sort data by clusters
                temp_df = pd.DataFrame({
                    'data': data,
                    'cluster': cluster_colors,
                    'element':  self.clustering_results['element'],
                    'PC': pc_name
                })

                sorted_df = temp_df.sort_values(by='cluster')
                sorted_data = sorted_df['data']
                sorted_cluster_colors = sorted_df['cluster'].values
                sorted_labels = sorted_df['element'].values
                self.plot_pca_barchart(sorted_data, pc_name, axes[i], sorted_cluster_colors, sorted_labels)
                graph_data.append(sorted_df)
                
            else:
                temp_df = pd.DataFrame({
                    'data': data,
                    'cluster': cluster_colors,
                    'element':  self.clustering_results['element'],
                    'PC': pc_name
                })
                graph_data.append(temp_df)
                labels = temp_df['element'].values
                self.plot_pca_barchart(data, pc_name, axes[i], cluster_colors, labels)

        self.graph_data_df = pd.concat(graph_data)

        # Add a big title over all subplots
        fig.suptitle(f'{cluster_result} PC Cluster Charts by Elements', fontsize=16)

        plt.tight_layout()
        
        # Add the plot to the tkinter canvas
        canvas = FigureCanvasTkAgg(fig, master=content_frame)
        canvas.draw()

        # Add navigation toolbar to the canvas
        toolbar = NavigationToolbar2Tk(canvas, content_frame)
        toolbar.update()
        toolbar.pack(side=tk.TOP, fill=tk.X)
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True, in_=content_frame)


    def save_loadings_to_excel(self):
        # Save the loadings to an Excel file
        if self.graph_data_df is None:
            print("No data to save. Ensure clusters are calculated and graph is generated.")
            return
        
        # Prompt user to select save location and filename
        file_name = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_name:
            return

        # Create a pivot table with the clustering results
        pivot_data = self.graph_data_df.pivot_table(index='element', columns='PC', values='data', sort=False)
        pivot_data.columns = [f'{col}' for col in pivot_data.columns]

        # Save the pivot table to an Excel file
        with pd.ExcelWriter(file_name, engine='openpyxl') as excel_writer:
            pivot_data.to_excel(excel_writer, index=True)

        # Load the saved workbook and worksheet
        wb = load_workbook(file_name)
        ws = wb.active

        # Define color map for clusters
        colormap = plt.get_cmap('Set2')
        cluster_colors = ['FF{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255)) for r, g, b, _ in [colormap(i) for i in range(colormap.N)]]
        cluster_fill = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in cluster_colors]

        # Apply cluster colors to the Excel file
        for i in range(len(self.graph_data_df)):
            element = self.graph_data_df.iloc[i]['element']
            cluster = self.graph_data_df.iloc[i]['cluster']
            row = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
            for cell in row:
                if cell[0].value == element:
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row=cell[0].row, column=col).fill = cluster_fill[cluster]

        wb.save(file_name)
        print(f"Loadings saved to {file_name}")


